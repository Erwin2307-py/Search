# modules/email_module.py - VOLLSTÄNDIGE VERSION MIT MEHREREN EMAIL-EMPFÄNGERN UND EXCEL-INTEGRATION
import streamlit as st
import datetime
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import time
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import ssl
from typing import List, Dict, Any, Tuple
import json
from pathlib import Path
import threading

def module_email():
    """VOLLSTÄNDIGE FUNKTION - Email-Modul mit mehreren Empfängern und Excel-Integration"""
    st.title("📧 Wissenschaftliches Paper-Suche & Email-System")
    st.success("✅ Vollständiges Modul mit mehreren Email-Empfängern und Excel-Integration geladen!")
    
    # Session State initialisieren
    initialize_session_state()
    
    # Erweiterte Tabs
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "📊 Dashboard", 
        "🔍 Paper-Suche", 
        "📧 Email-Konfiguration",
        "📋 Excel-Management",
        "🤖 Automatische Suchen",
        "📈 Statistiken",
        "⚙️ System-Einstellungen"
    ])
    
    with tab1:
        show_dashboard()
    
    with tab2:
        show_advanced_paper_search()
    
    with tab3:
        show_email_config()
    
    with tab4:
        show_excel_template_management()
    
    with tab5:
        show_automatic_search_system()
    
    with tab6:
        show_detailed_statistics()
    
    with tab7:
        show_system_settings()

def initialize_session_state():
    """Vollständige Session State Initialisierung mit mehreren Email-Empfängern"""
    # Erstelle notwendige Ordner
    for folder in ["excel_templates", "saved_searches", "search_history", "config"]:
        if not os.path.exists(folder):
            os.makedirs(folder)
    
    # Email-Einstellungen - ERWEITERT für mehrere Empfänger
    if "email_settings" not in st.session_state:
        st.session_state["email_settings"] = {
            "sender_email": "",
            "recipient_emails": "",  # MEHRERE EMPFÄNGER (komma-getrennt)
            "smtp_server": "smtp.gmail.com",
            "smtp_port": 587,
            "sender_password": "",
            "use_tls": True,
            "auto_notifications": True,
            "min_papers": 1,
            "notification_frequency": "Bei jeder Suche",
            "subject_template": "🔬 {count} neue Papers für '{search_term}' - {frequency}",
            "message_template": """📧 Automatische Paper-Benachrichtigung

📅 Datum: {date}
🔍 Suchbegriff: '{search_term}'
📊 Neue Papers: {count}
⏰ Häufigkeit: {frequency}

📋 Neue Papers:
{new_papers_list}

📎 Excel-Datei wurde aktualisiert: {excel_file}

Mit freundlichen Grüßen,
Ihr automatisches Paper-Überwachung-System"""
        }
    
    # Excel-Template System
    if "excel_template" not in st.session_state:
        st.session_state["excel_template"] = {
            "file_path": "excel_templates/master_papers.xlsx",
            "auto_create_sheets": True,
            "sheet_naming": "topic_based",
            "max_sheets": 50
        }
    
    # Such-Historie
    if "search_history" not in st.session_state:
        st.session_state["search_history"] = []
    
    # Email-Historie
    if "email_history" not in st.session_state:
        st.session_state["email_history"] = []
    
    # Automatische Suchen
    if "automatic_searches" not in st.session_state:
        st.session_state["automatic_searches"] = {}
    
    # System-Status
    if "system_status" not in st.session_state:
        st.session_state["system_status"] = {
            "total_searches": 0,
            "total_papers": 0,
            "total_emails": 0,
            "last_search": None,
            "excel_sheets": 0,
            "unique_papers": 0
        }
    
    # Store current search results for manual email sending
    if "current_search_results" not in st.session_state:
        st.session_state["current_search_results"] = {}
    
    # Erstelle Master Excel-Datei falls nicht vorhanden
    create_master_excel_template()

def create_master_excel_template():
    """Erstellt Master Excel-Template mit Overview-Sheet und Excel-Integration"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    if not os.path.exists(template_path):
        try:
            wb = openpyxl.Workbook()
            
            # Overview Sheet
            overview_sheet = wb.active
            overview_sheet.title = "📊_Overview"
            
            # Header-Style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Overview Headers
            overview_headers = [
                "Sheet_Name", "Suchbegriff", "Anzahl_Papers", "Letztes_Update", 
                "Neue_Papers_Letzter_Run", "Status", "Erstellt_am"
            ]
            
            for col, header in enumerate(overview_headers, 1):
                cell = overview_sheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Spaltenbreite anpassen
            column_widths = [20, 25, 15, 18, 20, 12, 18]
            for col, width in enumerate(column_widths, 1):
                col_letter = get_column_letter(col)
                overview_sheet.column_dimensions[col_letter].width = width
            
            # Template Info Sheet
            info_sheet = wb.create_sheet("ℹ️_Template_Info")
            
            info_data = [
                ["📋 Excel Template Information", ""],
                ["", ""],
                ["Erstellt am:", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")],
                ["System:", "Wissenschaftliches Paper-Suche System"],
                ["Version:", "3.0 mit Excel-Integration & mehreren Email-Empfängern"],
                ["", ""],
                ["📖 Anleitung:", ""],
                ["• Jeder Suchbegriff bekommt ein eigenes Sheet", ""],
                ["• Das Overview-Sheet zeigt alle Suchanfragen", ""],
                ["• Neue Papers werden automatisch hinzugefügt", ""],
                ["• Email-Benachrichtigungen an mehrere Empfänger", ""],
                ["• Duplikate werden automatisch erkannt", ""],
            ]
            
            for row_idx, (key, value) in enumerate(info_data, 1):
                info_sheet.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                info_sheet.cell(row=row_idx, column=2, value=value)
            
            info_sheet.column_dimensions['A'].width = 30
            info_sheet.column_dimensions['B'].width = 40
            
            wb.save(template_path)
            st.session_state["system_status"]["excel_sheets"] = len(wb.sheetnames)
            
        except Exception as e:
            st.error(f"❌ Fehler beim Erstellen des Master-Templates: {str(e)}")

# =============== EXCEL-INTEGRATION FUNKTIONEN ===============

def load_master_workbook():
    """Lädt das Master Excel Workbook"""
    excel_path = st.session_state["excel_template"]["file_path"]
    try:
        return openpyxl.load_workbook(excel_path)
    except Exception as e:
        st.error(f"❌ Excel-Datei konnte nicht geladen werden: {str(e)}")
        return None

def paper_exists_in_workbook(pmid: str, wb) -> Tuple[bool, str]:
    """
    Prüft, ob ein Paper (anhand PMID) bereits im gesamten Workbook existiert
    Returns: (exists: bool, sheet_name: str)
    """
    if not pmid:
        return False, ""
    
    for sheet in wb.worksheets:
        if sheet.title.startswith(("📊", "ℹ️")):  # Überspringe Overview und Info-Sheets
            continue
        
        # Prüfe alle Zeilen im Sheet (erste Spalte = PMID)
        for row in sheet.iter_rows(min_row=2, max_col=1):  # Ab Zeile 2, nur erste Spalte
            cell_value = row[0].value
            if str(cell_value) == str(pmid):
                return True, sheet.title
    
    return False, ""

def get_or_create_sheet_for_search(search_term: str, wb) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Holt ein existierendes Sheet für den Suchbegriff oder erstellt ein neues
    """
    # Bereinige Suchbegriff für Sheet-Name (max 31 Zeichen, keine Sonderzeichen)
    sheet_name = generate_sheet_name(search_term)
    
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        # Erstelle neues Sheet
        sheet = wb.create_sheet(title=sheet_name)
        
        # Schreibe Header
        headers = ["PMID", "Titel", "Autoren", "Journal", "Jahr", "DOI", "URL", 
                   "Abstract", "Hinzugefügt_am", "Status", "Notizen"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Spaltenbreiten
        column_widths = [10, 50, 40, 30, 8, 20, 25, 80, 15, 10, 20]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width
        
        return sheet

def add_new_papers_to_excel(search_term: str, all_papers: List[Dict]) -> Tuple[int, List[Dict]]:
    """
    Fügt neue Papers zur Excel hinzu und gibt zurück: (anzahl_neue, neue_papers_liste)
    """
    wb = load_master_workbook()
    if not wb:
        return 0, []
    
    # Hole oder erstelle Sheet für diesen Suchbegriff
    sheet = get_or_create_sheet_for_search(search_term, wb)
    
    new_papers = []
    added_count = 0
    
    for paper in all_papers:
        pmid = paper.get("PMID", "")
        if not pmid:
            continue
        
        # Prüfe ob Paper bereits existiert
        exists, existing_sheet = paper_exists_in_workbook(pmid, wb)
        
        if not exists:
            # Neues Paper - füge hinzu
            current_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
            row_data = [
                pmid,
                paper.get("Title", "")[:500],  # Titel kürzen falls zu lang
                paper.get("Authors", "")[:200], # Autoren kürzen
                paper.get("Journal", ""),
                paper.get("Year", ""),
                paper.get("DOI", ""),
                paper.get("URL", ""),
                paper.get("Abstract", "")[:1000] + "..." if len(paper.get("Abstract", "")) > 1000 else paper.get("Abstract", ""),
                current_time,
                "NEU",
                ""
            ]
            
            sheet.append(row_data)
            new_papers.append(paper)
            added_count += 1
    
    # Speichere Excel
    try:
        wb.save(st.session_state["excel_template"]["file_path"])
        
        # Update Overview Sheet
        total_papers = sheet.max_row - 1  # -1 für Header
        update_overview_sheet_integrated(wb, sheet.title, search_term, total_papers, added_count)
        
        # Update System Status
        st.session_state["system_status"]["excel_sheets"] = len([s for s in wb.sheetnames if not s.startswith(("📊", "ℹ️"))])
        st.session_state["system_status"]["unique_papers"] += added_count
        
    except Exception as e:
        st.error(f"❌ Fehler beim Speichern der Excel: {str(e)}")
        return 0, []
    
    return added_count, new_papers

def update_overview_sheet_integrated(wb, sheet_name: str, search_term: str, total_papers: int, new_papers: int):
    """Aktualisiert das Overview-Sheet mit Suchstatistiken"""
    try:
        overview_sheet = wb["📊_Overview"]
        
        # Suche nach existierendem Eintrag
        search_row = None
        for row_num, row in enumerate(overview_sheet.iter_rows(min_row=2), start=2):
            if row[1].value == search_term:  # Spalte B = Suchbegriff
                search_row = row_num
                break
        
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        
        if search_row:
            # Update existierenden Eintrag
            overview_sheet.cell(row=search_row, column=3, value=total_papers)  # Anzahl Papers
            overview_sheet.cell(row=search_row, column=4, value=current_time)  # Letztes Update
            overview_sheet.cell(row=search_row, column=5, value=new_papers)    # Neue Papers
            overview_sheet.cell(row=search_row, column=6, value="✅ Aktiv")     # Status
        else:
            # Neuer Eintrag
            new_row = [
                sheet_name,         # Sheet Name
                search_term,        # Suchbegriff
                total_papers,       # Anzahl Papers
                current_time,       # Letztes Update
                new_papers,         # Neue Papers
                "✅ Aktiv",         # Status
                current_time        # Erstellt am
            ]
            overview_sheet.append(new_row)
        
        wb.save(st.session_state["excel_template"]["file_path"])
    
    except Exception as e:
        st.warning(f"⚠️ Fehler beim Aktualisieren des Overview-Sheets: {str(e)}")

def get_search_statistics_from_excel() -> Dict:
    """Holt Statistiken aus der Excel-Datei"""
    wb = load_master_workbook()
    if not wb:
        return {}
    
    stats = {
        "total_sheets": len([s for s in wb.sheetnames if not s.startswith(("📊", "ℹ️"))]),
        "total_searches": 0,
        "total_papers": 0,
        "search_terms": []
    }
    
    if "📊_Overview" in wb.sheetnames:
        overview_sheet = wb["📊_Overview"]
        
        for row in overview_sheet.iter_rows(min_row=2):
            if row[1].value:  # Suchbegriff existiert
                stats["total_searches"] += 1
                stats["total_papers"] += row[2].value or 0
                stats["search_terms"].append({
                    "term": row[1].value,
                    "papers": row[2].value or 0,
                    "last_update": row[3].value,
                    "new_papers": row[4].value or 0
                })
    
    return stats

# =============== MEHRERE EMAIL-EMPFÄNGER FUNKTIONEN ===============

def parse_recipient_emails(email_string: str) -> List[str]:
    """Parst Email-String und gibt Liste gültiger Emails zurück"""
    if not email_string:
        return []
    
    # Split by comma and clean
    emails = [email.strip() for email in email_string.split(",")]
    
    # Basic email validation
    valid_emails = []
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    for email in emails:
        if email and email_pattern.match(email):
            valid_emails.append(email)
    
    return valid_emails

def send_real_email_multiple(to_emails: List[str], subject: str, message: str, attachment_path: str = None) -> tuple:
    """Sendet echte Email über SMTP an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    
    sender_email = settings.get("sender_email", "")
    sender_password = settings.get("sender_password", "")
    smtp_server = settings.get("smtp_server", "smtp.gmail.com")
    smtp_port = settings.get("smtp_port", 587)
    use_tls = settings.get("use_tls", True)
    
    if not all([sender_email, sender_password]):
        return False, "❌ Email-Konfiguration unvollständig (Absender/Passwort)"
    
    if not to_emails:
        return False, "❌ Keine Empfänger-Emails konfiguriert"
    
    try:
        # SMTP Server Setup
        server = smtplib.SMTP(smtp_server, smtp_port)
        
        if use_tls:
            context = ssl.create_default_context()
            server.starttls(context=context)
        
        server.login(sender_email, sender_password)
        
        successful_sends = 0
        failed_sends = []
        
        # Send to each recipient
        for recipient in to_emails:
            try:
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient
                msg['Subject'] = subject
                
                msg.attach(MIMEText(message, 'plain', 'utf-8'))
                
                # Add attachment if provided
                if attachment_path and os.path.exists(attachment_path):
                    with open(attachment_path, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {os.path.basename(attachment_path)}'
                        )
                        msg.attach(part)
                
                server.send_message(msg)
                successful_sends += 1
                
            except Exception as e:
                failed_sends.append(f"{recipient}: {str(e)}")
        
        server.quit()
        
        if successful_sends == len(to_emails):
            return True, f"✅ Email erfolgreich an alle {successful_sends} Empfänger gesendet"
        elif successful_sends > 0:
            return True, f"⚠️ Email an {successful_sends}/{len(to_emails)} Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        else:
            return False, f"❌ Email an keinen Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        
    except smtplib.SMTPAuthenticationError:
        return False, "❌ SMTP-Authentifizierung fehlgeschlagen - Prüfen Sie Email/Passwort"
    except smtplib.SMTPServerDisconnected:
        return False, "❌ SMTP-Server-Verbindung unterbrochen"
    except Exception as e:
        return False, f"❌ Email-Fehler: {str(e)}"

# =============== HAUPTFUNKTIONEN ===============

def show_dashboard():
    """Dashboard mit anklickbaren Suchhistorie und Excel-Integration"""
    st.subheader("📊 Dashboard - Excel-Integrierte Übersicht")
    
    # Excel-Statistiken holen
    excel_stats = get_search_statistics_from_excel()
    status = st.session_state["system_status"]
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("🔍 Gesamt Suchen", excel_stats.get("total_searches", 0))
    
    with col2:
        st.metric("📄 Papers (Excel)", excel_stats.get("total_papers", 0))
    
    with col3:
        st.metric("📧 Gesendete Emails", status["total_emails"])
    
    with col4:
        recipients = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
        st.metric("📧 Email-Empfänger", recipients)
    
    # Letzte Aktivität
    if status["last_search"]:
        try:
            last_search_time = datetime.datetime.fromisoformat(status["last_search"])
            time_diff = datetime.datetime.now() - last_search_time
            hours = time_diff.seconds // 3600
            minutes = (time_diff.seconds % 3600) // 60
            st.info(f"🕒 Letzte Suche: vor {time_diff.days}d {hours}h {minutes}min")
        except:
            st.info("🕒 Letzte Suche: Unbekannt")
    
    # Excel-Download im Dashboard
    excel_path = st.session_state["excel_template"]["file_path"]
    if os.path.exists(excel_path):
        with open(excel_path, "rb") as file:
            st.download_button(
                "📎 Excel-Datenbank herunterladen",
                data=file.read(),
                file_name=f"paper_database_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Suchhistorie aus Excel
    st.markdown("---")
    st.subheader("📋 Excel-basierte Suchhistorie")
    
    if excel_stats.get("search_terms"):
        # Sortiere nach letztem Update
        recent_terms = sorted(excel_stats["search_terms"], key=lambda x: x.get("last_update", ""), reverse=True)
        
        for term_info in recent_terms:
            search_term = term_info["term"]
            papers = term_info["papers"]
            last_update = term_info.get("last_update", "")[:16].replace('T', ' ')
            new_papers = term_info.get("new_papers", 0)
            
            col_search1, col_search2, col_search3 = st.columns([3, 1, 1])
            
            with col_search1:
                if st.button(f"🔍 **{search_term}** ({papers} Papers, {new_papers} neue)", 
                           key=f"search_btn_{search_term}"):
                    show_search_details_from_excel(search_term, term_info)
            
            with col_search2:
                st.write(f"📅 {last_update}")
            
            with col_search3:
                if st.button("📊 Excel", key=f"excel_btn_{search_term}"):
                    show_excel_sheet_content(search_term)
        
        # Quick Actions
        st.markdown("---")
        st.subheader("⚡ Quick Actions")
        
        col_quick1, col_quick2, col_quick3 = st.columns(3)
        
        with col_quick1:
            if st.button("🔄 **Alle Suchen wiederholen**"):
                repeat_all_searches_from_excel()
        
        with col_quick2:
            if st.button("📧 **Status-Email senden**"):
                send_status_email_multiple()
        
        with col_quick3:
            if st.button("📁 **Excel öffnen**"):
                offer_excel_download()
    
    else:
        st.info("📭 Noch keine Suchen durchgeführt. Starten Sie im Tab 'Paper-Suche'!")

def show_advanced_paper_search():
    """Erweiterte Paper-Suche mit Excel-Integration und mehreren Email-Empfängern"""
    st.subheader("🔍 Excel-Integrierte Paper-Suche")
    
    # Excel- und Email-Status
    excel_stats = get_search_statistics_from_excel()
    email_status = is_email_configured()
    recipient_count = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
    
    col_info1, col_info2, col_info3, col_info4 = st.columns(4)
    with col_info1:
        st.metric("📊 Excel-Sheets", excel_stats.get("total_sheets", 0))
    with col_info2:
        st.metric("📄 Papers in Excel", excel_stats.get("total_papers", 0))
    with col_info3:
        st.metric("📧 Email-Empfänger", recipient_count)
    with col_info4:
        st.metric("🔍 Durchsuchungen", excel_stats.get("total_searches", 0))
    
    if email_status:
        st.success(f"✅ Email-System bereit für **{recipient_count} Empfänger** | Excel-Integration: ✅ Aktiv")
    else:
        st.info("ℹ️ Email-System nicht konfiguriert | Excel-Integration: ✅ Aktiv")
    
    # Such-Interface
    with st.form("advanced_search_form"):
        col_search1, col_search2 = st.columns([3, 1])
        
        with col_search1:
            search_query = st.text_input(
                "**🔍 PubMed Suchbegriff:**",
                placeholder="z.B. 'diabetes genetics', 'machine learning radiology', 'COVID-19 treatment'",
                help="Durchsucht Excel auf bereits bekannte Papers und fügt nur neue hinzu"
            )
        
        with col_search2:
            max_results = st.number_input(
                "Max. Ergebnisse", 
                min_value=10, 
                max_value=500, 
                value=100
            )
        
        # Erweiterte Optionen
        with st.expander("🔧 Erweiterte Excel- & Email-Optionen"):
            col_adv1, col_adv2, col_adv3 = st.columns(3)
            
            with col_adv1:
                date_filter = st.selectbox(
                    "📅 Zeitraum:",
                    ["Alle", "Letztes Jahr", "Letzte 2 Jahre", "Letzte 5 Jahre", "Letzte 10 Jahre"],
                    index=2
                )
            
            with col_adv2:
                force_email = st.checkbox(
                    "📧 Email erzwingen", 
                    value=False,
                    help="Sendet Email auch wenn keine neuen Papers gefunden"
                )
            
            with col_adv3:
                show_existing = st.checkbox(
                    "📊 Bereits bekannte Papers anzeigen", 
                    value=False,
                    help="Zeigt auch Papers an, die bereits in Excel vorhanden sind"
                )
        
        search_button = st.form_submit_button("🚀 **EXCEL-INTEGRIERTE PAPER-SUCHE**", type="primary")
    
    # Quick Search aus Excel-Historie
    if excel_stats.get("search_terms"):
        st.write("**⚡ Schnellsuche (aus Excel-Historie):**")
        recent_terms = sorted(excel_stats["search_terms"], key=lambda x: x.get("last_update", ""), reverse=True)[:5]
        
        cols = st.columns(min(len(recent_terms), 5))
        for i, term_info in enumerate(recent_terms):
            term = term_info["term"]
            papers = term_info["papers"]
            with cols[i]:
                if st.button(f"🔍 {term[:15]}... ({papers})", key=f"quick_{i}"):
                    execute_excel_integrated_search(term, 50, "Letzte 2 Jahre", False, False)
    
    # Hauptsuche ausführen
    if search_button and search_query:
        execute_excel_integrated_search(search_query, max_results, date_filter, force_email, show_existing)
    
    # *** MANUELLER EMAIL-VERSAND BEREICH ***
    show_manual_email_section()

def execute_excel_integrated_search(query: str, max_results: int, date_filter: str, force_email: bool, show_existing: bool):
    """Führt Excel-integrierte Paper-Suche durch"""
    st.markdown("---")
    st.subheader(f"🔍 **Excel-integrierte Suche:** '{query}'")
    
    # Progress Tracking
    progress_container = st.container()
    with progress_container:
        progress_bar = st.progress(0)
        status_text = st.empty()
    
    try:
        # 1. Lade Excel für Duplikatsprüfung
        status_text.text("📊 Lade Excel-Datei für Duplikatsprüfung...")
        progress_bar.progress(0.1)
        
        wb = load_master_workbook()
        if not wb:
            st.error("❌ Excel-Datei konnte nicht geladen werden!")
            return
        
        # 2. PubMed-Suche durchführen
        status_text.text("🔍 Durchsuche PubMed-Datenbank...")
        progress_bar.progress(0.3)
        
        advanced_query = build_advanced_search_query(query, date_filter)
        current_papers = perform_comprehensive_pubmed_search(advanced_query, max_results)
        
        progress_bar.progress(0.5)
        
        if not current_papers:
            st.error(f"❌ **Keine Papers für '{query}' gefunden!**")
            progress_bar.empty()
            status_text.empty()
            return
        
        # 3. Excel-Integration: Prüfe auf neue Papers
        status_text.text("🔍 Prüfe Papers gegen Excel-Datenbank...")
        progress_bar.progress(0.7)
        
        added_count, new_papers = add_new_papers_to_excel(query, current_papers)
        
        # 4. Ergebnisse anzeigen
        progress_bar.progress(0.9)
        status_text.text("📊 Bereite Ergebnisse vor...")
        
        if added_count > 0:
            st.success(f"🆕 **{added_count} NEUE Papers gefunden und zu Excel hinzugefügt!** (von {len(current_papers)} gesamt)")
            st.balloons()
            
            # Email senden bei neuen Papers
            if is_email_configured() and (force_email or should_send_email(added_count)):
                send_excel_integrated_email_multiple(query, new_papers, len(current_papers), added_count)
        else:
            st.info(f"ℹ️ **Keine neuen Papers** - Alle {len(current_papers)} Papers bereits in Excel vorhanden")
            
            # Email erzwingen wenn gewünscht
            if force_email and is_email_configured():
                send_excel_integrated_email_multiple(query, [], len(current_papers), 0)
        
        # 5. Detaillierte Ergebnisse anzeigen
        display_excel_integrated_results(current_papers, new_papers, query, added_count, show_existing)
        
        # 6. Für manuellen Email-Versand speichern
        st.session_state["current_search_results"] = {
            "search_term": query,
            "papers": current_papers,
            "new_papers": new_papers,
            "added_count": added_count,
            "timestamp": datetime.datetime.now().isoformat()
        }
        
        # 7. System-Status aktualisieren
        progress_bar.progress(1.0)
        status_text.text("✅ Excel-integrierte Suche abgeschlossen!")
        
        st.session_state["system_status"]["total_searches"] += 1
        st.session_state["system_status"]["total_papers"] += added_count
        st.session_state["system_status"]["last_search"] = datetime.datetime.now().isoformat()
        
        time.sleep(1)
        progress_bar.empty()
        status_text.empty()
        
    except Exception as e:
        progress_bar.empty()
        status_text.empty()
        st.error(f"❌ **Fehler bei der Excel-integrierten Suche:** {str(e)}")

def send_excel_integrated_email_multiple(search_term: str, new_papers: List[Dict], total_found: int, added_count: int):
    """Sendet Email für Excel-integrierte Suche an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not recipient_emails:
        st.warning("⚠️ Keine Email-Empfänger konfiguriert!")
        return
    
    # Subject generieren
    if added_count > 0:
        subject = f"📊 {added_count} neue Papers für '{search_term}' - Excel aktualisiert"
    else:
        subject = f"📊 Keine neuen Papers für '{search_term}' - Excel-Check durchgeführt"
    
    # Sheet-Name ermitteln
    sheet_name = generate_sheet_name(search_term)
    
    # Papers-Liste formatieren (nur neue)
    if new_papers:
        papers_list = ""
        for i, paper in enumerate(new_papers[:8], 1):
            title = paper.get("Title", "Unbekannt")[:70]
            authors = paper.get("Authors", "n/a")[:40]
            journal = paper.get("Journal", "n/a")
            year = paper.get("Year", "n/a")
            pmid = paper.get("PMID", "n/a")
            
            papers_list += f"\n{i}. **{title}...**\n"
            papers_list += f"   👥 {authors}...\n"
            papers_list += f"   📚 {journal} ({year}) | PMID: {pmid}\n\n"
        
        if len(new_papers) > 8:
            papers_list += f"... und {len(new_papers) - 8} weitere neue Papers (siehe Excel-Datei)\n"
    else:
        papers_list = "\nKeine neuen Papers gefunden - alle Papers bereits in Excel-Datenbank vorhanden.\n"
    
    # Message generieren
    message = f"""📊 **Excel-Integrierte Paper-Suche - Ergebnisse**

📅 **Datum:** {datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}
🔍 **Suchbegriff:** '{search_term}'
📊 **Gefundene Papers:** {total_found}
🆕 **Neue Papers:** {added_count}
📊 **Bereits bekannt:** {total_found - added_count}
📁 **Excel-Sheet:** {sheet_name}

{'-' * 60}
🆕 **NEUE PAPERS:**
{papers_list}

📎 **Excel-Integration:**
✅ Alle neuen Papers wurden automatisch zur Excel-Datei hinzugefügt
✅ Duplikate wurden automatisch erkannt und übersprungen
✅ Sheet für diesen Suchbegriff wurde aktualisiert
📋 Sheet-Name: {sheet_name}

📧 **Email-Info:**
📧 Versendet an: {len(recipient_emails)} Empfänger
{chr(10).join([f"   • {email}" for email in recipient_emails])}
📎 Excel-Datei als Anhang beigefügt

Mit freundlichen Grüßen,
Ihr Excel-integriertes Paper-Suche System"""
    
    # Excel als Anhang
    excel_path = st.session_state["excel_template"]["file_path"]
    attachment_path = excel_path if os.path.exists(excel_path) else None
    
    # Email senden
    with st.spinner(f"📧 Sende Excel-integrierte Email an {len(recipient_emails)} Empfänger..."):
        success, status_message = send_real_email_multiple(recipient_emails, subject, message, attachment_path)
    
    # Email-Historie
    email_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "type": "Excel-Integriert",
        "search_term": search_term,
        "recipients": recipient_emails,
        "recipient_count": len(recipient_emails),
        "subject": subject,
        "paper_count": added_count,
        "total_found": total_found,
        "success": success,
        "status": status_message,
        "has_attachment": attachment_path is not None,
        "sheet_name": sheet_name
    }
    
    st.session_state["email_history"].append(email_entry)
    
    # Ergebnis anzeigen
    if success:
        st.session_state["system_status"]["total_emails"] += 1
        st.success(f"📧 **Excel-integrierte Email erfolgreich versendet!**\n{status_message}")
        
        with st.expander("📋 Email-Details"):
            st.write(f"**📧 Empfänger:** {len(recipient_emails)}")
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
            st.write(f"**🆕 Neue Papers:** {added_count}")
            st.write(f"**📊 Gesamt gefunden:** {total_found}")
            st.write(f"**📁 Excel-Sheet:** {sheet_name}")
            st.write(f"**📎 Anhang:** {'✅ Excel-Datei' if attachment_path else '❌ Kein Anhang'}")
    else:
        st.error(f"❌ **Email-Fehler:** {status_message}")

def show_manual_email_section():
    """Manueller Email-Versand nach Suche für mehrere Empfänger"""
    if st.session_state.get("current_search_results"):
        st.markdown("---")
        st.subheader("📧 Manueller Email-Versand (Excel-Integriert)")
        
        current_results = st.session_state["current_search_results"]
        search_term = current_results.get("search_term", "")
        papers = current_results.get("papers", [])
        new_papers = current_results.get("new_papers", [])
        added_count = current_results.get("added_count", 0)
        
        if papers:
            col_email1, col_email2, col_email3 = st.columns(3)
            
            with col_email1:
                st.metric("📄 Verfügbare Papers", len(papers))
            
            with col_email2:
                st.metric("🆕 Neue Papers", added_count)
            
            with col_email3:
                recipient_count = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
                st.metric("📧 Empfänger", recipient_count)
            
            # Email-Optionen
            email_status = is_email_configured()
            
            if email_status and recipient_count > 0:
                col_send1, col_send2 = st.columns(2)
                
                with col_send1:
                    if st.button(f"📧 **Alle Papers emailen** ({len(papers)})", type="primary"):
                        send_manual_search_email_multiple(search_term, papers, "Alle Papers")
                
                with col_send2:
                    if added_count > 0 and st.button(f"📧 **Nur neue Papers emailen** ({added_count})", type="secondary"):
                        send_manual_search_email_multiple(search_term, new_papers, "Nur neue Papers")
            else:
                if not email_status:
                    st.warning("⚠️ **Email-Versand nicht möglich:** Konfigurieren Sie Email-Einstellungen im entsprechenden Tab")
                elif recipient_count == 0:
                    st.warning("⚠️ **Keine Empfänger konfiguriert:** Fügen Sie Email-Adressen in der Email-Konfiguration hinzu")

def send_manual_search_email_multiple(search_term: str, papers: List[Dict], email_type: str):
    """Sendet manuelle Email für Suchergebnisse an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not recipient_emails:
        st.error("❌ Keine Empfänger konfiguriert!")
        return
    
    # Subject generieren
    subject = f"📧 {email_type}: {len(papers)} Papers für '{search_term}' (Manuell)"
    
    # Papers-Liste formatieren
    papers_list = ""
    for i, paper in enumerate(papers[:15], 1):  # Erste 15 Papers
        title = paper.get("Title", "Unbekannt")[:70]
        authors = paper.get("Authors", "n/a")[:50]
        journal = paper.get("Journal", "n/a")
        year = paper.get("Year", "n/a")
        pmid = paper.get("PMID", "n/a")
        
        papers_list += f"\n{i}. **{title}...**\n"
        papers_list += f"   👥 {authors}...\n"
        papers_list += f"   📚 {journal} ({year}) | PMID: {pmid}\n\n"
    
    if len(papers) > 15:
        papers_list += f"... und {len(papers) - 15} weitere Papers (siehe Excel-Datei)\n"
    
    # Message generieren
    message = f"""📧 **Manueller Email-Versand - Paper-Suche**

📅 **Datum:** {datetime.datetime.now().strftime("%d.%m.%Y %H:%M")}
🔍 **Suchbegriff:** '{search_term}'
📊 **Typ:** {email_type}
📄 **Anzahl Papers:** {len(papers)}
📧 **Empfänger:** {len(recipient_emails)}

📧 **Empfänger-Liste:**
{chr(10).join([f"   • {email}" for email in recipient_emails])}

{'-' * 50}
📋 **PAPERS:**
{papers_list}

📎 **Excel-Datei:** Die aktualisierte Excel-Datei ist als Anhang beigefügt.

ℹ️ **Hinweis:** Diese Email wurde manuell über das Paper-Suche System versendet.

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""
    
    # Excel als Anhang
    excel_path = st.session_state["excel_template"]["file_path"]
    attachment_path = excel_path if os.path.exists(excel_path) else None
    
    # Email senden
    with st.spinner(f"📧 Sende Email an {len(recipient_emails)} Empfänger..."):
        success, status_message = send_real_email_multiple(recipient_emails, subject, message, attachment_path)
    
    # Email-Historie
    email_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "type": f"Manuell - {email_type}",
        "search_term": search_term,
        "recipients": recipient_emails,
        "recipient_count": len(recipient_emails),
        "subject": subject,
        "paper_count": len(papers),
        "success": success,
        "status": status_message,
        "has_attachment": attachment_path is not None
    }
    
    st.session_state["email_history"].append(email_entry)
    
    # Ergebnis anzeigen
    if success:
        st.session_state["system_status"]["total_emails"] += 1
        st.success(f"📧 **Email erfolgreich versendet!**\n{status_message}")
        st.balloons()
        
        # Details anzeigen
        with st.expander("📋 Email-Details anzeigen"):
            st.write(f"**📧 Empfänger ({len(recipient_emails)}):**")
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
            st.write(f"**📄 Papers:** {len(papers)}")
            st.write(f"**📎 Anhang:** {'✅ Excel-Datei' if attachment_path else '❌ Kein Anhang'}")
    else:
        st.error(f"❌ **Email-Fehler:** {status_message}")

def show_email_config():
    """Email-Konfiguration mit mehreren Empfängern"""
    st.subheader("📧 Email-Konfiguration (Mehrere Empfänger)")
    
    settings = st.session_state.get("email_settings", {})
    
    # Email-Setup Hilfe
    with st.expander("📖 Email-Setup Hilfe & Mehrere Empfänger"):
        st.info("""
        **Für Gmail (empfohlen):**
        1. ✅ 2-Faktor-Authentifizierung aktivieren
        2. ✅ App-Passwort erstellen (nicht normales Passwort!)
        3. ✅ SMTP: smtp.gmail.com, Port: 587, TLS: An
        
        **Mehrere Empfänger:**
        • Trennen Sie mehrere Email-Adressen mit Kommas
        • Beispiel: user1@gmail.com, user2@outlook.com, user3@company.de
        • Whitespaces werden automatisch entfernt
        
        **Für Outlook/Hotmail:**
        - SMTP: smtp-mail.outlook.com, Port: 587
        """)
    
    with st.form("email_config_form"):
        st.subheader("📬 Grundeinstellungen")
        
        col1, col2 = st.columns(2)
        
        with col1:
            sender_email = st.text_input(
                "Absender Email *", 
                value=settings.get("sender_email", ""),
                placeholder="absender@gmail.com"
            )
            
            smtp_server = st.text_input(
                "SMTP Server *",
                value=settings.get("smtp_server", "smtp.gmail.com")
            )
            
            auto_notifications = st.checkbox(
                "Automatische Benachrichtigungen", 
                value=settings.get("auto_notifications", True)
            )
        
        with col2:
            smtp_port = st.number_input(
                "SMTP Port *",
                value=settings.get("smtp_port", 587),
                min_value=1,
                max_value=65535
            )
            
            min_papers = st.number_input(
                "Min. Papers für Benachrichtigung", 
                value=settings.get("min_papers", 1),
                min_value=1,
                max_value=100
            )
            
            use_tls = st.checkbox(
                "TLS Verschlüsselung verwenden (empfohlen)",
                value=settings.get("use_tls", True)
            )
        
        # MEHRERE EMPFÄNGER - Text Area
        recipient_emails = st.text_area(
            "📧 Empfänger Email-Adressen * (mehrere mit Komma trennen)",
            value=settings.get("recipient_emails", ""),
            placeholder="empfaenger1@example.com, empfaenger2@gmail.com, empfaenger3@company.de",
            help="Mehrere Email-Adressen mit Komma trennen. Beispiel: user1@gmail.com, user2@outlook.com",
            height=80
        )
        
        sender_password = st.text_input(
            "Email Passwort / App-Passwort *",
            value=settings.get("sender_password", ""),
            type="password",
            help="Für Gmail: App-spezifisches Passwort verwenden!"
        )
        
        # Email-Vorlagen
        st.subheader("📝 Email-Vorlagen")
        
        subject_template = st.text_input(
            "Betreff-Vorlage",
            value=settings.get("subject_template", "🔬 {count} neue Papers für '{search_term}'"),
            help="Platzhalter: {count}, {search_term}, {frequency}"
        )
        
        message_template = st.text_area(
            "Nachricht-Vorlage",
            value=settings.get("message_template", """📧 Automatische Paper-Benachrichtigung

📅 Datum: {date}
🔍 Suchbegriff: '{search_term}'
📊 Neue Papers: {count}

📋 Neue Papers:
{new_papers_list}

📎 Excel-Datei: {excel_file}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""),
            height=200,
            help="Platzhalter: {date}, {search_term}, {count}, {frequency}, {new_papers_list}, {excel_file}"
        )
        
        if st.form_submit_button("💾 **Email-Einstellungen speichern**", type="primary"):
            # Validiere Email-Adressen
            recipient_list = parse_recipient_emails(recipient_emails)
            
            if not recipient_list:
                st.error("❌ Mindestens eine gültige Empfänger-Email erforderlich!")
            else:
                new_settings = {
                    "sender_email": sender_email,
                    "recipient_emails": recipient_emails,
                    "smtp_server": smtp_server,
                    "smtp_port": smtp_port,
                    "sender_password": sender_password,
                    "use_tls": use_tls,
                    "auto_notifications": auto_notifications,
                    "min_papers": min_papers,
                    "subject_template": subject_template,
                    "message_template": message_template,
                    "parsed_recipients": recipient_list  # Store parsed list
                }
                
                st.session_state["email_settings"] = new_settings
                st.success(f"✅ Email-Einstellungen gespeichert! **{len(recipient_list)} Empfänger** konfiguriert:")
                for i, email in enumerate(recipient_list, 1):
                    st.write(f"   {i}. 📧 {email}")
    
    # Zeige konfigurierte Empfänger
    if settings.get("recipient_emails"):
        recipient_list = parse_recipient_emails(settings.get("recipient_emails", ""))
        if recipient_list:
            st.info(f"📧 **Aktuell konfigurierte Empfänger ({len(recipient_list)}):**")
            cols = st.columns(min(len(recipient_list), 3))
            for i, email in enumerate(recipient_list):
                with cols[i % 3]:
                    st.write(f"✅ {email}")
    
    # Test-Email
    st.markdown("---")
    st.subheader("🧪 Email-System testen")
    
    col_test1, col_test2 = st.columns(2)
    
    with col_test1:
        if st.button("📧 **Test-Email an alle Empfänger senden**", type="primary"):
            send_test_email_multiple()
    
    with col_test2:
        if st.button("📊 **Email-Status prüfen**"):
            check_email_status_multiple()

def send_test_email_multiple():
    """Sendet Test-Email an alle konfigurierten Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not settings.get("sender_email") or not recipient_emails:
        st.error("❌ Email-Konfiguration unvollständig!")
        return
    
    subject = "🧪 Test-Email vom Paper-Suche System (Mehrere Empfänger)"
    message = f"""Dies ist eine Test-Email vom Paper-Suche System mit Unterstützung für mehrere Empfänger.

📅 Gesendet am: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
📧 Von: {settings.get('sender_email')}
📧 An: {len(recipient_emails)} Empfänger

Empfänger-Liste:
{chr(10).join([f"• {email}" for email in recipient_emails])}

✅ Wenn Sie diese Email erhalten, funktioniert das Email-System korrekt!

System-Informationen:
• SMTP Server: {settings.get('smtp_server')}
• Port: {settings.get('smtp_port')}
• TLS: {'Aktiviert' if settings.get('use_tls') else 'Deaktiviert'}
• Empfänger: {len(recipient_emails)}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""
    
    success, status_message = send_real_email_multiple(
        recipient_emails, 
        subject, 
        message
    )
    
    if success:
        st.success(f"✅ **Test-Email erfolgreich gesendet!** {status_message}")
        st.balloons()
    else:
        st.error(f"❌ **Test-Email fehlgeschlagen:** {status_message}")

def check_email_status_multiple():
    """Prüft Email-Status mit mehreren Empfängern"""
    settings = st.session_state.get("email_settings", {})
    
    st.write("**📊 Email-Konfiguration Status:**")
    
    # Prüfe Konfiguration
    sender_ok = bool(settings.get("sender_email"))
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    recipients_ok = len(recipient_emails) > 0
    password_ok = bool(settings.get("sender_password"))
    
    st.write(f"📧 Absender Email: {'✅' if sender_ok else '❌'} {settings.get('sender_email', 'Nicht konfiguriert')}")
    st.write(f"📧 Empfänger Emails: {'✅' if recipients_ok else '❌'} {len(recipient_emails)} konfiguriert")
    
    if recipients_ok:
        with st.expander(f"📧 Empfänger-Liste ({len(recipient_emails)})"):
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
    
    st.write(f"🔑 Passwort: {'✅' if password_ok else '❌'} {'Konfiguriert' if password_ok else 'Nicht konfiguriert'}")
    st.write(f"🔒 SMTP Server: {settings.get('smtp_server', 'smtp.gmail.com')}:{settings.get('smtp_port', 587)}")
    st.write(f"🔐 TLS: {'✅ Aktiviert' if settings.get('use_tls', True) else '❌ Deaktiviert'}")
    
    # Gesamtstatus
    if sender_ok and recipients_ok and password_ok:
        st.success(f"✅ **Email-System vollständig konfiguriert für {len(recipient_emails)} Empfänger!**")
    else:
        st.error("❌ **Email-System nicht vollständig konfiguriert!**")

# =============== WEITERE FUNKTIONEN ===============

def perform_comprehensive_pubmed_search(query: str, max_results: int) -> List[Dict[str, Any]]:
    """Umfassende PubMed-Suche"""
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    
    # 1. esearch - hole PMIDs
    search_url = f"{base_url}esearch.fcgi"
    params = {
        "db": "pubmed",
        "term": query,
        "retmode": "json",
        "retmax": max_results,
        "email": "research.system@papersearch.com",
        "tool": "ScientificPaperSearchSystem",
        "sort": "relevance"
    }
    
    try:
        with st.spinner("🔍 Verbinde zu PubMed..."):
            response = requests.get(search_url, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            pmids = data.get("esearchresult", {}).get("idlist", [])
            total_count = int(data.get("esearchresult", {}).get("count", 0))
            
            st.write(f"📊 **PubMed Datenbank:** {total_count:,} Papers verfügbar, {len(pmids)} werden abgerufen")
            
            if not pmids:
                return []
            
            # 2. efetch - hole Details in Batches
            return fetch_paper_details_batch(pmids)
            
    except requests.exceptions.RequestException as e:
        st.error(f"❌ **PubMed Verbindungsfehler:** {str(e)}")
        return []
    except Exception as e:
        st.error(f"❌ **PubMed Suchfehler:** {str(e)}")
        return []

def fetch_paper_details_batch(pmids: List[str], batch_size: int = 50) -> List[Dict[str, Any]]:
    """Holt Paper-Details in Batches für bessere Performance"""
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    all_papers = []
    
    # Teile PMIDs in Batches
    batches = [pmids[i:i + batch_size] for i in range(0, len(pmids), batch_size)]
    
    progress_bar = st.progress(0)
    batch_status = st.empty()
    
    for batch_idx, batch_pmids in enumerate(batches):
        try:
            batch_status.text(f"📥 Batch {batch_idx + 1}/{len(batches)}: {len(batch_pmids)} Papers...")
            
            params = {
                "db": "pubmed",
                "id": ",".join(batch_pmids),
                "retmode": "xml",
                "email": "research.system@papersearch.com",
                "tool": "ScientificPaperSearchSystem"
            }
            
            response = requests.get(base_url, params=params, timeout=60)
            response.raise_for_status()
            
            # Parse XML
            root = ET.fromstring(response.content)
            articles = root.findall(".//PubmedArticle")
            
            for article in articles:
                paper_data = parse_pubmed_article(article)
                if paper_data:
                    all_papers.append(paper_data)
            
            # Progress Update
            progress = (batch_idx + 1) / len(batches)
            progress_bar.progress(progress)
            
            # Rate limiting
            time.sleep(0.5)
            
        except Exception as e:
            st.warning(f"⚠️ Batch {batch_idx + 1} Fehler: {str(e)}")
            continue
    
    progress_bar.empty()
    batch_status.empty()
    
    return all_papers

def parse_pubmed_article(article) -> Dict[str, Any]:
    """Erweiterte Artikel-Parsing mit mehr Feldern"""
    try:
        # PMID
        pmid_elem = article.find(".//PMID")
        pmid = pmid_elem.text if pmid_elem is not None else ""
        
        # Title
        title_elem = article.find(".//ArticleTitle")
        title = title_elem.text if title_elem is not None else "Titel nicht verfügbar"
        
        # Abstract (alle Teile)
        abstract_parts = []
        for abstract_elem in article.findall(".//AbstractText"):
            if abstract_elem.text:
                label = abstract_elem.get("Label", "")
                text = abstract_elem.text
                if label and label.upper() not in ["UNLABELLED", "UNASSIGNED"]:
                    abstract_parts.append(f"**{label}:** {text}")
                else:
                    abstract_parts.append(text)
        
        abstract = "\n\n".join(abstract_parts) if abstract_parts else "Kein Abstract verfügbar"
        
        # Journal Info
        journal_elem = article.find(".//Journal/Title")
        journal = journal_elem.text if journal_elem is not None else "Journal unbekannt"
        
        # Publication Date
        year_elem = article.find(".//PubDate/Year")
        year = year_elem.text if year_elem is not None else "Unbekannt"
        
        # Authors
        authors = []
        for author in article.findall(".//Author"):
            lastname = author.find("LastName")
            forename = author.find("ForeName")
            
            if lastname is not None:
                author_name = lastname.text or ""
                if forename is not None:
                    author_name = f"{author_name}, {forename.text}"
                authors.append(author_name)
        
        authors_str = "; ".join(authors[:8])  # Erste 8 Autoren
        if len(authors) > 8:
            authors_str += f" et al. (+{len(authors) - 8} weitere)"
        
        # DOI
        doi = ""
        for article_id in article.findall(".//ArticleId"):
            if article_id.get("IdType") == "doi":
                doi = article_id.text
                break
        
        return {
            "PMID": pmid,
            "Title": title,
            "Abstract": abstract,
            "Journal": journal,
            "Year": year,
            "Authors": authors_str,
            "DOI": doi,
            "URL": f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/",
            "Search_Date": datetime.datetime.now().isoformat(),
            "Is_New": True,
            "Has_DOI": bool(doi)
        }
        
    except Exception as e:
        st.warning(f"⚠️ Fehler beim Parsen eines Artikels: {str(e)}")
        return None

def display_excel_integrated_results(all_papers: List[Dict], new_papers: List[Dict], query: str, added_count: int, show_existing: bool):
    """Zeigt Ergebnisse der Excel-integrierten Suche an"""
    
    # Statistiken
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
    with col_stat1:
        st.metric("📄 Gefunden", len(all_papers))
    with col_stat2:
        st.metric("🆕 Neue Papers", added_count)
    with col_stat3:
        st.metric("📊 Bereits bekannt", len(all_papers) - added_count)
    with col_stat4:
        st.metric("💾 In Excel gespeichert", added_count)
    
    # Neue Papers hervorheben
    if new_papers:
        st.subheader(f"🆕 Neue Papers ({len(new_papers)})")
        
        with st.expander(f"📋 Alle {len(new_papers)} neuen Papers anzeigen", expanded=True):
            for i, paper in enumerate(new_papers[:10], 1):  # Zeige erste 10
                with st.container():
                    col_paper1, col_paper2 = st.columns([3, 1])
                    
                    with col_paper1:
                        st.write(f"**{i}. {paper.get('Title', 'Unbekannt')[:100]}...**")
                        st.write(f"👥 {paper.get('Authors', 'n/a')[:80]}...")
                        st.write(f"📚 {paper.get('Journal', 'n/a')} ({paper.get('Year', 'n/a')})")
                        if paper.get('URL'):
                            st.markdown(f"🔗 [**PubMed**]({paper.get('URL')})")
                    
                    with col_paper2:
                        st.success("🆕 NEU")
                        st.write(f"PMID: {paper.get('PMID', 'n/a')}")
            
            if len(new_papers) > 10:
                st.info(f"... und {len(new_papers) - 10} weitere neue Papers (siehe Excel-Datei)")
    
    # Bereits bekannte Papers (optional)
    if show_existing and (len(all_papers) - added_count) > 0:
        existing_papers = [p for p in all_papers if p not in new_papers]
        
        with st.expander(f"📊 Bereits bekannte Papers ({len(existing_papers)})", expanded=False):
            for i, paper in enumerate(existing_papers[:5], 1):  # Zeige erste 5
                with st.container():
                    col_paper1, col_paper2 = st.columns([3, 1])
                    
                    with col_paper1:
                        st.write(f"**{i}. {paper.get('Title', 'Unbekannt')[:100]}...**")
                        st.write(f"👥 {paper.get('Authors', 'n/a')[:80]}...")
                        st.write(f"📚 {paper.get('Journal', 'n/a')} ({paper.get('Year', 'n/a')})")
                    
                    with col_paper2:
                        st.info("📊 BEKANNT")
                        st.write(f"PMID: {paper.get('PMID', 'n/a')}")
            
            if len(existing_papers) > 5:
                st.write(f"... und {len(existing_papers) - 5} weitere bereits bekannte Papers")

def generate_sheet_name(search_term: str) -> str:
    """Generiert gültigen Excel-Sheet-Namen"""
    # Excel Sheet Namen dürfen max 31 Zeichen haben und bestimmte Zeichen nicht enthalten
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    
    clean_name = search_term
    for char in invalid_chars:
        clean_name = clean_name.replace(char, '_')
    
    # Entferne multiple Unterstriche und trimme
    clean_name = re.sub(r'_+', '_', clean_name).strip('_')
    
    # Kürze auf 25 Zeichen (lasse Platz für eventuelle Suffixe)
    if len(clean_name) > 25:
        clean_name = clean_name[:25]
    
    return clean_name

def build_advanced_search_query(query: str, date_filter: str) -> str:
    """Erweiterte Suchanfrage mit Filtern"""
    query_parts = [query]
    
    if date_filter != "Alle":
        current_year = datetime.datetime.now().year
        if date_filter == "Letztes Jahr":
            query_parts.append(f"AND {current_year-1}:{current_year}[dp]")
        elif date_filter == "Letzte 2 Jahre":
            query_parts.append(f"AND {current_year-2}:{current_year}[dp]")
        elif date_filter == "Letzte 5 Jahre":
            query_parts.append(f"AND {current_year-5}:{current_year}[dp]")
        elif date_filter == "Letzte 10 Jahre":
            query_parts.append(f"AND {current_year-10}:{current_year}[dp]")
    
    return " ".join(query_parts)

def is_email_configured() -> bool:
    """Prüft Email-Konfiguration für mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    return (bool(settings.get("sender_email")) and 
            len(recipient_emails) > 0 and
            bool(settings.get("sender_password")))

def should_send_email(paper_count: int) -> bool:
    """Prüft ob Email gesendet werden soll"""
    settings = st.session_state.get("email_settings", {})
    return (settings.get("auto_notifications", False) and
            paper_count >= settings.get("min_papers", 1) and
            is_email_configured())

# =============== STATUS UND WIEDERHOLUNGSFUNKTIONEN ===============

def send_status_email_multiple():
    """Sendet Status-Email mit aktueller Übersicht an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not is_email_configured():
        st.error("❌ Email nicht konfiguriert! Bitte konfigurieren Sie die Email-Einstellungen.")
        return
    
    # System-Status sammeln
    status = st.session_state["system_status"]
    excel_stats = get_search_statistics_from_excel()
    email_history = st.session_state.get("email_history", [])
    
    # Subject
    subject = f"📊 System-Status Report - {datetime.datetime.now().strftime('%d.%m.%Y')}"
    
    # Message erstellen
    message = f"""📊 **SYSTEM-STATUS REPORT**
    
📅 **Berichts-Datum:** {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}

📈 **SYSTEM-STATISTIKEN:**
• 🔍 Gesamt Suchen: {excel_stats.get('total_searches', 0)}
• 📄 Papers in Excel: {excel_stats.get('total_papers', 0)}
• 📊 Excel Sheets: {excel_stats.get('total_sheets', 0)}
• 📧 Gesendete Emails: {len(email_history)}
• 📧 Email-Empfänger: {len(recipient_emails)}

📋 **LETZTE SUCHAKTIVITÄTEN (Excel-basiert):**"""

    # Letzte Suchen aus Excel hinzufügen
    if excel_stats.get("search_terms"):
        recent_searches = sorted(excel_stats["search_terms"], key=lambda x: x.get("last_update", ""), reverse=True)[:5]
        for i, search in enumerate(recent_searches, 1):
            term = search.get("term", "Unbekannt")
            papers = search.get("papers", 0)
            new_papers = search.get("new_papers", 0)
            last_update = search.get("last_update", "")[:16].replace('T', ' ')
            
            message += f"\n{i}. 🔍 {term} ({papers} Papers, {new_papers} neu) - {last_update}"
    
    message += f"""

📧 **EMAIL-EMPFÄNGER ({len(recipient_emails)}):**
{chr(10).join([f"• {email}" for email in recipient_emails])}

📎 **EXCEL-DATEI:** 
Die aktuelle Master Excel-Datei enthält {excel_stats.get('total_sheets', 0)} Sheets mit insgesamt {excel_stats.get('total_papers', 0)} Papers.

---
Dieser Report wurde automatisch generiert.
System: Paper-Suche & Email-System v3.0 (Excel-Integration + Mehrere Empfänger)"""
    
    # Email senden mit Excel-Anhang
    template_path = st.session_state["excel_template"]["file_path"]
    excel_path = template_path if os.path.exists(template_path) else None
    
    success, status_message = send_real_email_multiple(
        recipient_emails, 
        subject, 
        message,
        excel_path
    )
    
    # Email-Historie aktualisieren
    email_entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "type": "Status-Report",
        "recipients": recipient_emails,
        "recipient_count": len(recipient_emails),
        "subject": subject,
        "success": success,
        "status": status_message,
        "has_attachment": excel_path is not None
    }
    
    st.session_state["email_history"].append(email_entry)
    
    # Update System-Status
    if success:
        st.session_state["system_status"]["total_emails"] += 1
    
    # Ergebnis anzeigen
    if success:
        st.success(f"📧 **Status-Email erfolgreich an {len(recipient_emails)} Empfänger gesendet!**")
        st.balloons()
    else:
        st.error(f"❌ **Status-Email Fehler:** {status_message}")

def repeat_all_searches_from_excel():
    """Wiederholt alle Suchen basierend auf Excel-Daten"""
    excel_stats = get_search_statistics_from_excel()
    search_terms = excel_stats.get("search_terms", [])
    
    if not search_terms:
        st.info("📭 Keine Suchhistorie in Excel vorhanden.")
        return
    
    st.info(f"🔄 Wiederhole {len(search_terms)} Suchen basierend auf Excel-Daten...")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_new_papers = 0
    
    for i, term_info in enumerate(search_terms):
        search_term = term_info.get("term", "")
        if not search_term:
            continue
        
        try:
            status_text.text(f"🔍 Suche {i+1}/{len(search_terms)}: '{search_term}'...")
            
            # Führe Excel-integrierte Suche durch
            current_papers = perform_comprehensive_pubmed_search(search_term, 100)
            
            if current_papers:
                # Füge neue Papers zur Excel hinzu
                added_count, new_papers = add_new_papers_to_excel(search_term, current_papers)
                
                if added_count > 0:
                    # Sende Email wenn konfiguriert
                    if should_send_email(added_count):
                        send_excel_integrated_email_multiple(search_term, new_papers, len(current_papers), added_count)
                    
                    total_new_papers += added_count
                    st.write(f"✅ **{search_term}:** {added_count} neue Papers")
                else:
                    st.write(f"ℹ️ **{search_term}:** Keine neuen Papers")
            else:
                st.write(f"⚠️ **{search_term}:** Keine Papers gefunden")
            
            # Progress update
            progress_bar.progress((i + 1) / len(search_terms))
            time.sleep(1)  # Rate limiting
            
        except Exception as e:
            st.error(f"❌ Fehler bei '{search_term}': {str(e)}")
            continue
    
    progress_bar.empty()
    status_text.empty()
    
    # Ergebnis
    if total_new_papers > 0:
        st.success(f"🎉 **Wiederholung abgeschlossen!** {total_new_papers} neue Papers insgesamt gefunden!")
        st.balloons()
    else:
        st.info("ℹ️ **Wiederholung abgeschlossen.** Keine neuen Papers gefunden.")

# =============== WEITERE TAB-FUNKTIONEN ===============

def show_search_details_from_excel(search_term: str, term_info: Dict):
    """Zeigt Details einer Suchanfrage basierend auf Excel-Daten"""
    st.markdown("---")
    st.subheader(f"🔍 Excel-Details für: '{search_term}'")
    
    # Statistiken
    papers = term_info.get("papers", 0)
    new_papers = term_info.get("new_papers", 0)
    last_update = term_info.get("last_update", "Unbekannt")
    
    col_detail1, col_detail2, col_detail3 = st.columns(3)
    
    with col_detail1:
        st.metric("📄 Gesamt Papers", papers)
    
    with col_detail2:
        st.metric("🆕 Neue Papers (letzter Run)", new_papers)
    
    with col_detail3:
        st.metric("📅 Letztes Update", last_update[:16].replace('T', ' ') if last_update != "Unbekannt" else "Unbekannt")
    
    # Aktionen
    col_action1, col_action2 = st.columns(2)
    
    with col_action1:
        if st.button("🔄 Suche wiederholen", key=f"repeat_{search_term}"):
            execute_excel_integrated_search(search_term, 100, "Letzte 2 Jahre", False, False)
    
    with col_action2:
        if st.button("📊 Excel-Sheet anzeigen", key=f"show_excel_{search_term}"):
            show_excel_sheet_content(search_term)

def show_excel_sheet_content(search_term: str):
    """Zeigt Inhalt eines Excel-Sheets"""
    template_path = st.session_state["excel_template"]["file_path"]
    sheet_name = generate_sheet_name(search_term)
    
    try:
        if os.path.exists(template_path):
            xl_file = pd.ExcelFile(template_path)
            
            if sheet_name in xl_file.sheet_names:
                df = pd.read_excel(template_path, sheet_name=sheet_name)
                
                st.markdown("---")
                st.subheader(f"📊 Excel-Sheet: '{search_term}'")
                
                # Statistiken
                col_stat1, col_stat2, col_stat3 = st.columns(3)
                
                with col_stat1:
                    st.metric("📄 Gesamt Papers", len(df))
                
                with col_stat2:
                    new_papers = len(df[df.get("Status") == "NEU"]) if "Status" in df.columns else 0
                    st.metric("🆕 Neue Papers", new_papers)
                
                with col_stat3:
                    with_doi = len(df[df.get("DOI", "").astype(str).str.len() > 0]) if "DOI" in df.columns else 0
                    st.metric("🔗 Mit DOI", with_doi)
                
                # Anzeige der Papers
                st.write("**📋 Papers (erste 10):**")
                display_papers = df.head(10)
                
                for idx, (_, paper) in enumerate(display_papers.iterrows(), 1):
                    title = paper.get("Titel", "Unbekannt")
                    authors = paper.get("Autoren", "Unbekannt")
                    journal = paper.get("Journal", "Unbekannt")
                    year = paper.get("Jahr", "")
                    
                    with st.expander(f"📄 **{idx}.** {title[:60]}... ({year})"):
                        st.write(f"**👥 Autoren:** {authors}")
                        st.write(f"**📚 Journal:** {journal}")
                        if paper.get("URL"):
                            st.markdown(f"🔗 [**PubMed ansehen**]({paper.get('URL')})")
                
                if len(df) > 10:
                    st.info(f"... und {len(df) - 10} weitere Papers")
            else:
                st.error(f"❌ Sheet '{sheet_name}' nicht gefunden!")
        else:
            st.error("❌ Excel-Datei nicht gefunden!")
    
    except Exception as e:
        st.error(f"❌ Fehler beim Anzeigen des Sheet-Inhalts: {str(e)}")

def show_excel_template_management():
    """Excel-Template Management mit Excel-Integration"""
    st.subheader("📋 Excel-Template Management & Integration")
    
    template_path = st.session_state["excel_template"]["file_path"]
    excel_stats = get_search_statistics_from_excel()
    
    # Template Status
    if os.path.exists(template_path):
        file_size = os.path.getsize(template_path)
        file_date = datetime.datetime.fromtimestamp(os.path.getmtime(template_path))
        
        st.success(f"✅ **Master Excel-Template aktiv:** {template_path}")
        st.info(f"📊 **Größe:** {file_size:,} bytes | **Letzte Änderung:** {file_date.strftime('%d.%m.%Y %H:%M')}")
        
        # Excel-Statistiken anzeigen
        excel_stats = get_search_statistics_from_excel()
        if excel_stats:
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            with col_stat1:
                st.metric("📊 Excel-Sheets", excel_stats.get("total_sheets", 0))
            with col_stat2:
                st.metric("📄 Gesamt Papers", excel_stats.get("total_papers", 0))
            with col_stat3:
                st.metric("🔍 Durchsuchungen", excel_stats.get("total_searches", 0))
    else:
        st.error("❌ Master Excel-Template nicht gefunden!")
        if st.button("🔧 Template neu erstellen"):
            create_master_excel_template()
            st.rerun()
    
    # Excel-Aktionen
    col_excel1, col_excel2, col_excel3 = st.columns(3)
    
    with col_excel1:
        if st.button("📥 **Excel herunterladen**"):
            offer_excel_download()
    
    with col_excel2:
        if st.button("📊 **Sheet-Übersicht anzeigen**"):
            show_excel_sheets_overview()
    
    with col_excel3:
        if st.button("🔄 **Template zurücksetzen**"):
            if st.button("✅ Bestätigen", key="confirm_reset"):
                reset_excel_template()

def get_search_statistics_from_excel() -> Dict:
    """Holt Statistiken aus der Excel-Datei"""
    wb = load_master_workbook()
    if not wb:
        return {}
    
    stats = {
        "total_sheets": len([s for s in wb.sheetnames if not s.startswith(("📊", "ℹ️"))]),
        "total_searches": 0,
        "total_papers": 0,
        "search_terms": []
    }
    
    if "📊_Overview" in wb.sheetnames:
        overview_sheet = wb["📊_Overview"]
        
        for row in overview_sheet.iter_rows(min_row=2):
            if row[1].value:  # Suchbegriff existiert
                stats["total_searches"] += 1
                stats["total_papers"] += row[2].value or 0
                stats["search_terms"].append({
                    "term": row[1].value,
                    "papers": row[2].value or 0,
                    "last_update": row[3].value,
                    "new_papers": row[4].value or 0
                })
    
    return stats

def load_master_workbook():
    """Lädt das Master Excel Workbook"""
    excel_path = st.session_state["excel_template"]["file_path"]
    try:
        return openpyxl.load_workbook(excel_path)
    except Exception as e:
        st.error(f"❌ Excel-Datei konnte nicht geladen werden: {str(e)}")
        return None

def show_excel_sheets_overview():
    """Zeigt Übersicht aller Excel-Sheets"""
    st.markdown("---")
    st.subheader("📊 Excel-Sheets Übersicht")
    
    excel_stats = get_search_statistics_from_excel()
    
    if excel_stats.get("search_terms"):
        # Erstelle DataFrame für bessere Darstellung
        df_overview = pd.DataFrame(excel_stats["search_terms"])
        df_overview.columns = ["Suchbegriff", "Papers", "Letztes Update", "Neue Papers"]
        
        # Sortiere nach letztem Update
        df_overview = df_overview.sort_values("Letztes Update", ascending=False)
        
        st.dataframe(df_overview, use_container_width=True)
        
        # Zusammenfassung
        total_papers = df_overview["Papers"].sum()
        total_new = df_overview["Neue Papers"].sum()
        
        col_sum1, col_sum2, col_sum3 = st.columns(3)
        with col_sum1:
            st.metric("📊 Gesamt Sheets", len(df_overview))
        with col_sum2:
            st.metric("📄 Gesamt Papers", total_papers)
        with col_sum3:
            st.metric("🆕 Neue Papers", total_new)
    else:
        st.info("📭 Noch keine Excel-Sheets vorhanden.")

def offer_excel_download():
    """Bietet Master Excel-Datei zum Download an"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    if os.path.exists(template_path):
        try:
            with open(template_path, 'rb') as f:
                excel_data = f.read()
            
            filename = f"PaperSearch_Master_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            st.download_button(
                label="📥 **Master Excel-Datei herunterladen**",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Lädt die komplette Excel-Datei mit allen Sheets herunter"
            )
        
        except Exception as e:
            st.error(f"❌ Fehler beim Bereitstellen der Excel-Datei: {str(e)}")

def reset_excel_template():
    """Setzt Excel-Template zurück"""
    template_path = st.session_state["excel_template"]["file_path"]
    
    try:
        if os.path.exists(template_path):
            # Backup erstellen
            backup_path = f"{template_path}.backup_{int(time.time())}"
            os.rename(template_path, backup_path)
            st.info(f"📁 Backup erstellt: {backup_path}")
        
        create_master_excel_template()
        st.success("✅ Excel-Template zurückgesetzt!")
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Fehler beim Zurücksetzen: {str(e)}")

# =============== EMAIL-KONFIGURATION MIT MEHREREN EMPFÄNGERN ===============

def show_email_config():
    """Email-Konfiguration mit mehreren Empfängern"""
    st.subheader("📧 Email-Konfiguration (Mehrere Empfänger)")
    
    settings = st.session_state.get("email_settings", {})
    
    # Email-Setup Hilfe
    with st.expander("📖 Email-Setup Hilfe & Mehrere Empfänger"):
        st.info("""
        **Für Gmail (empfohlen):**
        1. ✅ 2-Faktor-Authentifizierung aktivieren
        2. ✅ App-Passwort erstellen (nicht normales Passwort!)
        3. ✅ SMTP: smtp.gmail.com, Port: 587, TLS: An
        
        **Mehrere Empfänger:**
        • Trennen Sie mehrere Email-Adressen mit Kommas
        • Beispiel: user1@gmail.com, user2@outlook.com, user3@company.de
        • Whitespaces werden automatisch entfernt
        
        **Für Outlook/Hotmail:**
        - SMTP: smtp-mail.outlook.com, Port: 587
        """)
    
    with st.form("email_config_form"):
        st.subheader("📬 Grundeinstellungen")
        
        col1, col2 = st.columns(2)
        
        with col1:
            sender_email = st.text_input(
                "Absender Email *", 
                value=settings.get("sender_email", ""),
                placeholder="absender@gmail.com"
            )
            
            smtp_server = st.text_input(
                "SMTP Server *",
                value=settings.get("smtp_server", "smtp.gmail.com")
            )
            
            auto_notifications = st.checkbox(
                "Automatische Benachrichtigungen", 
                value=settings.get("auto_notifications", True)
            )
        
        with col2:
            smtp_port = st.number_input(
                "SMTP Port *",
                value=settings.get("smtp_port", 587),
                min_value=1,
                max_value=65535
            )
            
            min_papers = st.number_input(
                "Min. Papers für Benachrichtigung", 
                value=settings.get("min_papers", 1),
                min_value=1,
                max_value=100
            )
            
            use_tls = st.checkbox(
                "TLS Verschlüsselung verwenden (empfohlen)",
                value=settings.get("use_tls", True)
            )
        
        # MEHRERE EMPFÄNGER - Text Area
        recipient_emails = st.text_area(
            "📧 Empfänger Email-Adressen * (mehrere mit Komma trennen)",
            value=settings.get("recipient_emails", ""),
            placeholder="empfaenger1@example.com, empfaenger2@gmail.com, empfaenger3@company.de",
            help="Mehrere Email-Adressen mit Komma trennen. Beispiel: user1@gmail.com, user2@outlook.com",
            height=80
        )
        
        sender_password = st.text_input(
            "Email Passwort / App-Passwort *",
            value=settings.get("sender_password", ""),
            type="password",
            help="Für Gmail: App-spezifisches Passwort verwenden!"
        )
        
        # Email-Vorlagen
        st.subheader("📝 Email-Vorlagen")
        
        subject_template = st.text_input(
            "Betreff-Vorlage",
            value=settings.get("subject_template", "🔬 {count} neue Papers für '{search_term}'"),
            help="Platzhalter: {count}, {search_term}, {frequency}"
        )
        
        message_template = st.text_area(
            "Nachricht-Vorlage",
            value=settings.get("message_template", """📧 Automatische Paper-Benachrichtigung

📅 Datum: {date}
🔍 Suchbegriff: '{search_term}'
📊 Neue Papers: {count}

📋 Neue Papers:
{new_papers_list}

📎 Excel-Datei: {excel_file}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""),
            height=200,
            help="Platzhalter: {date}, {search_term}, {count}, {frequency}, {new_papers_list}, {excel_file}"
        )
        
        if st.form_submit_button("💾 **Email-Einstellungen speichern**", type="primary"):
            # Validiere Email-Adressen
            recipient_list = parse_recipient_emails(recipient_emails)
            
            if not recipient_list:
                st.error("❌ Mindestens eine gültige Empfänger-Email erforderlich!")
            else:
                new_settings = {
                    "sender_email": sender_email,
                    "recipient_emails": recipient_emails,
                    "smtp_server": smtp_server,
                    "smtp_port": smtp_port,
                    "sender_password": sender_password,
                    "use_tls": use_tls,
                    "auto_notifications": auto_notifications,
                    "min_papers": min_papers,
                    "subject_template": subject_template,
                    "message_template": message_template,
                    "parsed_recipients": recipient_list  # Store parsed list
                }
                
                st.session_state["email_settings"] = new_settings
                st.success(f"✅ Email-Einstellungen gespeichert! **{len(recipient_list)} Empfänger** konfiguriert:")
                for i, email in enumerate(recipient_list, 1):
                    st.write(f"   {i}. 📧 {email}")
    
    # Zeige konfigurierte Empfänger
    if settings.get("recipient_emails"):
        recipient_list = parse_recipient_emails(settings.get("recipient_emails", ""))
        if recipient_list:
            st.info(f"📧 **Aktuell konfigurierte Empfänger ({len(recipient_list)}):**")
            cols = st.columns(min(len(recipient_list), 3))
            for i, email in enumerate(recipient_list):
                with cols[i % 3]:
                    st.write(f"✅ {email}")
    
    # Test-Email
    st.markdown("---")
    st.subheader("🧪 Email-System testen")
    
    col_test1, col_test2 = st.columns(2)
    
    with col_test1:
        if st.button("📧 **Test-Email an alle Empfänger senden**", type="primary"):
            send_test_email_multiple()
    
    with col_test2:
        if st.button("📊 **Email-Status prüfen**"):
            check_email_status_multiple()

def parse_recipient_emails(email_string: str) -> List[str]:
    """Parst Email-String und gibt Liste gültiger Emails zurück"""
    if not email_string:
        return []
    
    # Split by comma and clean
    emails = [email.strip() for email in email_string.split(",")]
    
    # Basic email validation
    valid_emails = []
    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    for email in emails:
        if email and email_pattern.match(email):
            valid_emails.append(email)
    
    return valid_emails

def send_real_email_multiple(to_emails: List[str], subject: str, message: str, attachment_path: str = None) -> tuple:
    """Sendet echte Email über SMTP an mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    
    sender_email = settings.get("sender_email", "")
    sender_password = settings.get("sender_password", "")
    smtp_server = settings.get("smtp_server", "smtp.gmail.com")
    smtp_port = settings.get("smtp_port", 587)
    use_tls = settings.get("use_tls", True)
    
    if not all([sender_email, sender_password]):
        return False, "❌ Email-Konfiguration unvollständig (Absender/Passwort)"
    
    if not to_emails:
        return False, "❌ Keine Empfänger-Emails konfiguriert"
    
    try:
        # SMTP Server Setup
        server = smtplib.SMTP(smtp_server, smtp_port)
        
        if use_tls:
            context = ssl.create_default_context()
            server.starttls(context=context)
        
        server.login(sender_email, sender_password)
        
        successful_sends = 0
        failed_sends = []
        
        # Send to each recipient
        for recipient in to_emails:
            try:
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = recipient
                msg['Subject'] = subject
                
                msg.attach(MIMEText(message, 'plain', 'utf-8'))
                
                # Add attachment if provided
                if attachment_path and os.path.exists(attachment_path):
                    with open(attachment_path, "rb") as attachment:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename= {os.path.basename(attachment_path)}'
                        )
                        msg.attach(part)
                
                server.send_message(msg)
                successful_sends += 1
                
            except Exception as e:
                failed_sends.append(f"{recipient}: {str(e)}")
        
        server.quit()
        
        if successful_sends == len(to_emails):
            return True, f"✅ Email erfolgreich an alle {successful_sends} Empfänger gesendet"
        elif successful_sends > 0:
            return True, f"⚠️ Email an {successful_sends}/{len(to_emails)} Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        else:
            return False, f"❌ Email an keinen Empfänger gesendet. Fehler: {'; '.join(failed_sends)}"
        
    except smtplib.SMTPAuthenticationError:
        return False, "❌ SMTP-Authentifizierung fehlgeschlagen - Prüfen Sie Email/Passwort"
    except smtplib.SMTPServerDisconnected:
        return False, "❌ SMTP-Server-Verbindung unterbrochen"
    except Exception as e:
        return False, f"❌ Email-Fehler: {str(e)}"

def send_test_email_multiple():
    """Sendet Test-Email an alle konfigurierten Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    if not settings.get("sender_email") or not recipient_emails:
        st.error("❌ Email-Konfiguration unvollständig!")
        return
    
    subject = "🧪 Test-Email vom Paper-Suche System (Mehrere Empfänger)"
    message = f"""Dies ist eine Test-Email vom Paper-Suche System mit Unterstützung für mehrere Empfänger.

📅 Gesendet am: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}
📧 Von: {settings.get('sender_email')}
📧 An: {len(recipient_emails)} Empfänger

Empfänger-Liste:
{chr(10).join([f"• {email}" for email in recipient_emails])}

✅ Wenn Sie diese Email erhalten, funktioniert das Email-System korrekt!

System-Informationen:
• SMTP Server: {settings.get('smtp_server')}
• Port: {settings.get('smtp_port')}
• TLS: {'Aktiviert' if settings.get('use_tls') else 'Deaktiviert'}
• Empfänger: {len(recipient_emails)}

Mit freundlichen Grüßen,
Ihr Paper-Suche System"""
    
    success, status_message = send_real_email_multiple(
        recipient_emails, 
        subject, 
        message
    )
    
    if success:
        st.success(f"✅ **Test-Email erfolgreich gesendet!** {status_message}")
        st.balloons()
    else:
        st.error(f"❌ **Test-Email fehlgeschlagen:** {status_message}")

def check_email_status_multiple():
    """Prüft Email-Status mit mehreren Empfängern"""
    settings = st.session_state.get("email_settings", {})
    
    st.write("**📊 Email-Konfiguration Status:**")
    
    # Prüfe Konfiguration
    sender_ok = bool(settings.get("sender_email"))
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    recipients_ok = len(recipient_emails) > 0
    password_ok = bool(settings.get("sender_password"))
    
    st.write(f"📧 Absender Email: {'✅' if sender_ok else '❌'} {settings.get('sender_email', 'Nicht konfiguriert')}")
    st.write(f"📧 Empfänger Emails: {'✅' if recipients_ok else '❌'} {len(recipient_emails)} konfiguriert")
    
    if recipients_ok:
        with st.expander(f"📧 Empfänger-Liste ({len(recipient_emails)})"):
            for i, email in enumerate(recipient_emails, 1):
                st.write(f"   {i}. {email}")
    
    st.write(f"🔑 Passwort: {'✅' if password_ok else '❌'} {'Konfiguriert' if password_ok else 'Nicht konfiguriert'}")
    st.write(f"🔒 SMTP Server: {settings.get('smtp_server', 'smtp.gmail.com')}:{settings.get('smtp_port', 587)}")
    st.write(f"🔐 TLS: {'✅ Aktiviert' if settings.get('use_tls', True) else '❌ Deaktiviert'}")
    
    # Gesamtstatus
    if sender_ok and recipients_ok and password_ok:
        st.success(f"✅ **Email-System vollständig konfiguriert für {len(recipient_emails)} Empfänger!**")
    else:
        st.error("❌ **Email-System nicht vollständig konfiguriert!**")

def is_email_configured() -> bool:
    """Prüft Email-Konfiguration für mehrere Empfänger"""
    settings = st.session_state.get("email_settings", {})
    recipient_emails = parse_recipient_emails(settings.get("recipient_emails", ""))
    
    return (bool(settings.get("sender_email")) and 
            len(recipient_emails) > 0 and
            bool(settings.get("sender_password")))

# =============== WEITERE FUNKTIONEN ===============

def show_detailed_statistics():
    """Detaillierte Statistiken mit Excel-Integration"""
    st.subheader("📈 Detaillierte Statistiken")
    
    status = st.session_state["system_status"]
    search_history = st.session_state.get("search_history", [])
    email_history = st.session_state.get("email_history", [])
    excel_stats = get_search_statistics_from_excel()
    
    # Hauptstatistiken
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("🔍 Suchen (Session)", status["total_searches"])
        st.metric("🔍 Suchen (Excel)", excel_stats.get("total_searches", 0))
    
    with col2:
        st.metric("📄 Papers (Session)", status["total_papers"])
        st.metric("📄 Papers (Excel)", excel_stats.get("total_papers", 0))
    
    with col3:
        st.metric("📧 Gesendete Emails", len(email_history))
        recipient_count = len(parse_recipient_emails(st.session_state.get("email_settings", {}).get("recipient_emails", "")))
        st.metric("📧 Email-Empfänger", recipient_count)
    
    with col4:
        st.metric("📊 Excel Sheets", excel_stats.get("total_sheets", 0))
        auto_searches = len(st.session_state.get("automatic_searches", {}))
        st.metric("🤖 Auto-Suchen", auto_searches)
    
    # Email-Statistiken detailliert
    if email_history:
        st.markdown("---")
        st.subheader("📧 Email-Statistiken")
        
        successful_emails = len([e for e in email_history if e.get("success", False)])
        success_rate = (successful_emails / len(email_history)) * 100
        
        col_email1, col_email2, col_email3 = st.columns(3)
        
        with col_email1:
            st.metric("📧 Gesamt Emails", len(email_history))
        with col_email2:
            st.metric("✅ Erfolgreich", successful_emails)
        with col_email3:
            st.metric("📊 Erfolgsrate", f"{success_rate:.1f}%")
        
        # Letzte Emails
        st.write("**📧 Letzte Email-Aktivitäten:**")
        recent_emails = sorted(email_history, key=lambda x: x.get("timestamp", ""), reverse=True)[:5]
        
        for i, email in enumerate(recent_emails, 1):
            timestamp = email.get("timestamp", "")[:16].replace('T', ' ')
            email_type = email.get("type", "Unbekannt")
            success_icon = "✅" if email.get("success", False) else "❌"
            recipient_count = email.get("recipient_count", 1)
            
            st.write(f"{i}. {success_icon} **{email_type}** ({recipient_count} Empfänger) - {timestamp}")
    
    # Excel-basierte Suchstatistiken
    if excel_stats.get("search_terms"):
        st.markdown("---")
        st.subheader("📊 Excel-basierte Suchstatistiken")
        
        # Top Suchbegriffe nach Papers
        top_searches = sorted(excel_stats["search_terms"], key=lambda x: x.get("papers", 0), reverse=True)[:5]
        
        st.write("**🔝 Top 5 Suchbegriffe (nach Papers):**")
        for i, search in enumerate(top_searches, 1):
            term = search.get("term", "Unbekannt")
            papers = search.get("papers", 0)
            new_papers = search.get("new_papers", 0)
            
            st.write(f"{i}. **{term}** - {papers} Papers ({new_papers} neue)")

def show_system_settings():
    """System-Einstellungen mit Excel-Integration"""
    st.subheader("⚙️ System-Einstellungen")
    
    # Excel-Template Einstellungen
    template_settings = st.session_state["excel_template"]
    
    with st.form("system_settings_form"):
        st.write("**📊 Excel-Template Einstellungen:**")
        
        col_set1, col_set2 = st.columns(2)
        
        with col_set1:
            auto_create_sheets = st.checkbox(
                "Automatische Sheet-Erstellung",
                value=template_settings.get("auto_create_sheets", True),
                help="Erstellt automatisch neue Sheets für jeden Suchbegriff"
            )
            
            max_sheets = st.number_input(
                "Maximale Anzahl Sheets",
                value=template_settings.get("max_sheets", 50),
                min_value=10,
                max_value=100,
                help="Maximale Anzahl von Sheets in der Excel-Datei"
            )
        
        with col_set2:
            sheet_naming = st.selectbox(
                "Sheet-Benennung",
                ["topic_based", "date_based", "custom"],
                index=0,
                help="Art der Sheet-Benennung"
            )
        
        if st.form_submit_button("💾 Einstellungen speichern"):
            st.session_state["excel_template"].update({
                "auto_create_sheets": auto_create_sheets,
                "max_sheets": max_sheets,
                "sheet_naming": sheet_naming
            })
            st.success("✅ System-Einstellungen gespeichert!")
    
    # System-Informationen
    st.markdown("---")
    st.subheader("ℹ️ System-Informationen")
    
    col_info1, col_info2 = st.columns(2)
    
    with col_info1:
        st.write("**📁 Pfade:**")
        st.code(f"Excel-Template: {st.session_state['excel_template']['file_path']}")
        st.code(f"Arbeitsverzeichnis: {os.getcwd()}")
    
    with col_info2:
        st.write("**🔧 Konfiguration:**")
        st.write(f"Auto-Sheets: {'✅' if template_settings.get('auto_create_sheets') else '❌'}")
        st.write(f"Max-Sheets: {template_settings.get('max_sheets', 50)}")
        st.write(f"Email-System: {'✅' if is_email_configured() else '❌'}")
    
    # System zurücksetzen
    st.markdown("---")
    st.subheader("🔄 System zurücksetzen")
    
    col_reset1, col_reset2, col_reset3 = st.columns(3)
    
    with col_reset1:
        if st.button("🗑️ Such-Historie löschen"):
            st.session_state["search_history"] = []
            st.success("Such-Historie gelöscht!")
    
    with col_reset2:
        if st.button("📧 Email-Historie löschen"):
            st.session_state["email_history"] = []
            st.success("Email-Historie gelöscht!")
    
    with col_reset3:
        if st.button("🤖 Auto-Suchen löschen"):
            st.session_state["automatic_searches"] = {}
            st.success("Automatische Suchen gelöscht!")

# =============== HILFSFUNKTIONEN ===============

def build_advanced_search_query(query: str, date_filter: str) -> str:
    """Erweiterte Suchanfrage mit Filtern"""
    query_parts = [query]
    
    if date_filter != "Alle":
        current_year = datetime.datetime.now().year
        if date_filter == "Letztes Jahr":
            query_parts.append(f"AND {current_year-1}:{current_year}[dp]")
        elif date_filter == "Letzte 2 Jahre":
            query_parts.append(f"AND {current_year-2}:{current_year}[dp]")
        elif date_filter == "Letzte 5 Jahre":
            query_parts.append(f"AND {current_year-5}:{current_year}[dp]")
        elif date_filter == "Letzte 10 Jahre":
            query_parts.append(f"AND {current_year-10}:{current_year}[dp]")
    
    return " ".join(query_parts)

def load_previous_search_results(query: str) -> List[Dict]:
    """Lädt vorherige Suchergebnisse aus Excel"""
    template_path = st.session_state["excel_template"]["file_path"]
    sheet_name = generate_sheet_name(query)
    
    if not os.path.exists(template_path):
        return []
    
    try:
        xl_file = pd.ExcelFile(template_path)
        if sheet_name not in xl_file.sheet_names:
            return []
        
        df = pd.read_excel(template_path, sheet_name=sheet_name)
        
        previous_papers = []
        for _, row in df.iterrows():
            if pd.notna(row.get("PMID")):
                paper = {
                    "PMID": str(row.get("PMID", "")),
                    "Title": str(row.get("Titel", "")),
                    "Authors": str(row.get("Autoren", "")),
                    "Journal": str(row.get("Journal", "")),
                    "Year": str(row.get("Jahr", ""))
                }
                previous_papers.append(paper)
        
        return previous_papers
        
    except Exception:
        return []

def identify_new_papers(current_papers: List[Dict], previous_papers: List[Dict]) -> List[Dict]:
    """Identifiziert neue Papers"""
    previous_pmids = set(paper.get("PMID", "") for paper in previous_papers if paper.get("PMID"))
    
    new_papers = []
    for paper in current_papers:
        current_pmid = paper.get("PMID", "")
        if current_pmid and current_pmid not in previous_pmids:
            paper["Is_New"] = True
            new_papers.append(paper)
        else:
            paper["Is_New"] = False
    
    return new_papers

def save_search_to_history(query: str, papers: List[Dict], new_papers: List[Dict]):
    """Speichert Suche in Historie"""
    search_entry = {
        "search_term": query,
        "timestamp": datetime.datetime.now().isoformat(),
        "paper_count": len(papers),
        "new_papers": len(new_papers),
        "date": datetime.datetime.now().date().isoformat()
    }
    
    st.session_state["search_history"].append(search_entry)

def update_system_status(paper_count: int):
    """Aktualisiert System-Status"""
    status = st.session_state["system_status"]
    status["total_searches"] += 1
    status["total_papers"] += paper_count
    status["last_search"] = datetime.datetime.now().isoformat()
    
    # Zähle Excel-Sheets
    excel_stats = get_search_statistics_from_excel()
    status["excel_sheets"] = excel_stats.get("total_sheets", 0)

def display_search_results(papers: List[Dict], new_papers: List[Dict], query: str, is_repeat: bool):
    """Zeigt Suchergebnisse an"""
    st.subheader(f"📋 Ergebnisse für: '{query}'")
    
    # Statistiken
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("📄 Gesamt Papers", len(papers))
    
    with col2:
        st.metric("🆕 Neue Papers", len(new_papers))
    
    with col3:
        with_abstract = len([p for p in papers if p.get("Abstract", "") != "Kein Abstract verfügbar"])
        st.metric("📝 Mit Abstract", with_abstract)
    
    with col4:
        with_doi = len([p for p in papers if p.get("DOI", "")])
        st.metric("🔗 Mit DOI", with_doi)
    
    # Papers anzeigen (erste 5)
    display_papers = papers[:5]
    
    for idx, paper in enumerate(display_papers, 1):
        is_new = paper.get("Is_New", False)
        status_icon = "🆕" if is_new else "📄"
        
        title = paper.get("Title", "Unbekannt")
        header = f"{status_icon} **{idx}.** {title[:60]}..."
        
        with st.expander(header):
            st.write(f"**📄 Titel:** {title}")
            st.write(f"**👥 Autoren:** {paper.get('Authors', 'n/a')}")
            st.write(f"**📚 Journal:** {paper.get('Journal', 'n/a')} ({paper.get('Year', 'n/a')})")
            st.write(f"**🆔 PMID:** {paper.get('PMID', 'n/a')}")
            
            if paper.get('DOI'):
                st.write(f"**🔗 DOI:** {paper.get('DOI')}")
            
            if paper.get('URL'):
                st.markdown(f"🔗 [**PubMed ansehen**]({paper.get('URL')})")
    
    if len(papers) > 5:
        st.info(f"... und {len(papers) - 5} weitere Papers (siehe Excel-Datei)")

def should_send_email(paper_count: int) -> bool:
    """Prüft ob Email gesendet werden soll"""
    settings = st.session_state.get("email_settings", {})
    return (settings.get("auto_notifications", False) and
            paper_count >= settings.get("min_papers", 1) and
            is_email_configured())
def show_automatic_search_system():
    """Automatisches Such-System (vereinfacht ohne schedule)"""
    st.subheader("🤖 Automatisches Such-System")
    
    st.info("""
    💡 **Hinweis:** Diese Version funktioniert ohne das 'schedule' Paket.
    Automatische Suchen können manuell ausgeführt werden.
    """)
    
    # Automatische Suchen verwalten
    auto_searches = st.session_state.get("automatic_searches", {})
    
    # Neue automatische Suche erstellen
    with st.expander("➕ Neue automatische Suche erstellen"):
        with st.form("create_auto_search"):
            col_auto1, col_auto2 = st.columns(2)
            
            with col_auto1:
                auto_search_term = st.text_input(
                    "Suchbegriff",
                    placeholder="z.B. 'diabetes genetics', 'COVID-19 treatment'"
                )
                
                auto_frequency = st.selectbox(
                    "Häufigkeit",
                    ["Täglich", "Wöchentlich", "Monatlich"],
                    index=1
                )
            
            with col_auto2:
                auto_max_papers = st.number_input(
                    "Max. Papers pro Suche",
                    min_value=10,
                    max_value=200,
                    value=50
                )
                
                auto_email_enabled = st.checkbox(
                    "Email-Benachrichtigungen",
                    value=True
                )
            
            if st.form_submit_button("🤖 **Automatische Suche erstellen**", type="primary"):
                if auto_search_term:
                    create_automatic_search(auto_search_term, auto_frequency, auto_max_papers, auto_email_enabled)
                else:
                    st.error("❌ Suchbegriff ist erforderlich!")
    
    # Bestehende automatische Suchen anzeigen
    if auto_searches:
        st.markdown("---")
        st.subheader(f"🤖 Konfigurierte automatische Suchen ({len(auto_searches)})")
        
        for search_id, search_config in auto_searches.items():
            search_term = search_config.get("search_term", "Unbekannt")
            frequency = search_config.get("frequency", "Unbekannt")
            last_run = search_config.get("last_run", "Nie")
            
            with st.expander(f"🤖 **{search_term}** ({frequency})"):
                col_config1, col_config2 = st.columns([2, 1])
                
                with col_config1:
                    st.write(f"**🔍 Suchbegriff:** {search_term}")
                    st.write(f"**⏰ Häufigkeit:** {frequency}")
                    st.write(f"**📧 Email:** {'✅' if search_config.get('email_enabled', False) else '❌'}")
                    st.write(f"**🕒 Letzter Lauf:** {last_run[:19] if last_run != 'Nie' else 'Nie'}")
                
                with col_config2:
                    if st.button("▶️ Jetzt ausführen", key=f"run_auto_{search_id}"):
                        run_automatic_search_simple(search_config)
                    
                    if st.button("🗑️ Löschen", key=f"delete_auto_{search_id}"):
                        delete_automatic_search(search_id)
                        st.rerun()
        
        # Globale Aktionen
        st.markdown("---")
        col_global1, col_global2 = st.columns(2)
        
        with col_global1:
            if st.button("▶️ **Alle automatischen Suchen ausführen**", type="primary"):
                run_all_automatic_searches_simple()
        
        with col_global2:
            if st.button("🔄 **Status aktualisieren**"):
                st.rerun()
    
    else:
        st.info("📭 Noch keine automatischen Suchen konfiguriert.")

def create_automatic_search(search_term: str, frequency: str, max_papers: int, email_enabled: bool):
    """Erstellt neue automatische Suche"""
    search_id = f"auto_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    search_config = {
        "search_id": search_id,
        "search_term": search_term,
        "frequency": frequency,
        "max_papers": max_papers,
        "email_enabled": email_enabled,
        "created_date": datetime.datetime.now().isoformat(),
        "last_run": "Nie",
        "total_runs": 0
    }
    
    st.session_state["automatic_searches"][search_id] = search_config
    
    st.success(f"✅ **Automatische Suche erstellt:** '{search_term}' ({frequency})")

def run_automatic_search_simple(search_config: Dict):
    """Führt eine automatische Suche aus (vereinfacht)"""
    search_term = search_config.get("search_term", "")
    max_papers = search_config.get("max_papers", 50)
    email_enabled = search_config.get("email_enabled", False)
    
    st.info(f"🤖 Führe automatische Suche aus: '{search_term}'")
    
    try:
        # Führe Excel-integrierte Suche durch
        execute_excel_integrated_search(search_term, max_papers, "Letzte 2 Jahre", email_enabled, False)
        
        # Update Konfiguration
        search_config["last_run"] = datetime.datetime.now().isoformat()
        search_config["total_runs"] = search_config.get("total_runs", 0) + 1
        
        st.success(f"✅ Automatische Suche für '{search_term}' abgeschlossen!")
        
    except Exception as e:
        st.error(f"❌ Fehler bei automatischer Suche '{search_term}': {str(e)}")

def run_all_automatic_searches_simple():
    """Führt alle automatischen Suchen aus (vereinfacht)"""
    auto_searches = st.session_state.get("automatic_searches", {})
    
    if not auto_searches:
        st.info("📭 Keine automatischen Suchen konfiguriert.")
        return
    
    st.info(f"🤖 Führe {len(auto_searches)} automatische Suchen aus...")
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    total_new_papers = 0
    
    for i, search_config in enumerate(auto_searches.values()):
        search_term = search_config.get("search_term", "")
        try:
            status_text.text(f"🔍 Automatische Suche {i+1}/{len(auto_searches)}: '{search_term}'...")
            
            # Führe Suche durch
            current_papers = perform_comprehensive_pubmed_search(search_term, search_config.get("max_papers", 50))
            
            if current_papers:
                # Füge neue Papers zur Excel hinzu
                added_count, new_papers = add_new_papers_to_excel(search_term, current_papers)
                
                if added_count > 0:
                    # Sende Email wenn konfiguriert
                    if search_config.get("email_enabled", False) and should_send_email(added_count):
                        send_excel_integrated_email_multiple(search_term, new_papers, len(current_papers), added_count)
                    
                    total_new_papers += added_count
                    st.write(f"✅ **{search_term}:** {added_count} neue Papers")
                else:
                    st.write(f"ℹ️ **{search_term}:** Keine neuen Papers")
                
                # Update Konfiguration
                search_config["last_run"] = datetime.datetime.now().isoformat()
                search_config["total_runs"] = search_config.get("total_runs", 0) + 1
            else:
                st.write(f"⚠️ **{search_term}:** Keine Papers gefunden")
            
            # Progress update
            progress_bar.progress((i + 1) / len(auto_searches))
            time.sleep(1)  # Rate limiting
            
        except Exception as e:
            st.error(f"❌ Fehler bei automatischer Suche '{search_term}': {str(e)}")
            continue
    
    progress_bar.empty()
    status_text.empty()
    
    # Ergebnis
    if total_new_papers > 0:
        st.success(f"🎉 **Alle automatischen Suchen abgeschlossen!** {total_new_papers} neue Papers insgesamt gefunden!")
        st.balloons()
    else:
        st.info("ℹ️ **Alle automatischen Suchen abgeschlossen.** Keine neuen Papers gefunden.")

def delete_automatic_search(search_id: str):
    """Löscht automatische Suche"""
    if search_id in st.session_state["automatic_searches"]:
        search_term = st.session_state["automatic_searches"][search_id].get("search_term", "Unbekannt")
        del st.session_state["automatic_searches"][search_id]
        st.success(f"🗑️ Automatische Suche '{search_term}' gelöscht!")

if __name__ == "__main__":
    module_email()

