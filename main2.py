import streamlit as st
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import re
import datetime
import sys
import concurrent.futures
import os
import PyPDF2
import openai
import time
import json
import pdfplumber
import io

from typing import Dict, Any, Optional
from dotenv import load_dotenv
from PIL import Image
from scholarly import scholarly

# Neu: Excel / openpyxl-Import
import openpyxl

# Neuer Import für die Übersetzung mit google_trans_new
from google_trans_new import google_translator

# ------------------------------------------------------------------
# Umgebungsvariablen laden (für OPENAI_API_KEY, falls vorhanden)
# ------------------------------------------------------------------
load_dotenv()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# ------------------------------------------------------------------
# Streamlit-Konfiguration
# ------------------------------------------------------------------
st.set_page_config(page_title="Streamlit Multi-Modul Demo", layout="wide")

# ------------------------------------------------------------------
# Login-Funktionalität
# ------------------------------------------------------------------
def login():
    st.title("Login")
    user_input = st.text_input("Username")
    pass_input = st.text_input("Password", type="password")
    if st.button("Login"):
        if (
            user_input == st.secrets["login"]["username"]
            and pass_input == st.secrets["login"]["password"]
        ):
            st.session_state["logged_in"] = True
        else:
            st.error("Login failed. Please check your credentials!")

if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login()
    st.stop()

# ------------------------------------------------------------------
# 1) Gemeinsame Funktionen & Klassen
# ------------------------------------------------------------------
def clean_html_except_br(text):
    """Removes all HTML tags except <br>."""
    cleaned_text = re.sub(r'</?(?!br\b)[^>]*>', '', text)
    return cleaned_text

def translate_text_openai(text, source_language, target_language, api_key):
    """Übersetzt Text über OpenAI-ChatCompletion."""
    import openai
    openai.api_key = api_key
    prompt_system = (
        f"You are a translation engine from {source_language} to {target_language} for a biotech company called Novogenia "
        f"that focuses on lifestyle and health genetics and health analyses. The outputs you provide will be used directly as "
        f"the translated text blocks. Please translate as accurately as possible in the context of health and lifestyle reporting. "
        f"If there is no appropriate translation, the output should be 'TBD'. Keep the TAGS and do not add additional punctuation."
    )
    prompt_user = f"Translate the following text from {source_language} to {target_language}:\n'{text}'"
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": prompt_system},
                {"role": "user", "content": prompt_user}
            ],
            temperature=0
        )
        translation = response.choices[0].message.content.strip()
        # Removes leading/trailing quotes
        if translation and translation[0] in ["'", '"', "‘", "„"]:
            translation = translation[1:]
            if translation and translation[-1] in ["'", '"']:
                translation = translation[:-1]
        translation = clean_html_except_br(translation)
        return translation
    except Exception as e:
        st.warning("Translation error: " + str(e))
        return text

class CoreAPI:
    def __init__(self, api_key):
        self.base_url = "https://api.core.ac.uk/v3/"
        self.headers = {"Authorization": f"Bearer {api_key}"}

    def search_publications(self, query, filters=None, sort=None, limit=100):
        endpoint = "search/works"
        params = {"q": query, "limit": limit}
        if filters:
            filter_expressions = []
            for key, value in filters.items():
                filter_expressions.append(f"{key}:{value}")
            params["filter"] = ",".join(filter_expressions)
        if sort:
            params["sort"] = sort
        r = requests.get(
            self.base_url + endpoint,
            headers=self.headers,
            params=params,
            timeout=15
        )
        r.raise_for_status()
        return r.json()

def check_core_aggregate_connection(api_key="LmAMxdYnK6SDJsPRQCpGgwN7f5yTUBHF", timeout=15):
    """Check if CORE aggregator is reachable."""
    try:
        core = CoreAPI(api_key)
        result = core.search_publications("test", limit=1)
        return "results" in result
    except Exception:
        return False

def search_core_aggregate(query, api_key="LmAMxdYnK6SDJsPRQCpGgwN7f5yTUBHF"):
    """Simple search in CORE aggregator."""
    if not api_key:
        return []
    try:
        core = CoreAPI(api_key)
        raw = core.search_publications(query, limit=100)
        out = []
        results = raw.get("results", [])
        for item in results:
            title = item.get("title", "n/a")
            year = str(item.get("yearPublished", "n/a"))
            journal = item.get("publisher", "n/a")
            out.append({
                "PMID": "n/a",
                "Title": title,
                "Year": year,
                "Journal": journal
            })
        return out
    except Exception as e:
        st.error(f"CORE search error: {e}")
        return []

# ------------------------------------------------------------------
# 2) PubMed - Einfacher Check + Search
# ------------------------------------------------------------------
def check_pubmed_connection(timeout=10):
    """Quick connection test to PubMed."""
    test_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {"db": "pubmed", "term": "test", "retmode": "json"}
    try:
        r = requests.get(test_url, params=params, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return "esearchresult" in data
    except Exception:
        return False

def search_pubmed_simple(query):
    """Short search (title/journal/year) in PubMed."""
    esearch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {"db": "pubmed", "term": query, "retmode": "json", "retmax": 100}
    out = []
    try:
        r = requests.get(esearch_url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        idlist = data.get("esearchresult", {}).get("idlist", [])
        if not idlist:
            return out
        esummary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
        sum_params = {"db": "pubmed", "id": ",".join(idlist), "retmode": "json"}
        r2 = requests.get(esummary_url, params=sum_params, timeout=10)
        r2.raise_for_status()
        summary_data = r2.json().get("result", {})
        for pmid in idlist:
            info = summary_data.get(pmid, {})
            title = info.get("title", "n/a")
            pubdate = info.get("pubdate", "")
            year = pubdate[:4] if pubdate else "n/a"
            journal = info.get("fulljournalname", "n/a")
            out.append({
                "PMID": pmid,
                "Title": title,
                "Year": year,
                "Journal": journal
            })
        return out
    except Exception as e:
        st.error(f"Error searching PubMed: {e}")
        return []

def fetch_pubmed_abstract(pmid):
    """Fetches abstract via efetch for a given PubMed ID."""
    url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params = {"db": "pubmed", "id": pmid, "retmode": "xml"}
    try:
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        root = ET.fromstring(r.content)
        abs_text = []
        for elem in root.findall(".//AbstractText"):
            if elem.text:
                abs_text.append(elem.text.strip())
        if abs_text:
            return "\n".join(abs_text)
        else:
            return "(No abstract available)"
    except Exception as e:
        return f"(Error: {e})"

def fetch_pubmed_doi_and_link(pmid: str) -> (str, str):
    """
    Attempts to retrieve the DOI and PubMed link for a given PMID via E-Summary/E-Fetch.
    Returns (doi, pubmed_link). If no DOI is found, returns ("n/a", link).
    """
    if not pmid or pmid == "n/a":
        return ("n/a", "")
    
    # PubMed link
    link = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
    
    # 1) esummary
    summary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
    params_sum = {"db": "pubmed", "id": pmid, "retmode": "json"}
    try:
        rs = requests.get(summary_url, params=params_sum, timeout=8)
        rs.raise_for_status()
        data = rs.json()
        result_obj = data.get("result", {}).get(pmid, {})
        eloc = result_obj.get("elocationid", "")
        if eloc and eloc.startswith("doi:"):
            doi_ = eloc.split("doi:", 1)[1].strip()
            if doi_:
                return (doi_, link)
    except Exception:
        pass
    
    # 2) efetch
    efetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params_efetch = {"db": "pubmed", "id": pmid, "retmode": "xml"}
    try:
        r_ef = requests.get(efetch_url, params=params_efetch, timeout=8)
        r_ef.raise_for_status()
        root = ET.fromstring(r_ef.content)
        doi_found = "n/a"
        for aid in root.findall(".//ArticleId"):
            id_type = aid.attrib.get("IdType", "")
            if id_type.lower() == "doi":
                doi_found = aid.text.strip() if aid.text else "n/a"
                break
        return (doi_found, link)
    except Exception:
        return ("n/a", link)

# ------------------------------------------------------------------
# 3) Europe PMC Check + Search
# ------------------------------------------------------------------
def check_europe_pmc_connection(timeout=10):
    """Check if Europe PMC is reachable."""
    test_url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    params = {"query": "test", "format": "json", "pageSize": 100}
    try:
        r = requests.get(test_url, params=params, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return "resultList" in data and "result" in data["resultList"]
    except Exception:
        return False

def search_europe_pmc_simple(query):
    """Simple search in Europe PMC."""
    url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
    params = {
        "query": query,
        "format": "json",
        "pageSize": 100,
        "resultType": "core"
    }
    out = []
    try:
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        if "resultList" not in data or "result" not in data["resultList"]:
            return out
        results = data["resultList"]["result"]
        for item in results:
            pmid = item.get("pmid", "n/a")
            title = item.get("title", "n/a")
            year = str(item.get("pubYear", "n/a"))
            journal = item.get("journalTitle", "n/a")
            out.append({
                "PMID": pmid if pmid else "n/a",
                "Title": title,
                "Year": year,
                "Journal": journal
            })
        return out
    except Exception as e:
        st.error(f"Europe PMC search error: {e}")
        return []

# ------------------------------------------------------------------
# 4) OpenAlex API
# ------------------------------------------------------------------
BASE_URL = "https://api.openalex.org"

def fetch_openalex_data(entity_type, entity_id=None, params=None):
    url = f"{BASE_URL}/{entity_type}"
    if entity_id:
        url += f"/{entity_id}"
    if params is None:
        params = {}
    params["mailto"] = "your_email@example.com"
    response = requests.get(url, params=params)
    if response.status_code == 200:
        return response.json()
    else:
        st.error(f"Fehler: {response.status_code} - {response.text}")
        return None

def search_openalex_simple(query):
    """Short version: fetches raw data, checks if anything is returned."""
    search_params = {"search": query}
    return fetch_openalex_data("works", params=search_params)

# ------------------------------------------------------------------
# 5) Google Scholar
# ------------------------------------------------------------------
class GoogleScholarSearch:
    def __init__(self):
        self.all_results = []
    def search_google_scholar(self, base_query):
        try:
            search_results = scholarly.search_pubs(base_query)
            for _ in range(5):
                result = next(search_results)
                title = result['bib'].get('title', "n/a")
                authors = result['bib'].get('author', "n/a")
                year = result['bib'].get('pub_year', "n/a")
                url_article = result.get('url_scholarbib', "n/a")
                abstract_text = result['bib'].get('abstract', "")
                self.all_results.append({
                    "Source": "Google Scholar",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": "n/a",
                    "Year": year,
                    "PMID": "n/a",
                    "DOI": "n/a",
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            st.error(f"Fehler bei der Google Scholar-Suche: {e}")

# ------------------------------------------------------------------
# 6) Semantic Scholar
# ------------------------------------------------------------------
def check_semantic_scholar_connection(timeout=10):
    """Connection test to Semantic Scholar."""
    try:
        url = "https://api.semanticscholar.org/graph/v1/paper/search"
        params = {"query": "test", "limit": 1, "fields": "title"}
        headers = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers, params=params, timeout=timeout)
        response.raise_for_status()
        return response.status_code == 200
    except Exception:
        return False

class SemanticScholarSearch:
    def __init__(self):
        self.all_results = []
    def search_semantic_scholar(self, base_query):
        try:
            url = "https://api.semanticscholar.org/graph/v1/paper/search"
            headers = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
            params = {"query": base_query, "limit": 5, "fields": "title,authors,year,abstract,doi,paperId"}
            response = requests.get(url, headers=headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            for paper in data.get("data", []):
                title = paper.get("title", "n/a")
                authors = ", ".join([author.get("name", "") for author in paper.get("authors", [])])
                year = paper.get("year", "n/a")
                doi = paper.get("doi", "n/a")
                paper_id = paper.get("paperId", "")
                abstract_text = paper.get("abstract", "")
                url_article = f"https://www.semanticscholar.org/paper/{paper_id}" if paper_id else "n/a"
                self.all_results.append({
                    "Source": "Semantic Scholar",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": "n/a",
                    "Year": year,
                    "PMID": "n/a",
                    "DOI": "n/a",
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            st.error(f"Semantic Scholar: {e}")

# ------------------------------------------------------------------
# 7) Excel Online Search - Placeholder
# ------------------------------------------------------------------

# ------------------------------------------------------------------
# 8) Weitere Module + Seiten
# ------------------------------------------------------------------
def module_paperqa2():
    st.subheader("PaperQA2 Module")
    st.write("This is the PaperQA2 module. You can add more settings and functions here.")
    question = st.text_input("Please enter your question:")
    if st.button("Submit question"):
        st.write("Answer: This is a dummy answer to the question:", question)

def page_home():
    st.title("Welcome to the Main Menu")
    st.write("Choose a module in the sidebar to proceed.")
    st.image("Bild1.jpg", caption="Willkommen!", use_container_width=False, width=600)

def page_codewords_pubmed():
    st.title("Codewords & PubMed Settings")
    from modules.codewords_pubmed import module_codewords_pubmed
    module_codewords_pubmed()
    if st.button("Back to Main Menu"):
        st.session_state["current_page"] = "Home"

def page_paper_selection():
    st.title("Paper Selection Settings")
    st.write("Define how you want to pick or exclude certain papers. (Dummy placeholder...)")
    if st.button("Back to Main Menu"):
        st.session_state["current_page"] = "Home"

def page_analysis():
    st.title("Analysis & Evaluation Settings")
    st.write("Set up your analysis parameters, thresholds, etc. (Dummy placeholder...)")
    if st.button("Back to Main Menu"):
        st.session_state["current_page"] = "Home"

def page_extended_topics():
    st.title("Extended Topics")
    st.write("Access advanced or extended topics for further research. (Dummy placeholder...)")
    if st.button("Back to Main Menu"):
        st.session_state["current_page"] = "Home"

def page_paperqa2():
    st.title("PaperQA2")
    module_paperqa2()
    if st.button("Back to Main Menu"):
        st.session_state["current_page"] = "Home"

def page_excel_online_search():
    st.title("Excel Online Search")
    # Placeholder, or import existing code if needed

def page_online_api_filter():
    st.title("Online-API_Filter (Combined)")
    st.write("Here, you can combine API selection and filtering in one step.")
    from modules.online_api_filter import module_online_api_filter
    module_online_api_filter()
    if st.button("Back to Main Menu"):
        st.session_state["current_page"] = "Home"

# ------------------------------------------------------------------
# Important Classes for Analysis
# ------------------------------------------------------------------
class PaperAnalyzer:
    def __init__(self, model="gpt-3.5-turbo"):
        self.model = model
    
    def extract_text_from_pdf(self, pdf_file):
        """Extracts raw text via PyPDF2."""
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        return text
    
    def analyze_with_openai(self, text, prompt_template, api_key):
        """Helper function to call OpenAI via ChatCompletion."""
        import openai
        openai.api_key = api_key
        if len(text) > 15000:
            text = text[:15000] + "..."
        prompt = prompt_template.format(text=text)
        response = openai.ChatCompletion.create(
            model=self.model,
            messages=[
                {"role": "system", "content": "You are an expert in scientific paper analysis."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=1500
        )
        return response.choices[0].message.content
    
    def summarize(self, text, api_key):
        """Creates a summary in German."""
        prompt = (
            "Erstelle eine strukturierte Zusammenfassung des folgenden wissenschaftlichen Papers. "
            "Gliedere sie in mindestens vier klar getrennte Abschnitte (z.B. 1. Hintergrund, 2. Methodik, 3. Ergebnisse, 4. Schlussfolgerungen). "
            "Verwende maximal 500 Wörter:\n\n{text}"
        )
        return self.analyze_with_openai(text, prompt, api_key)
    
    def extract_key_findings(self, text, api_key):
        """Extract the 5 most important findings."""
        prompt = (
            "Extrahiere die 5 wichtigsten Erkenntnisse aus diesem wissenschaftlichen Paper. "
            "Liste sie mit Bulletpoints auf:\n\n{text}"
        )
        return self.analyze_with_openai(text, prompt, api_key)
    
    def identify_methods(self, text, api_key):
        """Identify methods and techniques used in the paper."""
        prompt = (
            "Identifiziere und beschreibe die im Paper verwendeten Methoden und Techniken. "
            "Gib zu jeder Methode eine kurze Erklärung:\n\n{text}"
        )
        return self.analyze_with_openai(text, prompt, api_key)
    
    def evaluate_relevance(self, text, topic, api_key):
        """Rates relevance to the topic on a scale of 1-10."""
        prompt = (
            f"Bewerte die Relevanz dieses Papers für das Thema '{topic}' auf einer Skala von 1-10. "
            f"Begründe deine Bewertung:\n\n{{text}}"
        )
        return self.analyze_with_openai(text, prompt, api_key)

class AlleleFrequencyFinder:
    """Class for retrieving and displaying allele frequencies from various sources (Ensembl primarily)."""
    def __init__(self):
        self.ensembl_server = "https://rest.ensembl.org"
        self.max_retries = 3
        self.retry_delay = 2  # seconds between retries

    def get_allele_frequencies(self, rs_id: str, retry_count: int = 0) -> Optional[Dict[str, Any]]:
        """Fetches allele frequencies from Ensembl."""
        if not rs_id.startswith("rs"):
            rs_id = f"rs{rs_id}"
        endpoint = f"/variation/human/{rs_id}?pops=1"
        url = f"{self.ensembl_server}{endpoint}"
        try:
            response = requests.get(url, headers={"Content-Type": "application/json"}, timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError:
            if response.status_code == 500 and retry_count < self.max_retries:
                time.sleep(self.retry_delay)
                return self.get_allele_frequencies(rs_id, retry_count + 1)
            elif response.status_code == 404:
                return None
            else:
                return None
        except requests.exceptions.RequestException:
            if retry_count < self.max_retries:
                time.sleep(self.retry_delay)
                return self.get_allele_frequencies(rs_id, retry_count + 1)
            return None
    
    def try_alternative_source(self, rs_id: str) -> Optional[Dict[str, Any]]:
        return None
    
    def build_freq_info_text(self, data: Dict[str, Any]) -> str:
        """Generates a short text about allele frequencies in ENGLISH for the Excel."""
        if not data:
            return "No data from Ensembl"
        maf = data.get("MAF", None)
        pops = data.get("populations", [])
        out = []
        out.append(f"MAF={maf}" if maf else "MAF=n/a")
        if pops:
            max_pop = 2
            for i, pop in enumerate(pops):
                if i >= max_pop:
                    break
                pop_name = pop.get('population', 'N/A')
                allele = pop.get('allele', 'N/A')
                freq = pop.get('frequency', 'N/A')
                out.append(f"{pop_name}:{allele}={freq}")
        else:
            out.append("No population data found.")
        return " | ".join(out)

def split_summary(summary_text):
    """Attempts to split 'Ergebnisse' and 'Schlussfolgerungen' from a German summary."""
    pattern = re.compile(
        r'(Ergebnisse(?:\:|\s*\n)|Resultate(?:\:|\s*\n))(?P<results>.*?)(Schlussfolgerungen(?:\:|\s*\n)|Fazit(?:\:|\s*\n))(?P<conclusion>.*)',
        re.IGNORECASE | re.DOTALL
    )
    match = pattern.search(summary_text)
    if match:
        ergebnisse = match.group('results').strip()
        schlussfolgerungen = match.group('conclusion').strip()
        return ergebnisse, schlussfolgerungen
    else:
        return summary_text, ""

def parse_cohort_info(summary_text: str) -> dict:
    """Parses rough info about the cohort (number of patients, origin, etc.) from a German summary."""
    info = {"study_size": "", "origin": ""}
    pattern_both = re.compile(
        r"(\d+)\s*Patient(?:en)?(?:[^\d]+)(\d+)\s*gesunde\s*Kontroll(?:personen)?",
        re.IGNORECASE
    )
    m_both = pattern_both.search(summary_text)
    if m_both:
        p_count = m_both.group(1)
        c_count = m_both.group(2)
        info["study_size"] = f"{p_count} Patienten / {c_count} Kontrollpersonen"
    else:
        pattern_single_p = re.compile(r"(\d+)\s*Patient(?:en)?", re.IGNORECASE)
        m_single_p = pattern_single_p.search(summary_text)
        if m_single_p:
            info["study_size"] = f"{m_single_p.group(1)} Patienten"
    pattern_origin = re.compile(r"in\s*der\s+(\S+)\s+Bevölkerung", re.IGNORECASE)
    m_orig = pattern_origin.search(summary_text)
    if m_orig:
        info["origin"] = m_orig.group(1).strip()
    return info

# ------------------------------------------------------------------
# (Bereits vorhanden) fetch_pubmed_doi_and_link
# ------------------------------------------------------------------

# ------------------------------------------------------------------
# Function for ChatGPT-based scoring search
# ------------------------------------------------------------------
def chatgpt_online_search_with_genes(papers, codewords, genes, top_k=100):
    openai.api_key = st.secrets.get("OPENAI_API_KEY", "")
    if not openai.api_key:
        st.error("No 'OPENAI_API_KEY' in st.secrets.")
        return []
    scored_results = []
    total = len(papers)
    progress = st.progress(0)
    status_text = st.empty()
    genes_str = ", ".join(genes) if genes else ""
    for idx, paper in enumerate(papers, start=1):
        current_title = paper.get("Title", "n/a")
        status_text.text(f"Processing Paper {idx}/{total}: {current_title}")
        progress.progress(idx / total)
        title = paper.get("Title", "n/a")
        abstract = paper.get("Abstract", "n/a")
        prompt = f"""
Codewords: {codewords}
Genes: {genes_str}

Paper:
Title: {title}
Abstract: {abstract}

Give me a number from 0 to 100 (relevance), taking both codewords and genes into account.
"""
        try:
            resp = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=20,
                temperature=0
            )
            raw_text = resp.choices[0].message.content.strip()
            match = re.search(r'(\d+)', raw_text)
            if match:
                score = int(match.group(1))
            else:
                score = 0
        except Exception as e:
            st.error(f"ChatGPT error during scoring: {e}")
            score = 0
        new_item = dict(paper)
        new_item["Relevance"] = score
        scored_results.append(new_item)
    status_text.empty()
    progress.empty()
    scored_results.sort(key=lambda x: x["Relevance"], reverse=True)
    return scored_results[:top_k]

# ------------------------------------------------------------------
# Function for analyzing commonalities & contradictions
# ------------------------------------------------------------------
def analyze_papers_for_commonalities_and_contradictions(pdf_texts: Dict[str, str], api_key: str, model: str, method_choice: str = "Standard"):
    import openai
    openai.api_key = api_key

    # 1) Extract claims per paper
    all_claims = {}
    for fname, txt in pdf_texts.items():
        prompt_claims = f"""
Lies den folgenden Ausschnitt eines wissenschaftlichen Papers (maximal 2000 Tokens).
Extrahiere bitte die wichtigsten 3-5 "Aussagen" (Claims), die das Paper aufstellt.
Nutze als Ausgabe ein kompaktes JSON-Format, z.B:
[
  {{"claim": "Aussage 1"}},
  {{"claim": "Aussage 2"}}
]
Text: {txt[:6000]}
"""
        try:
            resp_claims = openai.ChatCompletion.create(
                model=model,
                messages=[{"role": "user", "content": prompt_claims}],
                temperature=0.3,
                max_tokens=700
            )
            raw = resp_claims.choices[0].message.content.strip()
            try:
                claims_list = json.loads(raw)
            except Exception:
                claims_list = [{"claim": raw}]
            if not isinstance(claims_list, list):
                claims_list = [claims_list]
            all_claims[fname] = claims_list
        except Exception as e:
            st.error(f"Error extracting claims in {fname}: {e}")
            all_claims[fname] = []

    merged_claims = []
    for fname, cllist in all_claims.items():
        for cobj in cllist:
            ctext = cobj.get("claim", "(leer)")
            merged_claims.append({
                "paper": fname,
                "claim": ctext
            })
    big_input_str = json.dumps(merged_claims, ensure_ascii=False, indent=2)

    # 2) Identify commonalities + contradictions
    if method_choice == "ContraCrow":
        final_prompt = f"""
Nutze die ContraCrow-Methodik, um die folgenden Claims (Aussagen) aus mehreren wissenschaftlichen PDF-Papers zu analysieren. 
Die ContraCrow-Methodik fokussiert sich darauf, systematisch Gemeinsamkeiten und klare Widersprüche zu identifizieren.
Bitte identifiziere:
1) Die zentralen gemeinsamen Aussagen, die in den Papers auftreten.
2) Klare Widersprüche zwischen den Aussagen der verschiedenen Papers.

Antworte ausschließlich in folgendem JSON-Format (ohne zusätzliche Erklärungen):
{{
  "commonalities": [
    "Gemeinsamkeit 1",
    "Gemeinsamkeit 2"
  ],
  "contradictions": [
    {{"paperA": "...", "claimA": "...", "paperB": "...", "claimB": "...", "reason": "..." }},
    ...
  ]
}}

Hier die Claims:
{big_input_str}
"""
    else:
        final_prompt = f"""
Hier sind verschiedene Claims (Aussagen) aus mehreren wissenschaftlichen PDF-Papers im JSON-Format.
Bitte identifiziere:
1) Gemeinsamkeiten zwischen den Papers (Wo überschneiden oder ergänzen sich die Aussagen?)
2) Mögliche Widersprüche (Welche Aussagen widersprechen sich klar?)

Antworte NUR in folgendem JSON-Format (ohne weitere Erklärungen):
{{
  "commonalities": [
    "Gemeinsamkeit 1",
    "Gemeinsamkeit 2"
  ],
  "contradictions": [
    {{"paperA": "...", "claimA": "...", "paperB": "...", "claimB": "...", "reason": "..."}}
  ]
}}

Hier die Claims:
{big_input_str}
"""

    try:
        resp_final = openai.ChatCompletion.create(
            model=model,
            messages=[{"role": "user", "content": final_prompt}],
            temperature=0.0,
            max_tokens=1500
        )
        raw2 = resp_final.choices[0].message.content.strip()
        return raw2
    except Exception as e:
        return f"Fehler bei Gemeinsamkeiten/Widersprüche: {e}"

# ------------------------------------------------------------------
# Page: Analyze Paper (inkl. PaperQA Multi-Paper Analyzer)
# ------------------------------------------------------------------
def page_analyze_paper():
    st.title("Analyze Paper - Integrated")
    
    if "api_key" not in st.session_state:
        st.session_state["api_key"] = OPENAI_API_KEY or ""
    
    st.sidebar.header("Settings - PaperAnalyzer")
    new_key_value = st.sidebar.text_input("OpenAI API Key", type="password", value=st.session_state["api_key"])
    st.session_state["api_key"] = new_key_value
    
    model = st.sidebar.selectbox(
        "OpenAI Model",
        ["gpt-3.5-turbo", "gpt-3.5-turbo-16k", "gpt-4"],
        index=0
    )
    
    analysis_method = st.sidebar.selectbox("Analysis Method (Commonalities & Contradictions)", ["Standard GPT", "ContraCrow"])
    
    compare_mode = st.sidebar.checkbox("Compare all papers together (exclude outliers)?")
    
    theme_mode = st.sidebar.radio("Determine main theme", ["Manually", "GPT"])
    
    action = st.sidebar.radio(
        "Analysis Type",
        ["Zusammenfassung", "Wichtigste Erkenntnisse", "Methoden & Techniken", "Relevanz-Bewertung", "Tabellen & Grafiken"],
        index=0
    )
    
    user_defined_theme = ""
    if theme_mode == "Manually":
        user_defined_theme = st.sidebar.text_input("Manual main theme (if Compare-Mode is active)")
    
    topic = st.sidebar.text_input("Topic for relevance rating (if relevant)")
    output_lang = st.sidebar.selectbox("Output Language", ["Deutsch", "Englisch", "Portugiesisch", "Serbisch"], index=0)
    
    uploaded_files = st.file_uploader("Upload PDF files", type="pdf", accept_multiple_files=True)
    analyzer = PaperAnalyzer(model=model)
    api_key = st.session_state["api_key"]
    
    if "paper_texts" not in st.session_state:
        st.session_state["paper_texts"] = {}
    
    if "relevant_papers_compare" not in st.session_state:
        st.session_state["relevant_papers_compare"] = None
    if "theme_compare" not in st.session_state:
        st.session_state["theme_compare"] = ""
    
    def do_outlier_logic(paper_map: dict) -> (list, str):
        """Determines which papers are thematically relevant and possibly a shared main theme."""
        if theme_mode == "Manually":
            main_theme = user_defined_theme.strip()
            if not main_theme:
                st.error("Please provide a manual main theme!")
                return ([], "")
            snippet_list = []
            for name, txt_data in paper_map.items():
                snippet = txt_data[:700].replace("\n", " ")
                snippet_list.append(f'{{"filename": "{name}", "snippet": "{snippet}"}}')
            big_snippet = ",\n".join(snippet_list)
            big_input = f"""
Der Nutzer hat folgendes Hauptthema definiert: '{main_theme}'.

Hier sind mehrere Paper in JSON-Form. Entscheide pro Paper, ob es zu diesem Thema passt oder nicht.
Gib mir am Ende ein JSON-Format zurück:

{{
  "theme": "you repeat the user-defined theme",
  "papers": [
    {{"filename": "...", "relevant": true/false, "reason": "Short reason"}}
  ]
}}

Only return the JSON, no extra explanation.

[{big_snippet}]
"""
            try:
                openai.api_key = api_key
                scope_resp = openai.ChatCompletion.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "You check paper snippets for relevance to the user theme."},
                        {"role": "user", "content": big_input}
                    ],
                    temperature=0.0,
                    max_tokens=1800
                )
                scope_decision = scope_resp.choices[0].message.content
            except Exception as e1:
                st.error(f"GPT error in Compare-Mode (Manual): {e1}")
                return ([], "")
            st.markdown("#### GPT-Output (Outlier-Check / Manual):")
            st.code(scope_decision, language="json")
            json_str = scope_decision.strip()
            if json_str.startswith("```"):
                json_str = re.sub(r"```[\w]*\n?", "", json_str)
                json_str = re.sub(r"\n?```", "", json_str)
            try:
                data_parsed = json.loads(json_str)
                papers_info = data_parsed.get("papers", [])
            except Exception as parse_e:
                st.error(f"Error parsing JSON: {parse_e}")
                return ([], "")
            st.write(f"**Main theme (Manual)**: {main_theme}")
            relevant_papers_local = []
            st.write("**Paper classification**:")
            for p in papers_info:
                fname = p.get("filename", "?")
                rel = p.get("relevant", False)
                reason = p.get("reason", "(none)")
                if rel:
                    relevant_papers_local.append(fname)
                    st.success(f"{fname} => relevant. Reason: {reason}")
                else:
                    st.warning(f"{fname} => NOT relevant. Reason: {reason}")
            return (relevant_papers_local, main_theme)
        else:
            snippet_list = []
            for name, txt_data in paper_map.items():
                snippet = txt_data[:700].replace("\n", " ")
                snippet_list.append(f'{{"filename": "{name}", "snippet": "{snippet}"}}')
            big_snippet = ",\n".join(snippet_list)
            big_input = f"""
Hier sind mehrere Paper in JSON-Form. Bitte ermittele das gemeinsame Hauptthema.
Dann antworte mir in folgendem JSON-Format: 
{{
  "main_theme": "Brief description of the shared topic",
  "papers": [
    {{"filename":"...","relevant":true/false,"reason":"Short reason"}}
  ]
}}

Only output this JSON, no further explanation:

[{big_snippet}]
"""
            try:
                openai.api_key = api_key
                scope_resp = openai.ChatCompletion.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "You are an assistant that thematically filters papers."},
                        {"role": "user", "content": big_input}
                    ],
                    temperature=0.0,
                    max_tokens=1800
                )
                scope_decision = scope_resp.choices[0].message.content
            except Exception as e1:
                st.error(f"GPT error in Compare-Mode: {e1}")
                return ([], "")
            st.markdown("#### GPT-Output (Outlier-Check / GPT):")
            st.code(scope_decision, language="json")
            json_str = scope_decision.strip()
            if json_str.startswith("```"):
                json_str = re.sub(r"```[\w]*\n?", "", json_str)
                json_str = re.sub(r"\n?```", "", json_str)
            try:
                data_parsed = json.loads(json_str)
                main_theme = data_parsed.get("main_theme", "No theme extracted.")
                papers_info = data_parsed.get("papers", [])
            except Exception as parse_e:
                st.error(f"Error parsing JSON: {parse_e}")
                return ([], "")
            st.write(f"**Main theme (GPT)**: {main_theme}")
            relevant_papers_local = []
            st.write("**Paper classification**:")
            for p in papers_info:
                fname = p.get("filename", "?")
                rel = p.get("relevant", False)
                reason = p.get("reason", "(none)")
                if rel:
                    relevant_papers_local.append(fname)
                    st.success(f"{fname} => relevant. Reason: {reason}")
                else:
                    st.warning(f"{fname} => NOT relevant. Reason: {reason}")
            return (relevant_papers_local, main_theme)

    if uploaded_files and api_key:
        if compare_mode:
            st.write("### Compare-Mode: Exclude Outlier Papers")
            if st.button("Start Compare-Analysis"):
                paper_map = {}
                for fpdf in uploaded_files:
                    txt = analyzer.extract_text_from_pdf(fpdf)
                    if txt.strip():
                        paper_map[fpdf.name] = txt
                    else:
                        st.warning(f"No text extracted from {fpdf.name} (skipped).")
                if not paper_map:
                    st.error("No usable papers.")
                    return
                relevant_papers, discovered_theme = do_outlier_logic(paper_map)
                st.session_state["relevant_papers_compare"] = relevant_papers
                st.session_state["theme_compare"] = discovered_theme
                if not relevant_papers:
                    st.error("No relevant papers remain after outlier-check.")
                    return
                combined_text = ""
                for rp in relevant_papers:
                    combined_text += f"\n=== {rp} ===\n{paper_map[rp]}"
                if action == "Tabellen & Grafiken":
                    final_result = "Tables & figures not implemented in combined Compare-Mode."
                else:
                    if action == "Zusammenfassung":
                        final_result = analyzer.summarize(combined_text, api_key)
                    elif action == "Wichtigste Erkenntnisse":
                        final_result = analyzer.extract_key_findings(combined_text, api_key)
                    elif action == "Methoden & Techniken":
                        final_result = analyzer.identify_methods(combined_text, api_key)
                    elif action == "Relevanz-Bewertung":
                        if not topic:
                            st.error("Please provide a topic!")
                            return
                        final_result = analyzer.evaluate_relevance(combined_text, topic, api_key)
                    else:
                        final_result = "(No analysis type selected.)"
                if output_lang != "Deutsch":
                    lang_map = {"Englisch": "English", "Portugiesisch": "Portuguese", "Serbisch": "Serbian"}
                    target_lang = lang_map.get(output_lang, "English")
                    final_result = translate_text_openai(final_result, "German", target_lang, api_key)
                st.subheader("Result of Compare-Mode:")
                st.write(final_result)
        else:
            st.write("### Single or Multi-Mode (no outlier-check)")
            
            pdf_options = ["(All)"] + [f"{i+1}) {f.name}" for i, f in enumerate(uploaded_files)]
            selected_pdf = st.selectbox("Select a PDF for single analysis or '(All)'", pdf_options)
            
            col_analysis, col_contradiction = st.columns(2)

            with col_analysis:
                if st.button("Start Analysis (Single-Mode)"):
                    if selected_pdf == "(All)":
                        files_to_process = uploaded_files
                    else:
                        idx = pdf_options.index(selected_pdf) - 1
                        if idx < 0:
                            st.warning("No file selected.")
                            return
                        files_to_process = [uploaded_files[idx]]
                    final_result_text = []
                    for fpdf in files_to_process:
                        text_data = ""
                        if action != "Tabellen & Grafiken":
                            with st.spinner(f"Extracting text from {fpdf.name}..."):
                                text_data = analyzer.extract_text_from_pdf(fpdf)
                                if not text_data.strip():
                                    st.error(f"No text extracted from {fpdf.name}.")
                                    continue
                                st.success(f"Text extracted from {fpdf.name}!")
                                st.session_state["paper_text"] = text_data[:15000]
                        result = ""
                        if action == "Zusammenfassung":
                            with st.spinner(f"Creating summary for {fpdf.name}..."):
                                result = analyzer.summarize(text_data, api_key)
                        elif action == "Wichtigste Erkenntnisse":
                            with st.spinner(f"Extracting key findings from {fpdf.name}..."):
                                result = analyzer.extract_key_findings(text_data, api_key)
                        elif action == "Methoden & Techniken":
                            with st.spinner(f"Identifying methods for {fpdf.name}..."):
                                result = analyzer.identify_methods(text_data, api_key)
                        elif action == "Relevanz-Bewertung":
                            if not topic:
                                st.error("Please provide a topic!")
                                return
                            with st.spinner(f"Evaluating relevance of {fpdf.name}..."):
                                result = analyzer.evaluate_relevance(text_data, topic, api_key)
                        elif action == "Tabellen & Grafiken":
                            with st.spinner(f"Searching for tables/figures in {fpdf.name}..."):
                                all_tables_text = []
                                try:
                                    with pdfplumber.open(fpdf) as pdf_:
                                        for page_number, page in enumerate(pdf_.pages, start=1):
                                            st.markdown(f"### Page {page_number} in {fpdf.name}")
                                            tables = page.extract_tables()
                                            if tables:
                                                st.markdown("**Tables on this page**")
                                                for table_idx, table_data in enumerate(tables, start=1):
                                                    if not table_data:
                                                        st.write("Empty table detected.")
                                                        continue
                                                    first_row = table_data[0]
                                                    data_rows = table_data[1:]
                                                    if not data_rows:
                                                        st.write("Only a header present.")
                                                        data_rows = table_data
                                                        first_row = [f"Col_{i}" for i in range(len(data_rows[0]))]
                                                    import pandas as pd
                                                    new_header = []
                                                    used_cols = {}
                                                    for col in first_row:
                                                        col_str = col if col else "N/A"
                                                        if col_str not in used_cols:
                                                            used_cols[col_str] = 1
                                                            new_header.append(col_str)
                                                        else:
                                                            used_cols[col_str] += 1
                                                            new_header.append(f"{col_str}.{used_cols[col_str]}")
                                                    if any(len(row) != len(new_header) for row in data_rows):
                                                        st.write("Warning: inconsistent column count.")
                                                        df = pd.DataFrame(table_data)
                                                    else:
                                                        df = pd.DataFrame(data_rows, columns=new_header)
                                                    st.write(f"**Table {table_idx}** in {fpdf.name}:")
                                                    st.dataframe(df)
                                                    table_str = df.to_csv(index=False)
                                                    all_tables_text.append(f"Page {page_number} - Table {table_idx}\n{table_str}\n")
                                            else:
                                                st.write("No tables here.")
                                            images = page.images
                                            if images:
                                                st.markdown("**Images/Figures on this page**")
                                                for img_index, img_dict in enumerate(images, start=1):
                                                    xref = img_dict.get("xref")
                                                    if xref is not None:
                                                        extracted_img = page.extract_image(xref)
                                                        if extracted_img:
                                                            image_data = extracted_img["image"]
                                                            image = Image.open(io.BytesIO(image_data))
                                                            st.write(f"**Image {img_index}** in {fpdf.name}:")
                                                            st.image(image, use_column_width=True)
                                                        else:
                                                            st.write(f"Image {img_index} could not be extracted.")
                                            else:
                                                st.write("No images here.")
                                    # Simple fulltext search for "Table"
                                    st.markdown(f"### Fulltext-Search 'Table' in {fpdf.name}")
                                    try:
                                        text_all_pages = ""
                                        with pdfplumber.open(fpdf) as pdf2:
                                            for pg in pdf2.pages:
                                                t_ = pg.extract_text() or ""
                                                text_all_pages += t_ + "\n"
                                        lines = text_all_pages.splitlines()
                                        matches = [ln for ln in lines if "Table" in ln]
                                        if matches:
                                            st.write("Lines containing 'Table':")
                                            for ln in matches:
                                                st.write(f"- {ln}")
                                        else:
                                            st.write("No mention of 'Table'.")
                                    except Exception as e2:
                                        st.warning(f"Error in fulltext-search 'Table': {e2}")
                                    if len(all_tables_text) > 0:
                                        combined_tables_text = "\n".join(all_tables_text)
                                        if len(combined_tables_text) > 14000:
                                            combined_tables_text = combined_tables_text[:14000] + "..."
                                        gpt_prompt = (
                                            "Please analyze the following tables from a scientific PDF. "
                                            "Summarize the key insights and (if possible) give a short interpretation "
                                            "in the context of lifestyle and health genetics:\n\n"
                                            f"{combined_tables_text}"
                                        )
                                        try:
                                            openai.api_key = api_key
                                            gpt_resp = openai.ChatCompletion.create(
                                                model=model,
                                                messages=[
                                                    {"role": "system", "content": "You are an expert in PDF table analysis."},
                                                    {"role": "user", "content": gpt_prompt}
                                                ],
                                                temperature=0.3,
                                                max_tokens=1000
                                            )
                                            result = gpt_resp.choices[0].message.content
                                        except Exception as e2:
                                            st.error(f"Error in GPT table analysis: {str(e2)}")
                                            result = "(Error in GPT evaluation.)"
                                    else:
                                        result = f"No tables detected in {fpdf.name}."
                                except Exception as e_:
                                    st.error(f"Error in {fpdf.name}: {str(e_)}")
                                    result = f"(Error in {fpdf.name})"
                        if action != "Tabellen & Grafiken" and result:
                            if output_lang != "Deutsch":
                                lang_map = {"Englisch": "English", "Portugiesisch": "Portuguese", "Serbisch": "Serbian"}
                                target_lang = lang_map.get(output_lang, "English")
                                result = translate_text_openai(result, "German", target_lang, api_key)
                        final_result_text.append(f"**Result for {fpdf.name}:**\n\n{result}")
                    st.subheader("Result of (Multi-)Analysis (Single-Mode):")
                    combined_output = "\n\n---\n\n".join(final_result_text)
                    st.markdown(combined_output)

            with col_contradiction:
                st.write("Contradiction Analysis (Uploaded Papers)")
                if st.button("Start Contradiction Analysis now"):
                    if "paper_texts" not in st.session_state or not st.session_state["paper_texts"]:
                        st.session_state["paper_texts"] = {}
                        for upf in uploaded_files:
                            t_ = analyzer.extract_text_from_pdf(upf)
                            if t_.strip():
                                st.session_state["paper_texts"][upf.name] = t_
                    paper_texts = st.session_state["paper_texts"]
                    if not paper_texts:
                        st.error("No texts for contradiction analysis (uploaded PDFs).")
                        return
                    with st.spinner("Analyzing uploaded papers for commonalities & contradictions..."):
                        result_json_str = analyze_papers_for_commonalities_and_contradictions(
                            pdf_texts=paper_texts,
                            api_key=api_key,
                            model=model,
                            method_choice="ContraCrow" if analysis_method == "ContraCrow" else "Standard"
                        )
                        st.subheader("Result (JSON)")
                        st.code(result_json_str, language="json")
                        try:
                            data_js = json.loads(result_json_str)
                            common = data_js.get("commonalities", [])
                            contras = data_js.get("contradictions", [])
                            st.write("## Commonalities")
                            if common:
                                for c in common:
                                    st.write(f"- {c}")
                            else:
                                st.info("No commonalities detected.")
                            st.write("## Contradictions")
                            if contras:
                                for i, cobj in enumerate(contras, start=1):
                                    st.write(f"Contradiction {i}:")
                                    st.write(f"- **Paper A**: {cobj.get('paperA')} => {cobj.get('claimA')}")
                                    st.write(f"- **Paper B**: {cobj.get('paperB')} => {cobj.get('claimB')}")
                                    st.write(f"  Reason: {cobj.get('reason','(none)')}")
                            else:
                                st.info("No contradictions detected.")
                        except Exception as e:
                            st.warning(f"GPT output could not be parsed as valid JSON.\nError: {e}")
    
    else:
        if not api_key:
            st.warning("Please enter an OpenAI API Key!")
        elif not uploaded_files:
            st.info("Please upload one or more PDF files.")

    st.write("---")
    st.write("## All Analyses & Excel Export (Multi-PDF)")
    user_relevance_score = st.text_input("Manual Relevance Score (1-10)?")

    if "excel_downloads" not in st.session_state:
        st.session_state["excel_downloads"] = []

    # ------------------------------------------------------------------
    # NEW: GenotypeFinder (we use a separate class to compute genotype frequencies)
    # ------------------------------------------------------------------
    class GenotypeFinder:
        def __init__(self):
            self.ensembl_server = "https://rest.ensembl.org"
        
        def get_variant_info(self, rs_id):
            """Fetches detailed info about a variation from Ensembl."""
            if not rs_id.startswith("rs"):
                rs_id = f"rs{rs_id}"
            ext = f"/variation/human/{rs_id}?pops=1"
            url = f"{self.ensembl_server}{ext}"
            try:
                r = requests.get(url, headers={"Content-Type": "application/json"}, timeout=10)
                r.raise_for_status()
                return r.json()
            except Exception:
                return None
        
        def calculate_genotype_frequency(self, data, genotype):
            """
            Calculates genotype frequency based on allele frequencies and Hardy-Weinberg.
            data: JSON data from the Ensembl API
            genotype: genotype string (e.g., 'AA','AG','GG')
            Returns: Dict of population => genotype frequency
            """
            if not data or 'populations' not in data:
                return {}
            if len(genotype) != 2:
                return {}
            
            allele1, allele2 = genotype[0], genotype[1]
            results = {}
            
            # We'll only look at 1000GENOMES populations to keep it consistent
            for population in data['populations']:
                pop_name = population.get('population', 'Unknown')
                if '1000GENOMES' not in pop_name:
                    continue
                
                # Gather allele frequencies
                allele_freqs = {}
                for pop2 in data['populations']:
                    if pop2.get('population') == pop_name:
                        a = pop2.get('allele', '')
                        f = pop2.get('frequency', 0)
                        allele_freqs[a] = f
                
                if allele1 not in allele_freqs or allele2 not in allele_freqs:
                    continue
                
                # HW assumption
                if allele1 == allele2:
                    genotype_freq = allele_freqs[allele1] ** 2
                else:
                    genotype_freq = 2 * allele_freqs[allele1] * allele_freqs[allele2]
                
                results[pop_name] = genotype_freq
            
            return results
    
    def build_genotype_freq_text(freq_dict: Dict[str, float]) -> str:
        """Convert genotype frequency dict into an English multiline text."""
        if not freq_dict:
            return "No genotype frequency data found."
        lines = []
        if "1000GENOMES:phase_3:ALL" in freq_dict:
            lines.append(f"Global population: {freq_dict['1000GENOMES:phase_3:ALL']:.4f}")
            lines.append("---")
        for pop, freq in sorted(freq_dict.items()):
            if pop == "1000GENOMES:phase_3:ALL":
                continue
            lines.append(f"{pop}: {freq:.4f}")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Button: "Do all analyses & save to Excel (Multi)"
    # ------------------------------------------------------------------
    if uploaded_files and api_key:
        if st.button("Do all analyses & save to Excel (Multi)"):
            st.session_state["excel_downloads"].clear()
            with st.spinner("Analyzing all uploaded PDFs (for Excel)..."):
                analyzer = PaperAnalyzer(model=model)
                
                if compare_mode:
                    if not st.session_state["relevant_papers_compare"]:
                        paper_map_auto = {}
                        for fpdf in uploaded_files:
                            txt = analyzer.extract_text_from_pdf(fpdf)
                            if txt.strip():
                                paper_map_auto[fpdf.name] = txt
                        if not paper_map_auto:
                            st.error("No usable papers.")
                            return
                        relevant_papers_auto, discovered_theme_auto = do_outlier_logic(paper_map_auto)
                        st.session_state["relevant_papers_compare"] = relevant_papers_auto
                        st.session_state["theme_compare"] = discovered_theme_auto
                    relevant_list_for_excel = st.session_state["relevant_papers_compare"] or []
                    if not relevant_list_for_excel:
                        st.error("No relevant papers remain after outlier-check for Excel.")
                        return
                    selected_files_for_excel = [f for f in uploaded_files if f.name in relevant_list_for_excel]
                else:
                    selected_files_for_excel = uploaded_files

                gf = GenotypeFinder()

                for fpdf in selected_files_for_excel:
                    text = analyzer.extract_text_from_pdf(fpdf)
                    if not text.strip():
                        st.error(f"No text extracted from {fpdf.name} (possibly no OCR). Skipping...")
                        continue
                    
                    summary_de = analyzer.summarize(text, api_key)
                    key_findings_result = analyzer.extract_key_findings(text, api_key)
                    
                    main_theme_for_excel = st.session_state.get("theme_compare", "N/A")
                    if not compare_mode and theme_mode == "Manually":
                        main_theme_for_excel = user_defined_theme or "N/A"
                    
                    if not topic:
                        relevance_result = "(No topic => no relevance rating)"
                    else:
                        relevance_result = analyzer.evaluate_relevance(text, topic, api_key)
                    
                    methods_result = analyzer.identify_methods(text, api_key)
                    
                    # Attempt to find a gene or variant in the text (very basic example)
                    pattern_obvious = re.compile(r"in the\s+([A-Za-z0-9_-]+)\s+gene", re.IGNORECASE)
                    match_text = re.search(pattern_obvious, text)
                    gene_via_text = match_text.group(1) if match_text else None
                    
                    rs_pat = r"(rs\d+)"
                    found_rs_match = re.search(rs_pat, text)
                    rs_num = found_rs_match.group(1) if found_rs_match else None
                    
                    genotype_regex = r"\b([ACGT]{2,3})\b"
                    lines = text.split("\n")
                    found_pairs = []
                    for line in lines:
                        matches = re.findall(genotype_regex, line)
                        if matches:
                            for m in matches:
                                found_pairs.append((m, line.strip()))
                    unique_geno_pairs = []
                    for gp in found_pairs:
                        if gp not in unique_geno_pairs:
                            unique_geno_pairs.append(gp)
                    
                    aff = AlleleFrequencyFinder()
                    allele_freq_info = "No rsID found"
                    if rs_num:
                        data_allele = aff.get_allele_frequencies(rs_num)
                        if not data_allele:
                            data_allele = aff.try_alternative_source(rs_num)
                        if data_allele:
                            allele_freq_info = aff.build_freq_info_text(data_allele)
                    
                    ergebnisse, schlussfolgerungen = split_summary(summary_de)
                    cohort_data = parse_cohort_info(summary_de)
                    study_size = cohort_data.get("study_size", "")
                    origin = cohort_data.get("origin", "")
                    if study_size or origin:
                        cohort_info = (study_size + (", " + origin if origin else "")).strip(", ")
                    else:
                        cohort_info = ""
                    
                    pub_year_match = re.search(r"\b(20[0-9]{2})\b", text)
                    year_for_excel = pub_year_match.group(1) if pub_year_match else "n/a"

                    pmid_pattern = re.compile(r"\bPMID:\s*(\d+)\b", re.IGNORECASE)
                    pmid_match = pmid_pattern.search(text)
                    pmid_found = pmid_match.group(1) if pmid_match else "n/a"

                    doi_final = "n/a"
                    link_pubmed = ""
                    if pmid_found != "n/a":
                        doi_final, link_pubmed = fetch_pubmed_doi_and_link(pmid_found)

                    # Translate to English for Excel
                    ergebnisse_en = translate_text_openai(ergebnisse, "German", "English", api_key) if ergebnisse else ""
                    schlussfolgerungen_en = translate_text_openai(schlussfolgerungen, "German", "English", api_key) if schlussfolgerungen else ""
                    cohort_info_en = translate_text_openai(cohort_info, "German", "English", api_key) if cohort_info else ""
                    key_findings_result_en = translate_text_openai(key_findings_result, "German", "English", api_key) if key_findings_result else ""

                    try:
                        wb = openpyxl.load_workbook("vorlage_paperqa2.xlsx")
                    except FileNotFoundError:
                        st.error("Template 'vorlage_paperqa2.xlsx' was not found!")
                        return
                    ws = wb.active

                    # Fill the main theme & date
                    ws["D2"].value = main_theme_for_excel
                    ws["J2"].value = datetime.datetime.now().strftime("%Y-%m-%d")

                    # Fill gene / rsNumber
                    ws["D5"].value = gene_via_text if gene_via_text else ""
                    ws["D6"].value = rs_num if rs_num else ""
                    
                    # Up to 3 genotype hits
                    for i in range(3):
                        row_i = 10 + i
                        if i < len(unique_geno_pairs):
                            genotype_str = unique_geno_pairs[i][0]
                            ws[f"D{row_i}"].value = genotype_str
                            if rs_num:
                                data_gf = gf.get_variant_info(rs_num)
                                gfreq = gf.calculate_genotype_frequency(data_gf, genotype_str)
                                gf_text = build_genotype_freq_text(gfreq)
                                ws[f"E{row_i}"].value = gf_text
                            else:
                                ws[f"E{row_i}"].value = "No rsID => no genotype frequency"
                        else:
                            ws[f"D{row_i}"] = ""
                            ws[f"E{row_i}"] = ""

                    # Publication year, cohort, key findings
                    ws["C20"].value = year_for_excel
                    ws["D20"].value = cohort_info_en
                    ws["E20"].value = key_findings_result_en

                    # Fill separated summary results (English)
                    ws["G21"].value = ergebnisse_en
                    ws["G22"].value = schlussfolgerungen_en

                    # Fill PMID, link, and DOI
                    ws["J21"].value = pmid_found if pmid_found != "n/a" else ""
                    ws["J22"].value = link_pubmed if link_pubmed else ""
                    ws["I22"].value = doi_final if doi_final != "n/a" else ""

                    output_buffer = io.BytesIO()
                    wb.save(output_buffer)
                    output_buffer.seek(0)
                    
                    xlsx_name = f"analysis_{fpdf.name.replace('.pdf','')}.xlsx"
                    st.session_state["excel_downloads"].append({
                        "label": f"Download Excel for {fpdf.name}",
                        "data": output_buffer.getvalue(),
                        "file_name": xlsx_name
                    })

    if "excel_downloads" in st.session_state and st.session_state["excel_downloads"]:
        st.write("## Generated Excel Downloads:")
        for dl in st.session_state["excel_downloads"]:
            st.download_button(
                label=dl["label"],
                data=dl["data"],
                file_name=dl["file_name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    st.write("---")
    st.write("## Single Analysis of Papers Selected After ChatGPT Scoring")
    
    # Button for scoring
    if st.button("Perform Scoring now"):
        if "search_results" in st.session_state and st.session_state["search_results"]:
            codewords_str = st.session_state.get("codewords", "")
            selected_genes = st.session_state.get("selected_genes", [])
            scored_list = chatgpt_online_search_with_genes(
                papers=st.session_state["search_results"],
                codewords=codewords_str,
                genes=selected_genes,
                top_k=200
            )
            st.session_state["scored_list"] = scored_list
            st.success("Scored papers saved to st.session_state['scored_list']!")
        else:
            st.info("No (previous) search results found, so no scoring possible.")
    
    if "scored_list" not in st.session_state or not st.session_state["scored_list"]:
        st.info("No scored papers yet. Please click 'Perform Scoring now' first.")
        return
    
    st.subheader("Single Analysis of the ChatGPT-Scored Papers")
    scored_titles = [paper["Title"] for paper in st.session_state["scored_list"]]
    chosen_title = st.selectbox(
        "Select a paper from the scoring list:",
        options=["(Please choose)"] + scored_titles
    )
    
    analysis_choice_for_scored_paper = st.selectbox(
        "Which analysis do you want to perform?",
        ["(No selection)", "Zusammenfassung", "Wichtigste Erkenntnisse", "Methoden & Techniken", "Relevanz-Bewertung"]
    )
    
    if chosen_title != "(Please choose)":
        selected_paper = next((p for p in st.session_state["scored_list"] if p["Title"] == chosen_title), None)
        if selected_paper:
            st.write("**Title:** ", selected_paper.get("Title", "n/a"))
            st.write("**Source:** ", selected_paper.get("Source", "n/a"))
            st.write("**PubMed ID:** ", selected_paper.get("PubMed ID", "n/a"))
            st.write("**Year:** ", selected_paper.get("Year", "n/a"))
            st.write("**Publisher:** ", selected_paper.get("Publisher", "n/a"))
            st.write("**Abstract:**")
            abstract = selected_paper.get("Abstract") or ""
            if abstract.strip():
                st.markdown(f"> {abstract}")
            else:
                st.warning(f"No abstract for {selected_paper.get('Title', 'Unnamed')}.")
            
            if st.button("Perform Analysis for this Paper"):
                if not abstract.strip():
                    st.error("No abstract present, cannot analyze.")
                    return
                if analysis_choice_for_scored_paper == "Zusammenfassung":
                    res = analyzer.summarize(abstract, api_key)
                elif analysis_choice_for_scored_paper == "Wichtigste Erkenntnisse":
                    res = analyzer.extract_key_findings(abstract, api_key)
                elif analysis_choice_for_scored_paper == "Methoden & Techniken":
                    res = analyzer.identify_methods(abstract, api_key)
                elif analysis_choice_for_scored_paper == "Relevanz-Bewertung":
                    if not topic:
                        st.error("Please enter a topic in the sidebar.")
                        return
                    res = analyzer.evaluate_relevance(abstract, topic, api_key)
                else:
                    st.info("No valid analysis choice selected.")
                    return

                if res and output_lang != "Deutsch" and analysis_choice_for_scored_paper != "(No selection)":
                    lang_map = {
                        "Englisch": "English",
                        "Portugiesisch": "Portuguese",
                        "Serbisch": "Serbian"
                    }
                    target_lang = lang_map.get(output_lang, "English")
                    res = translate_text_openai(res, "German", target_lang, api_key)
                
                st.write("### Analysis Result:")
                st.write(res)
        else:
            st.warning("Paper not found (unexpected error).")

    st.write("---")
    st.header("PaperQA Multi-Paper Analyzer: Commonalities & Contradictions (Scored Papers)")
    if st.button("Perform Analysis (Scored Papers)"):
        if "scored_list" in st.session_state and st.session_state["scored_list"]:
            paper_texts = {}
            for paper in st.session_state["scored_list"]:
                title = paper.get("Title", "Unnamed")
                abstract = paper.get("Abstract") or ""
                if abstract.strip():
                    paper_texts[title] = abstract
                else:
                    st.warning(f"No abstract for {title}.")
            if not paper_texts:
                st.error("No texts for the analysis.")
            else:
                with st.spinner("Analyzing scored papers for commonalities & contradictions..."):
                    result_json_str = analyze_papers_for_commonalities_and_contradictions(
                        paper_texts,
                        api_key,
                        model,
                        method_choice="ContraCrow" if analysis_method == "ContraCrow" else "Standard"
                    )
                    st.subheader("Result (JSON)")
                    st.code(result_json_str, language="json")
                    try:
                        data_js = json.loads(result_json_str)
                        common = data_js.get("commonalities", [])
                        contras = data_js.get("contradictions", [])
                        st.write("## Commonalities")
                        if common:
                            for c in common:
                                st.write(f"- {c}")
                        else:
                            st.info("No commonalities found.")
                        st.write("## Contradictions")
                        if contras:
                            for i, cobj in enumerate(contras, start=1):
                                st.write(f"Contradiction {i}:")
                                st.write(f"- **Paper A**: {cobj.get('paperA')} => {cobj.get('claimA')}")
                                st.write(f"- **Paper B**: {cobj.get('paperB')} => {cobj.get('claimB')}")
                                st.write(f"  Reason: {cobj.get('reason','(none)')}")
                        else:
                            st.info("No contradictions found.")
                    except Exception as e:
                        st.warning("GPT output could not be parsed as valid JSON.")

# ------------------------------------------------------------------
# NEUE KLASSE & FUNKTION FÜR KI-INHALTSERKENNUNG (AIContentDetector)
# ------------------------------------------------------------------
class AIContentDetector:
    def __init__(self, api_key=None, api_provider=None):
        self.api_key = api_key
        self.api_provider = api_provider
        self.detection_methods = {
            "pattern_analysis": self.analyze_patterns,
            "consistency_check": self.check_consistency,
            "citation_verification": self.verify_citations,
            "api_detection": self.detect_with_api
        }
    
    def analyze_patterns(self, text):
        """Untersucht typische KI-Schreibmuster"""
        # Einige einfache Heuristiken/Regex
        patterns = {
            "wiederholende_phrasen": r'(\b\w+\s+\w+\b)(?=.*\1)',  # Beispiel: wiederholte Wortgruppen
            "gleichmäßiger_ton": r'(jedoch|allerdings|dennoch|daher|folglich|somit)',  # Signalwörter
            "generische_übergänge": r'\b(zunächst|anschließend|abschließend|zusammenfassend)\b'
        }
        
        scores = {}
        for name, pattern in patterns.items():
            matches = re.findall(pattern, text, re.IGNORECASE)
            # Frequenz pro 100 Wörter
            density = len(matches) / (len(text.split()) / 100 + 1e-8)
            # Eine einfache Skalierung  (kein echter wissenschaftl. Ansatz)
            scores[name] = min(100, density * 5)
        
        return sum(scores.values()) / len(scores) if scores else 0
    
    def check_consistency(self, text):
        """Prüft auf konsistente Schreibweise und Ton"""
        paragraphs = text.split('\n\n')
        if len(paragraphs) < 3:
            return 50  # Zu wenig Text für eine sinnvolle Analyse
        
        # Satzlängenvariation
        sentences = re.split(r'[.!?]+', text)
        lengths = [len(s.split()) for s in sentences if s.strip()]
        if not lengths:
            return 50
        
        avg_length = sum(lengths) / len(lengths)
        variation = sum(abs(l - avg_length) for l in lengths) / len(lengths)
        
        # Niedrige Variation => KI-typisch (Heuristik)
        consistency_score = 100 - min(100, variation * 10)
        return consistency_score
    
    def verify_citations(self, text):
        """Überprüft Zitate auf Plausibilität (sehr einfach)"""
        citation_pattern = r'\(([^)]+\d{4}[^)]*)\)'
        citations = re.findall(citation_pattern, text)
        
        if not citations:
            return 60  # Keine Zitate gefunden => neutraler Wert
        
        # Einfache Heuristik: sind viele Zitate im gleichen Format?
        formats = {}
        for citation in citations:
            format_key = re.sub(r'[A-Za-z\s]', 'X', citation)
            format_key = re.sub(r'\d', '9', format_key)
            formats[format_key] = formats.get(format_key, 0) + 1
        
        uniformity = max(formats.values()) / len(citations) * 100
        return uniformity
    
    def detect_with_api(self, text):
        """Verwendet externe APIs (Originality.ai oder Scribbr)"""
        if not self.api_key:
            return 50  # Keine API => Mittelwert
        
        # Originality.ai
        if self.api_provider == "originality":
            try:
                response = requests.post(
                    "https://api.originality.ai/api/v1/scan/ai",
                    headers={"X-OAI-API-KEY": self.api_key},
                    json={"content": text}
                )
                if response.status_code == 200:
                    result = response.json()
                    # 'score.ai' => 0-1
                    return result.get("score", {}).get("ai", 0.5) * 100
            except Exception as e:
                print(f"Originality.ai API-Fehler: {e}")
        
        # Scribbr (Beispiel, es gibt keine offizielle Public-API-Doku)
        elif self.api_provider == "scribbr":
            try:
                response = requests.post(
                    "https://api.scribbr.com/v1/ai-detection",  # fiktiver Endpunkt
                    headers={"Authorization": f"Bearer {self.api_key}"},
                    json={"text": text}
                )
                if response.status_code == 200:
                    result = response.json()
                    # Annahme: "ai_probability" => 0-100
                    return result.get("ai_probability", 50)
            except Exception as e:
                print(f"Scribbr API-Fehler: {e}")
        
        # Fallback
        return 50
    
    def analyze_text(self, text):
        """Führt eine komplette Analyse durch."""
        scores = {}
        for method_name, method_func in self.detection_methods.items():
            scores[method_name] = method_func(text)
        
        # Gewichtung
        weights = {
            "pattern_analysis": 0.20,
            "consistency_check": 0.20,
            "citation_verification": 0.10,
            "api_detection": 0.50
        }
        
        weighted_score = sum(scores[m] * weights[m] for m in scores)
        return {
            "gesamtbewertung": round(weighted_score, 2),
            "einzelbewertungen": {m: round(scores[m], 2) for m in scores},
            "interpretation": self.interpret_score(weighted_score)
        }
    
    def interpret_score(self, score):
        """Interpretation der KI-Wahrscheinlichkeit"""
        if score < 30:
            return "Wahrscheinlich von Menschen geschrieben"
        elif score < 60:
            return "Unklare Herkunft, könnte teilweise KI-unterstützt sein"
        elif score < 85:
            return "Wahrscheinlich KI-unterstützt oder überarbeitet"
        else:
            return "Sehr wahrscheinlich vollständig KI-generiert"

# ------------------------------------------------------------------
# NEUES MODULE/PAGE: KI-Inhaltserkennung via AIContentDetector
# ------------------------------------------------------------------
def page_ai_content_detection():
    """Seite zur Erkennung von KI-Textinhalten (Paper etc.)."""
    st.title("KI-Inhaltserkennung (AI Content Detector)")
    
    st.info("Hier kannst du Text eingeben oder eine Datei hochladen, um eine KI-Analyse durchzuführen.")
    
    # API-Infos für Originality oder Scribbr
    api_key_input = st.text_input("API Key (optional)", value="", type="password")
    provider_option = st.selectbox("API-Anbieter", ["Kein API-Einsatz", "originality", "scribbr"], index=0)
    
    # Eingabemethode
    input_mode = st.radio("Eingabemethode für den Text:", ["Direkte Eingabe", "Textdatei hochladen"])
    
    text_data = ""
    if input_mode == "Direkte Eingabe":
        text_data = st.text_area("Gib hier deinen Text ein:", height=200)
    else:
        uploaded_text_file = st.file_uploader("Text-Datei wählen (.txt, .md, etc.)", type=["txt","md","csv","json"])
        if uploaded_text_file is not None:
            try:
                text_data = uploaded_text_file.read().decode("utf-8", errors="ignore")
            except Exception as e:
                st.error(f"Fehler beim Lesen der Datei: {e}")
                return
    
    if st.button("KI-Analyse starten"):
        if not text_data.strip():
            st.warning("Bitte Text eingeben oder Datei hochladen.")
            return
        
        # Detector instanziieren
        if provider_option == "Kein API-Einsatz":
            detector = AIContentDetector(api_key=None, api_provider=None)
        else:
            detector = AIContentDetector(api_key=api_key_input, api_provider=provider_option.lower())
        
        with st.spinner("Analyse läuft..."):
            result = detector.analyze_text(text_data)
        
        st.subheader("Ergebnis der KI-Analyse")
        gesamtbewertung = result["gesamtbewertung"]
        interpretation = result["interpretation"]
        einzelbewertungen = result["einzelbewertungen"]
        
        st.metric("KI-Wahrscheinlichkeit (gesamt)", f"{gesamtbewertung} %", help=interpretation)
        st.write("**Interpretation:** ", interpretation)
        
        st.write("### Einzelbewertungen")
        for method, score in einzelbewertungen.items():
            st.write(f"- **{method}**: {score} %")
        
        if provider_option != "Kein API-Einsatz":
            st.write(f"Verwendeter API-Dienst: **{provider_option}**")
        else:
            st.write("**Hinweis:** Keine externe API genutzt, nur lokale Heuristiken.")

# ------------------------------------------------------------------
# Seite: Genotype Frequency Finder
# ------------------------------------------------------------------
def page_genotype_finder():
    """
    A separate page to look up genotype frequencies via Ensembl for a user-provided rsID & genotype.
    """
    st.title("Genotype Frequency Finder")

    class GenotypeFinder:
        def __init__(self):
            self.ensembl_server = "https://rest.ensembl.org"

        def get_variant_info(self, rs_id):
            if not rs_id.startswith("rs"):
                rs_id = f"rs{rs_id}"
            ext = f"/variation/human/{rs_id}?pops=1"
            url = f"{self.ensembl_server}{ext}"
            try:
                r = requests.get(url, headers={"Content-Type": "application/json"}, timeout=10)
                r.raise_for_status()
                return r.json()
            except:
                return None
        
        def calculate_genotype_frequency(self, data, genotype):
            if not data or 'populations' not in data:
                return {}
            if len(genotype) < 2:
                return {}
            allele1, allele2 = genotype[0], genotype[1]
            results = {}
            for pop in data['populations']:
                pop_name = pop.get('population', '')
                if '1000GENOMES' not in pop_name:
                    continue
                allele_freq_map = {}
                for pop2 in data['populations']:
                    if pop2.get('population') == pop_name:
                        a_ = pop2.get('allele')
                        f_ = pop2.get('frequency')
                        allele_freq_map[a_] = f_
                if allele1 in allele_freq_map and allele2 in allele_freq_map:
                    if allele1 == allele2:
                        freq_g = allele_freq_map[allele1] ** 2
                    else:
                        freq_g = 2 * allele_freq_map[allele1] * allele_freq_map[allele2]
                    results[pop_name] = freq_g
            return results
    
    def build_genotype_freq_text(freq_dict: Dict[str, float]) -> str:
        if not freq_dict:
            return "No genotype frequency data found."
        lines = []
        if "1000GENOMES:phase_3:ALL" in freq_dict:
            lines.append(f"Global population (ALL): {freq_dict['1000GENOMES:phase_3:ALL']:.4f}")
            lines.append("---")
        for pop, val in sorted(freq_dict.items()):
            if pop == "1000GENOMES:phase_3:ALL":
                continue
            lines.append(f"{pop}: {val:.4f}")
        return "\n".join(lines)

    st.write("Look up genotype frequencies for a given rsID (from Ensembl).")
    rs_input = st.text_input("Enter an rsID (e.g., 'rs1234'):", "")
    genotype_input = st.text_input("Enter a genotype (e.g., 'AA','AC','CC','AG', etc.):", "")

    if st.button("Check Frequencies"):
        if not rs_input.strip():
            st.warning("Please enter an rsID.")
            return
        gf = GenotypeFinder()
        data = gf.get_variant_info(rs_input.strip())
        if not data:
            st.error(f"No data found for {rs_input.strip()}. Are you sure it's correct?")
            return
        freq_dict = gf.calculate_genotype_frequency(data, genotype_input.strip().upper())
        freq_text = build_genotype_freq_text(freq_dict)
        st.subheader("Result:")
        st.write(freq_text)

# ------------------------------------------------------------------
# Sidebar Navigation & Chatbot
# ------------------------------------------------------------------
def sidebar_module_navigation():
    st.sidebar.title("Module Navigation")

    pages = {
        "Home": page_home,
        "Online-API_Filter": page_online_api_filter,
        "3) Codewords & PubMed": page_codewords_pubmed,
        "Analyze Paper": page_analyze_paper,
        "Genotype Frequency Finder": page_genotype_finder,
        "AI-Content Detection": page_ai_content_detection  # <-- NEU HINZUGEFÜGT
    }

    for label, page in pages.items():
        if st.sidebar.button(label, key=label):
            st.session_state["current_page"] = label
    
    if "current_page" not in st.session_state:
        st.session_state["current_page"] = "Home"
    return pages.get(st.session_state["current_page"], page_home)

def answer_chat(question: str) -> str:
    """Simple example: uses Paper text (if available) from st.session_state + GPT."""
    api_key = st.session_state.get("api_key", "")
    paper_text = st.session_state.get("paper_text", "")
    if not api_key:
        return f"(No API-Key) Echo: {question}"
    if not paper_text.strip():
        sys_msg = "You are a helpful assistant for general questions."
    else:
        sys_msg = (
            "You are a helpful assistant, and here is a paper as context:\n\n"
            + paper_text[:12000] + "\n\n"
            "Please use it to answer questions as expertly as possible."
        )
    openai.api_key = api_key
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": sys_msg},
                {"role": "user", "content": question}
            ],
            temperature=0.3,
            max_tokens=400
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"OpenAI error: {e}"

def main():
    # -------- LAYOUT: Left Modules, Right Chatbot --------
    col_left, col_right = st.columns([4, 1])
    
    with col_left:
        # Navigation
        page_fn = sidebar_module_navigation()
        if page_fn is not None:
            page_fn()
    
    with col_right:
        st.subheader("Chatbot")
        if "chat_history" not in st.session_state:
            st.session_state["chat_history"] = []
        user_input = st.text_input("Your question here", key="chatbot_right_input")
        if st.button("Send (Chat)", key="chatbot_right_send"):
            if user_input.strip():
                st.session_state["chat_history"].append(("user", user_input))
                bot_answer = answer_chat(user_input)
                st.session_state["chat_history"].append(("bot", bot_answer))
        
        st.markdown(
            """
            <style>
            .scrollable-chat {
                max-height: 400px; 
                overflow-y: auto; 
                border: 1px solid #CCC;
                padding: 8px;
                margin-top: 10px;
                border-radius: 4px;
                background-color: #f9f9f9;
            }
            .message {
                padding: 0.5rem 1rem;
                border-radius: 15px;
                margin-bottom: 0.5rem;
                max-width: 80%;
                word-wrap: break-word;
            }
            .user-message {
                background-color: #e3f2fd;
                margin-left: auto;
                border-bottom-right-radius: 0;
            }
            .assistant-message {
                background-color: #f0f0f0;
                margin-right: auto;
                border-bottom-left-radius: 0;
            }
            </style>
            """,
            unsafe_allow_html=True
        )
        st.markdown('<div class="scrollable-chat" id="chat-container">', unsafe_allow_html=True)
        for role, msg_text in st.session_state["chat_history"]:
            if role == "user":
                st.markdown(
                    f'<div class="message user-message"><strong>You:</strong> {msg_text}</div>',
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f'<div class="message assistant-message"><strong>Bot:</strong> {msg_text}</div>',
                    unsafe_allow_html=True
                )
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Auto-scroll JS
        st.markdown(
            """
            <script>
                function scrollToBottom() {
                    var container = document.getElementById('chat-container');
                    if(container) {
                        container.scrollTop = container.scrollHeight;
                    }
                }
                document.addEventListener('DOMContentLoaded', function() {
                    scrollToBottom();
                });
                const observer = new MutationObserver(function(mutations) {
                    scrollToBottom();
                });
                setTimeout(function() {
                    var container = document.getElementById('chat-container');
                    if(container) {
                        observer.observe(container, { childList: true });
                        scrollToBottom();
                    }
                }, 1000);
            </script>
            """,
            unsafe_allow_html=True
        )

# ------------------------------------------------------------------
# Actually run the Streamlit app
# ------------------------------------------------------------------
if __name__ == '__main__':
    main()
import os
import sys
import subprocess
import webbrowser
import base64
import time
import datetime
from importlib import metadata  # ab Python 3.8
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from urllib.parse import quote_plus
from fpdf import FPDF
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import fitz  # PyMuPDF
import re
import openpyxl
import xml.etree.ElementTree as ET

# Neu: Laden der .env-Variablen
from dotenv import load_dotenv

# Eigene ausgelagerte Module für Analyse & erweiterte Themen
from analysis_and_review import AnalysisWindow
from extended_topics import ExtendedTopicsWindow

# .env-Datei laden (muss im selben Verzeichnis wie main.py liegen)
load_dotenv()

# API-Keys aus den Environment-Variablen lesen
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY")
PERPLEXITY_API_KEY = os.environ.get("PERPLEXITY_API_KEY")
CHATGPT_API_KEY = os.environ.get("CHATGPT_API_KEY")

# ChatGPT-Key global setzen
import openai
openai.api_key = CHATGPT_API_KEY


#############################################
# DBSnpAPI-Klasse (NEUE Logik)
#############################################
class DBSnpAPI:
    def __init__(self, email: str, api_key: str):
        self.base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
        self.params = {
            "db": "snp",
            "retmode": "xml",
            "email": email,
            "api_key": api_key,
            "tool": "DBSnpPythonClient"
        }
        self.namespace = {'ns': 'https://www.ncbi.nlm.nih.gov/SNP/docsum'}

    def get_snp_info(self, rs_id: str):
        try:
            search_params = {"term": f"{rs_id}[RS]", "retmax": "1"}
            search_response = requests.get(
                f"{self.base_url}esearch.fcgi",
                params={**self.params, **search_params},
                timeout=10
            )
            search_response.raise_for_status()
            search_root = ET.fromstring(search_response.content)
            snp_id = search_root.findtext("IdList/Id")
            if not snp_id:
                print("SNP-ID nicht gefunden")
                return None

            fetch_response = requests.get(
                f"{self.base_url}efetch.fcgi",
                params={**self.params, "id": snp_id},
                timeout=10
            )
            fetch_response.raise_for_status()
            return self.parse_xml(fetch_response.content)

        except Exception as e:
            print(f"Fehler: {str(e)}")
            return None

    def parse_xml(self, xml_content):
        root = ET.fromstring(xml_content)
        snp_info = {}
        doc_summary = root.find(".//ns:DocumentSummary", self.namespace)
        if doc_summary is not None:
            snp_info["rs_id"] = f"rs{doc_summary.findtext('ns:SNP_ID', '', self.namespace)}"
            snp_info["chromosome"] = doc_summary.findtext("ns:CHR", "", self.namespace)
            snp_info["position"] = doc_summary.findtext("ns:CHRPOS", "", self.namespace)
            snp_info["alleles"] = doc_summary.findtext("ns:SPDI", "", self.namespace)

            clin_sig = doc_summary.findtext("ns:CLINICAL_SIGNIFICANCE", "", self.namespace)
            snp_info["clinical_significance"] = clin_sig.split(",") if clin_sig else []

            snp_info["gene"] = doc_summary.findtext("ns:GENES/ns:GENE_E/ns:NAME", "", self.namespace)

            mafs = []
            for maf in doc_summary.findall(".//ns:MAF", self.namespace):
                study = maf.findtext("ns:STUDY", "", self.namespace)
                freq = maf.findtext("ns:FREQ", "", self.namespace)
                mafs.append(f"{study}: {freq}")
            snp_info["mafs"] = mafs

        return snp_info


#############################################
# CoreAPI-Klasse für CORE Aggregate
#############################################
class CoreAPI:
    def __init__(self, api_key):
        self.base_url = "https://api.core.ac.uk/v3/"
        self.headers = {"Authorization": f"Bearer {api_key}"}

    def search_publications(self, query, filters=None, sort=None, limit=10):
        endpoint = "search/works"
        params = {"q": query, "limit": limit}
        if filters:
            filter_expressions = []
            for key, value in filters.items():
                filter_expressions.append(f"{key}:{value}")
            params["filter"] = ",".join(filter_expressions)
        if sort:
            params["sort"] = sort
        r = requests.get(
            self.base_url + endpoint,
            headers=self.headers,
            params=params,
            timeout=15
        )
        r.raise_for_status()
        return r.json()


#############################################
# Hilfsfunktion: URL in neuem Chrome-Fenster öffnen
#############################################
def open_in_new_chrome_window(url):
    try:
        subprocess.Popen(["chrome", "--new-window", url])
    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler beim Öffnen in Chrome: {e}")


#############################################
# PubMed-Verbindungsprüfung
#############################################
def check_pubmed_connection(timeout=10):
    test_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
    params = {"db": "pubmed", "term": "test", "retmode": "json"}
    try:
        r = requests.get(test_url, params=params, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        if "esearchresult" in data:
            return True
        else:
            return False
    except Exception:
        return False


#############################################
# Haupt-App-Klasse
#############################################
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("API-Suche und PDF-Erstellung")
        self.root.geometry("1600x900")

        self.blink_state = True
        self.pubmed_id_search = False

        # Online-Filter
        self.filter_local_var = tk.BooleanVar(value=True)
        self.filter_perplexity_var = tk.BooleanVar(value=False)
        self.filter_deepseek_var = tk.BooleanVar(value=False)
        self.filter_chatgpt_var = tk.BooleanVar(value=False)

        # ChatGPT-spezifisch
        self.chatgpt_genotype_var = tk.BooleanVar(value=True)
        self.chatgpt_phenotype_var = tk.BooleanVar(value=True)
        self.chatgpt_snp_var = tk.BooleanVar(value=True)

        # Extra-Filter
        self.filter_extra_var = tk.BooleanVar(value=False)
        self.extra_filter_string = tk.StringVar(value="")

        # Gene (Excel)
        self.filter_gene_excel_var = tk.BooleanVar(value=False)
        self.gene_excel_selected_sheet = None
        self.gene_excel_genes = []

        self.font_name = "DejaVu"
        self.font_file = "DejaVuSansCondensed.ttf"
        self._ensure_font_available()

        # PubMed-Check
        if check_pubmed_connection():
            self_connection_msg = "[INFO] Verbindung zu PubMed erfolgreich hergestellt."
        else:
            self_connection_msg = "[WARNUNG] Keine Verbindung zu PubMed."

        self.all_results = []
        self.page_size = 50
        self.current_page = 0

        self.available_excels = {
            "Genetik": r"C:\Users\ErwinSchimak\Desktop\Reserach\Journals\Genetik.xlsx",
            "Medizin": r"C:\Users\ErwinSchimak\Desktop\Reserach\Journals\Medizin.xlsx"
        }
        self.current_excel_path = self.available_excels["Genetik"]
        self.journals, self.journal_lookup = self._load_journals_from_excel(self.current_excel_path)

        self.excel_details = []
        self.excel_main = []

        # PubMed session
        self.pubmed_session = requests.Session()
        retries = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            raise_on_status=False
        )
        adapter = HTTPAdapter(max_retries=retries)
        self.pubmed_session.mount("http://", adapter)
        self.pubmed_session.mount("https://", adapter)

        # ------------- Frame: API Selection -------------
        api_frame = tk.LabelFrame(self.root, text="Wähle APIs aus", padx=10, pady=10)
        api_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

        self.api_vars = {
            "Europe PMC": tk.BooleanVar(value=True),
            "PubMed": tk.BooleanVar(value=False),
            "Ensembl REST": tk.BooleanVar(value=False),
            "UniProt": tk.BooleanVar(value=False),
            "OpenAlex": tk.BooleanVar(value=False),
            "Scopus": tk.BooleanVar(value=False),
            "CORE Aggregate": tk.BooleanVar(value=False),
            "Unpaywall": tk.BooleanVar(value=False),
            "NCBI e": tk.BooleanVar(value=False),
            "PLOS": tk.BooleanVar(value=False),
            "Hathitrust": tk.BooleanVar(value=False),
            "OpenCitations": tk.BooleanVar(value=False),
            "DataCite": tk.BooleanVar(value=False),
            "Google Scholar": tk.BooleanVar(value=False),
            "Semantic Scholar": tk.BooleanVar(value=False)
        }
        for api_name, var in self.api_vars.items():
            cb = tk.Checkbutton(api_frame, text=api_name, variable=var)
            cb.pack(side=tk.LEFT, padx=5)

        # ------------- Frame: Online Filter (erweitert) -------------
        filter_options_frame = tk.LabelFrame(self.root, text="Online-Filter Optionen", padx=10, pady=10)
        filter_options_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

        tk.Label(filter_options_frame, text="Wähle, welche Online-Filter angewendet werden sollen:").pack(side=tk.LEFT)
        tk.Checkbutton(filter_options_frame, text="Local (Schlüsselwörter)", variable=self.filter_local_var).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(filter_options_frame, text="Perplexity API", variable=self.filter_perplexity_var).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(filter_options_frame, text="DeepSeek API", variable=self.filter_deepseek_var).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(filter_options_frame, text="ChatGPT API", variable=self.filter_chatgpt_var).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(filter_options_frame, text="Genotype", variable=self.chatgpt_genotype_var).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(filter_options_frame, text="Phenotype", variable=self.chatgpt_phenotype_var).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(filter_options_frame, text="SNP", variable=self.chatgpt_snp_var).pack(side=tk.LEFT, padx=5)

        # Gene (Excel)
        tk.Checkbutton(filter_options_frame, text="Gene (Excel)", variable=self.filter_gene_excel_var, command=self._open_excel_sheet_for_genes).pack(side=tk.LEFT, padx=10)

        # Extra-Filter
        extra_filter_frame = tk.Frame(filter_options_frame)
        extra_filter_frame.pack(side=tk.LEFT, padx=20)
        tk.Checkbutton(extra_filter_frame, text="Extra-Filter aktivieren:", variable=self.filter_extra_var).pack(side=tk.LEFT)
        self.extra_filter_entry = tk.Entry(extra_filter_frame, textvariable=self.extra_filter_string, width=15)
        self.extra_filter_entry.pack(side=tk.LEFT, padx=5)
        btn_set_extra = tk.Button(extra_filter_frame, text="Extra Filterbegriff setzen", command=self._set_extra_filter)
        btn_set_extra.pack(side=tk.LEFT)

        # ------------- Frame: Code Words + Filter -------------
        top_frame = tk.Frame(self.root)
        top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        tk.Label(top_frame, text="Codewörter (OR-Suche):").pack(side=tk.LEFT)
        self.codeword_entry = tk.Entry(top_frame, width=50)
        self.codeword_entry.pack(side=tk.LEFT, padx=5)
        btn_add_code = tk.Button(top_frame, text="+ Codewort", command=self.add_code_word_entry)
        btn_add_code.pack(side=tk.LEFT, padx=5)

        tk.Label(top_frame, text="Genotyp:").pack(side=tk.LEFT, padx=5)
        self.genotype_var = tk.StringVar()
        genotype_options = ["", "A/A", "A/G", "A/T", "C/C", "C/G", "C/T", "G/G", "G/T", "T/T"]
        self.genotype_combo = ttk.Combobox(top_frame, textvariable=self.genotype_var, values=genotype_options, width=6, state="readonly")
        self.genotype_combo.pack(side=tk.LEFT, padx=5)
        self.genotype_combo.set("")

        btn_search = tk.Button(top_frame, text="Suchen", command=self.search_articles)
        btn_search.pack(side=tk.LEFT, padx=5)
        btn_clear = tk.Button(top_frame, text="Suchergebnisse löschen", command=self.clear_search_results)
        btn_clear.pack(side=tk.LEFT, padx=5)
        self.codeword_entry.bind("<Return>", self.search_articles)

        self.words_frame = tk.Frame(self.root)
        self.words_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        self.code_word_entries = []
        self.add_code_word_entry()

        # ------------- PubMed ID Search -------------
        pubmed_id_frame = tk.Frame(self.root)
        pubmed_id_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        tk.Label(pubmed_id_frame, text="PubMed ID(s) (durch Kommas trennen):").pack(side=tk.LEFT)
        self.pubmed_id_entry = tk.Entry(pubmed_id_frame, width=50)
        self.pubmed_id_entry.pack(side=tk.LEFT, padx=5)
        btn_pubmed_id_search = tk.Button(pubmed_id_frame, text="PubMed ID suchen", command=self.search_pubmed_ids)
        btn_pubmed_id_search.pack(side=tk.LEFT, padx=5)

        # ------------- Frame: Search and Filter Criteria -------------
        criteria_frame = tk.Frame(self.root)
        criteria_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)
        self.apply_search_criteria = tk.BooleanVar(value=False)

        # --- Suchkriterien ---
        such_frame = tk.LabelFrame(criteria_frame, text="Suchkriterien", padx=10, pady=10)
        such_frame.pack(side=tk.LEFT, padx=10, pady=10)
        chk_such = tk.Checkbutton(such_frame, text="Suchkriterien anwenden", variable=self.apply_search_criteria)
        chk_such.grid(row=0, column=0, columnspan=2, sticky="w")

        self.search_populationsgroesse = tk.BooleanVar(value=False)
        tk.Checkbutton(such_frame, text="Populationsgröße (falls vorhanden)", variable=self.search_populationsgroesse).grid(row=1, column=0, sticky="w", padx=5, pady=2)

        tk.Label(such_frame, text="Erscheinungsjahr:").grid(row=2, column=0, sticky="w")
        years_such = list(range(1975, 2026))
        self.search_year_var = tk.StringVar()
        self.search_year_combo = ttk.Combobox(such_frame, textvariable=self.search_year_var, values=years_such, state="readonly")
        self.search_year_combo.grid(row=2, column=1, padx=5, pady=2)
        self.search_year_combo.set("")

        tk.Label(such_frame, text="Journal (aus Excel):").grid(row=3, column=0, sticky="w")
        self.search_journal_var = tk.StringVar()
        self.search_journal_combo = ttk.Combobox(such_frame, textvariable=self.search_journal_var, values=[], state="readonly")
        self.search_journal_combo.grid(row=3, column=1, padx=5, pady=2)
        self.search_journal_combo.set("")

        tk.Label(such_frame, text="Bereich (Excel-Datei):").grid(row=4, column=0, sticky="w")
        self.search_bereich_var = tk.StringVar()
        self.search_bereich_combo = ttk.Combobox(such_frame, textvariable=self.search_bereich_var, values=list(self.available_excels.keys()), state="readonly")
        self.search_bereich_combo.grid(row=4, column=1, padx=5, pady=2)
        self.search_bereich_combo.set("Genetik")
        self.search_bereich_combo.bind("<<ComboboxSelected>>", self._on_search_excel_file_change)

        # --- Filterkriterien ---
        filter_frame = tk.LabelFrame(criteria_frame, text="Filterkriterien", padx=10, pady=10)
        filter_frame.pack(side=tk.LEFT, padx=10, pady=10)

        tk.Label(filter_frame, text="Bereich (Excel-Datei):").grid(row=0, column=0, sticky="w")
        self.bereich_var = tk.StringVar()
        self.bereich_combo = ttk.Combobox(filter_frame, textvariable=self.bereich_var, values=list(self.available_excels.keys()), state="readonly")
        self.bereich_combo.grid(row=0, column=1, padx=5, pady=5)
        self.bereich_combo.set("Genetik")
        self.bereich_combo.bind("<<ComboboxSelected>>", self._on_excel_file_change)

        self.var_populationsgroesse = tk.BooleanVar(value=False)
        tk.Checkbutton(filter_frame, text="Populationsgröße (falls vorhanden)", variable=self.var_populationsgroesse).grid(row=1, column=0, padx=5, pady=5, sticky="w")

        tk.Label(filter_frame, text="Erscheinungsjahr:").grid(row=2, column=0, sticky="w")
        years_filter = list(range(1975, 2026))
        self.year_var = tk.StringVar()
        self.year_combo = ttk.Combobox(filter_frame, textvariable=self.year_var, values=years_filter, state="readonly")
        self.year_combo.grid(row=2, column=1, padx=5, pady=5)
        self.year_combo.set("")

        tk.Label(filter_frame, text="Journal (aus Excel):").grid(row=3, column=0, sticky="w")
        self.journal_var = tk.StringVar()
        self.journal_combo = ttk.Combobox(filter_frame, textvariable=self.journal_var, values=self.journals, state="readonly")
        self.journal_combo.grid(row=3, column=1, padx=5, pady=5)
        self.journal_combo.set("")

        btn_auto_select = tk.Button(filter_frame, text="Auto-Select", command=self.auto_select_papers)
        btn_auto_select.grid(row=4, column=0, columnspan=2, pady=5)

        # --- UniProt ---
        uniprot_frame = tk.LabelFrame(criteria_frame, text="UniProt-Auswahl", padx=10, pady=10)
        uniprot_frame.pack(side=tk.LEFT, padx=10, pady=10)
        tk.Label(uniprot_frame, text="Protein:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.uniprot_protein_var = tk.StringVar()
        protein_list = ["TP53", "MYH7", "EGFR", "ACE2", "GFP", "Hämoglobin", "Myoglobin", "Insulin"]
        self.uniprot_protein_combo = ttk.Combobox(uniprot_frame, textvariable=self.uniprot_protein_var, values=protein_list, state="readonly", width=20)
        self.uniprot_protein_combo.grid(row=0, column=1, padx=5, pady=2)
        self.uniprot_protein_combo.set("")
        btn_uniprot_search = tk.Button(uniprot_frame, text="UniProt suchen", command=self.search_uniprot_direct)
        btn_uniprot_search.grid(row=1, column=0, columnspan=2, pady=5)
        btn_uniprot_manual = tk.Button(uniprot_frame, text="Manuell eingeben", command=self.open_manual_uniprot)
        btn_uniprot_manual.grid(row=2, column=0, columnspan=2, pady=5)

        # --- Ensembl ---
        ensembl_frame = tk.LabelFrame(criteria_frame, text="Ensembl-Auswahl", padx=10, pady=10)
        ensembl_frame.pack(side=tk.LEFT, padx=10, pady=10)
        tk.Label(ensembl_frame, text="Spezies:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.ensembl_species_var = tk.StringVar()
        self.ensembl_species_var.set("homo_sapiens")
        ttk.Entry(ensembl_frame, textvariable=self.ensembl_species_var, width=15).grid(row=0, column=1, padx=5, pady=2)
        tk.Label(ensembl_frame, text="Gen-Symbol:").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        self.ensembl_gene_var = tk.StringVar()
        ttk.Entry(ensembl_frame, textvariable=self.ensembl_gene_var, width=15).grid(row=1, column=1, padx=5, pady=2)
        btn_ensembl_search = tk.Button(ensembl_frame, text="Ensembl suchen", command=self.search_ensembl_direct)
        btn_ensembl_search.grid(row=2, column=0, columnspan=2, pady=5)

        # --- dbSNP ---
        dbsnp_frame = tk.LabelFrame(criteria_frame, text="NCBI dbSNP-Auswahl", padx=10, pady=10)
        dbsnp_frame.pack(side=tk.LEFT, padx=10, pady=10)
        tk.Label(dbsnp_frame, text="RS-ID:").grid(row=0, column=0, sticky="w", padx=5, pady=2)
        self.dbsnp_rs_entry = tk.Entry(dbsnp_frame, width=15)
        self.dbsnp_rs_entry.grid(row=0, column=1, padx=5, pady=2)
        btn_dbsnp_search = tk.Button(dbsnp_frame, text="dbSNP suchen", command=self.search_dbsnp_direct)
        btn_dbsnp_search.grid(row=1, column=0, columnspan=2, pady=5)

        # ------------- Roter Balken mit Log -------------
        self.found_frame = tk.Frame(self.root, bg="red", highlightthickness=2, highlightcolor="red", bd=2, relief="solid")
        self.found_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)

        self.total_label = tk.Label(self.found_frame, text="Insgesamt 0 Paper gefunden", bg="red", fg="white")
        self.total_label.pack(side=tk.LEFT, padx=10)

        # Blauer Zähler
        self.blue_label = tk.Label(self.found_frame, text="", bg="blue", fg="white", width=40, height=2)
        self.blue_label.pack(side=tk.RIGHT, padx=10)

        page_button_frame = tk.Frame(self.found_frame, bg="red")
        page_button_frame.pack(side=tk.RIGHT)

        self.prev_button = tk.Button(page_button_frame, text="<< Zurück", command=self.prev_page)
        self.prev_button.pack(side=tk.LEFT, padx=5)
        self.next_button = tk.Button(page_button_frame, text="Weiter >>", command=self.next_page)
        self.next_button.pack(side=tk.LEFT, padx=5)

        # ScrolledText für Log
        self.log_text = scrolledtext.ScrolledText(self.found_frame, wrap="word", width=60, height=5, bg="#EEE")
        self.log_text.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(5, 5))

        self.update_log(self_connection_msg)

        # ------------- Verschieben in rechte Treeview -------------
        move_button_frame = tk.Frame(self.root)
        move_button_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        btn_add_selected = tk.Button(move_button_frame, text=">> Auswählen", command=self.add_selected_papers)
        btn_add_selected.pack(side=tk.LEFT, padx=5)
        btn_remove_selected = tk.Button(move_button_frame, text="<< Entfernen", command=self.remove_selected_papers)
        btn_remove_selected.pack(side=tk.LEFT, padx=5)

        # ------------- Trees Frame (Left + Right) -------------
        trees_frame = tk.Frame(self.root)
        trees_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)

        left_tree_frame = tk.Frame(trees_frame)
        left_tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        tk.Label(left_tree_frame, text="Gefundene Paper (max. 50 pro Seite)").pack(side=tk.TOP)
        left_cols = ("Source", "Title", "Authors/Description", "Journal/Organism", "Year", "PMID", "DOI", "URL")
        self.left_tree = ttk.Treeview(left_tree_frame, columns=left_cols, show="headings", selectmode="extended")
        for c in left_cols:
            self.left_tree.heading(c, text=c)
            if c == "Title":
                self.left_tree.column(c, width=300)
            elif c == "Authors/Description":
                self.left_tree.column(c, width=250)
            elif c == "Journal/Organism":
                self.left_tree.column(c, width=150)
            elif c == "URL":
                self.left_tree.column(c, width=200)
            else:
                self.left_tree.column(c, width=100)
        self.left_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.left_tree.bind("<Double-1>", self.show_abstract_popup)

        right_tree_frame = tk.Frame(trees_frame)
        right_tree_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        tk.Label(right_tree_frame, text="Ausgewählte Paper").pack(side=tk.TOP)
        right_cols = ("Source", "Title", "Authors/Description", "Journal/Organism", "Year", "PMID", "DOI", "URL")
        self.right_tree = ttk.Treeview(right_tree_frame, columns=right_cols, show="headings", selectmode="extended")
        for c in right_cols:
            self.right_tree.heading(c, text=c)
            if c == "Title":
                self.right_tree.column(c, width=300)
            elif c == "Authors/Description":
                self.right_tree.column(c, width=250)
            elif c == "Journal/Organism":
                self.right_tree.column(c, width=150)
            elif c == "URL":
                self.right_tree.column(c, width=200)
            else:
                self.right_tree.column(c, width=100)
        self.right_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ------------- Bottom Frame (Buttons) -------------
        bottom_button_frame = tk.Frame(self.root)
        bottom_button_frame.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=5)

        btn_pdf = tk.Button(bottom_button_frame, text="Abstract herunterladen & Webseite als PDF speichern", command=self.download_papers)
        btn_pdf.pack(side=tk.LEFT, padx=(0, 10))

        btn_open_page = tk.Button(bottom_button_frame, text="Seite öffnen (nur Browser)", command=self.open_article_page)
        btn_open_page.pack(side=tk.LEFT)

        btn_execute_mods = tk.Button(bottom_button_frame, text="Ausgewählte Module ausführen", command=self.execute_selected_modules)
        btn_execute_mods.pack(side=tk.LEFT, padx=10)

        btn_excel_all = tk.Button(bottom_button_frame, text="Komplette Excel-Liste erstellen", command=self.create_complete_excel_list)
        btn_excel_all.pack(side=tk.LEFT, padx=10)

        btn_analysis = tk.Button(bottom_button_frame, text="Analyse & Bewertung", command=self.open_analysis_window, bg="green", fg="white")
        btn_analysis.pack(side=tk.LEFT, padx=10)

        btn_extended = tk.Button(bottom_button_frame, text="Erweiterte Themen & Upload", command=self.open_extended_topic_window, bg="blue", fg="white")
        btn_extended.pack(side=tk.LEFT, padx=10)

        self.blink_label = tk.Label(self.root, text="", width=2, height=1)
        self.blink_label.place(relx=1.0, rely=1.0, x=-50, y=-40, anchor="center")
        self._blink_lamp()

    def open_analysis_window(self):
        selected_papers = []
        for item_id in self.right_tree.get_children():
            vals = self.right_tree.item(item_id, "values")
            paper = {
                "Source": vals[0],
                "Title": vals[1],
                "Authors/Description": vals[2],
                "Journal/Organism": vals[3],
                "Year": vals[4],
                "PMID": vals[5],
                "DOI": vals[6],
                "URL": vals[7],
            }
            selected_papers.append(paper)

        AnalysisWindow(self.root, selected_papers)

    def open_extended_topic_window(self):
        ExtendedTopicsWindow(self.root, openai_api_key=CHATGPT_API_KEY)

    def update_log(self, msg):
        self.log_text.insert(tk.END, f"{msg}\n")
        self.log_text.see(tk.END)

    def _open_excel_sheet_for_genes(self):
        if not self.filter_gene_excel_var.get():
            self.gene_excel_genes.clear()
            self.gene_excel_selected_sheet = None
            return
        file_path = filedialog.askopenfilename(title="Excel-Datei für Gene auswählen", filetypes=[("Excel-Dateien", "*.xlsx *.xls")])
        if not file_path:
            self.update_log("Gene (Excel): Keine Datei gewählt.")
            return
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet_names = wb.sheetnames
            sheet_win = tk.Toplevel(self.root)
            sheet_win.title("Sheet auswählen")

            tk.Label(sheet_win, text=f"Bitte wähle ein Sheet aus '{os.path.basename(file_path)}':").pack(padx=10, pady=5)
            sheet_var = tk.StringVar(value=sheet_names[0])
            combo_sheet = ttk.Combobox(sheet_win, textvariable=sheet_var, values=sheet_names, state="readonly")
            combo_sheet.pack(padx=10, pady=5)

            def confirm_sheet():
                self.gene_excel_selected_sheet = combo_sheet.get()
                sheet = wb[self.gene_excel_selected_sheet]
                self.gene_excel_genes.clear()
                row_idx = 3
                col_letter = "C"
                while True:
                    cell_id = f"{col_letter}{row_idx}"
                    val = sheet[cell_id].value
                    if val is None:
                        break
                    gene_str = str(val).strip()
                    if gene_str:
                        self.gene_excel_genes.append(gene_str)
                    row_idx += 1

                self.update_log(f"Gene (Excel): Sheet '{self.gene_excel_selected_sheet}' eingelesen, {len(self.gene_excel_genes)} Einträge.")
                sheet_win.destroy()

            btn_ok = tk.Button(sheet_win, text="OK", command=confirm_sheet)
            btn_ok.pack(padx=10, pady=5)
        except Exception as e:
            self.update_log(f"Fehler beim Öffnen der Excel-Datei: {e}")

    def _set_extra_filter(self):
        term = self.extra_filter_entry.get().strip()
        if term:
            self.extra_filter_string.set(term)
            self.update_log(f"Extra-Filterbegriff gesetzt: '{term}'")
        else:
            self.extra_filter_string.set("")
            self.update_log("Extra-Filterbegriff zurückgesetzt (leer)")

    def _blink_lamp(self):
        any_filter_active = (
            self.filter_local_var.get()
            or self.filter_perplexity_var.get()
            or self.filter_deepseek_var.get()
            or self.filter_chatgpt_var.get()
            or self.filter_extra_var.get()
            or self.filter_gene_excel_var.get()
        )
        color_on = "green" if any_filter_active else "red"
        color_off = "white"

        if self.blink_state:
            self.blink_label.config(bg=color_on)
        else:
            self.blink_label.config(bg=color_off)
        self.blink_state = not self.blink_state
        self.root.after(500, self._blink_lamp)

    def open_manual_uniprot(self):
        top = tk.Toplevel(self.root)
        top.title("Manuelle UniProt-Eingabe")
        tk.Label(top, text="Protein manuell eingeben:").pack(padx=10, pady=5)
        manual_entry = tk.Entry(top, width=30)
        manual_entry.pack(padx=10, pady=5)

        def set_value():
            value = manual_entry.get().strip()
            if value:
                self.uniprot_protein_var.set(value)
                self.uniprot_protein_combo.set(value)
            top.destroy()

        btn_ok = tk.Button(top, text="OK", command=set_value)
        btn_ok.pack(padx=10, pady=5)

    def add_code_word_entry(self):
        entry = tk.Entry(self.words_frame, width=30)
        entry.pack(side=tk.LEFT, padx=2, pady=2)
        self.code_word_entries.append(entry)

    def clear_search_results(self):
        self.all_results.clear()
        self.current_page = 0
        self.excel_details.clear()
        self.excel_main.clear()
        for ch in self.left_tree.get_children():
            self.left_tree.delete(ch)
        for ch in self.right_tree.get_children():
            self.right_tree.delete(ch)

        self.found_frame.config(bg="red")
        self.total_label.config(text="Insgesamt 0 Paper gefunden", bg="red", fg="white")
        self.log_text.delete("1.0", tk.END)

    def clear_codewords(self):
        self.codeword_entry.delete(0, tk.END)
        for entry in self.code_word_entries:
            entry.delete(0, tk.END)

    def _on_excel_file_change(self, event):
        chosen_key = self.bereich_var.get()
        if chosen_key in self.available_excels:
            self.current_excel_path = self.available_excels[chosen_key]
            self.journals, self.journal_lookup = self._load_journals_from_excel(self.current_excel_path)
            self.journal_combo["values"] = self.journals
            self.journal_combo.set("")

    def _on_search_excel_file_change(self, event):
        chosen_key = self.search_bereich_var.get()
        if chosen_key in self.available_excels:
            new_xlsx = self.available_excels[chosen_key]
            jlist, jlookup = self._load_journals_from_excel(new_xlsx)
            self.search_journal_combo["values"] = jlist
            self.search_journal_combo.set("")

    def _ensure_font_available(self):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        ttf_path = os.path.join(script_dir, self.font_file)
        if not os.path.exists(ttf_path):
            url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSansCondensed.ttf"
            try:
                resp = requests.get(url, timeout=15)
                resp.raise_for_status()
                with open(ttf_path, "wb") as f:
                    f.write(resp.content)
            except Exception as e:
                raise RuntimeError(f"Fehler beim Laden {self.font_file}: {e}")

    def _load_journals_from_excel(self, excel_path):
        jlist = []
        jlookup = {}
        if not os.path.exists(excel_path):
            return jlist, jlookup
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            data_temp = []
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
                if row and row[0] and row[1]:
                    data_temp.append((row[0], row[1]))
            data_temp.sort(key=lambda x: x[0])
            for (r, t) in data_temp:
                combined = f"{r}) {t}"
                jlist.append(combined)
                jlookup[combined] = t
        except Exception:
            pass
        return jlist, jlookup

    def get_query_string(self):
        code_words = []
        for entry in self.code_word_entries:
            w = entry.get().strip()
            if w:
                code_words.append(w)
        main_word = self.codeword_entry.get().strip()
        if main_word:
            code_words.append(main_word)
        selected_geno = self.genotype_var.get().strip()
        if selected_geno:
            code_words.append(selected_geno)
        if not code_words:
            return ""
        return " OR ".join(code_words)

    def search_articles(self, event=None):
        self.pubmed_id_search = False
        query = self.get_query_string()
        if not query:
            messagebox.showwarning("Hinweis", "Bitte mindestens ein Suchwort oder Genotyp eingeben!")
            return

        selected_apis = [api for api, var in self.api_vars.items() if var.get()]
        if not selected_apis:
            messagebox.showwarning("Hinweis", "Mindestens eine API auswählen!")
            return

        self.all_results.clear()
        self.current_page = 0
        self.excel_details.clear()
        self.excel_main.clear()
        for ch in self.left_tree.get_children():
            self.left_tree.delete(ch)
        for ch in self.right_tree.get_children():
            self.right_tree.delete(ch)

        if self.apply_search_criteria.get():
            self.search_pop = bool(self.search_populationsgroesse.get())
            y = self.search_year_var.get().strip()
            if y:
                self.search_year = f"{y}:{y}[dp]"
            else:
                self.search_year = ""
            self.search_journal = self.search_journal_var.get().strip()
        else:
            self.search_pop = False
            self.search_year = ""
            self.search_journal = ""

        self.update_log(f"Starte Suche für: {query}")

        for api in selected_apis:
            self.update_log(f"Rufe API '{api}' auf ...")
            if api == "Europe PMC":
                self.search_europe_pmc(query)
            elif api == "PubMed":
                self.search_pubmed(query)
            elif api == "OpenAlex":
                self.search_openalex(query)
            elif api == "Google Scholar":
                self.search_google_scholar(query)
            elif api == "Semantic Scholar":
                self.search_semantic_scholar(query)
            elif api == "CORE Aggregate":
                self.search_core_aggregate(query)

        use_filter = (
            self.filter_local_var.get()
            or self.filter_perplexity_var.get()
            or self.filter_deepseek_var.get()
            or self.filter_chatgpt_var.get()
            or self.filter_extra_var.get()
            or self.filter_gene_excel_var.get()
        )
        if use_filter:
            messagebox.showinfo(
                "Filter",
                "Die Online-Filterung wird durchgeführt. Dies kann einige Zeit dauern.\n"
                "Bitte auf 'OK' klicken, um fortzufahren..."
            )
            original_count = len(self.all_results)
            filtered_results = []

            for i, paper in enumerate(self.all_results, start=1):
                self.blue_label.config(text=f"({i}/{original_count}) -> {paper.get('Title','')[:30]}")
                self.root.update()

                info_line = f"Online-Filter: Prüfe Paper {i}/{original_count} -> '{paper.get('Title','')}'"
                self.update_log(info_line)

                keep_it = self.online_filter_paper(paper)
                if keep_it:
                    filtered_results.append(paper)
                else:
                    self.update_log(f"Ergebnis Filter für '{paper.get('Title','')}' -> False")

            self.blue_label.config(text="Filter abgeschlossen")

            if not filtered_results:
                messagebox.showinfo("Online Filter", "Keine Ergebnisse nach Online-Filterung gefunden. Es werden alle Paper angezeigt.")
            else:
                self.all_results = filtered_results

            if messagebox.askyesno("Codewörter löschen?", "Möchten Sie die eingegebenen Codewörter für die nächste Suche entfernen?"):
                self.clear_codewords()

            messagebox.showinfo("Neue Suche möglich", "Die Filterung ist abgeschlossen. Du kannst jetzt erneut auf 'Suchen' klicken oder die Codewörter anpassen.")

        self.update_log("Suche abgeschlossen.")
        self._display_page()

    def online_filter_paper(self, paper):
        url = paper.get("URL", "")
        if not url or url == "n/a":
            self.update_log(f"[Filter] Keine gültige URL -> PAPER gefiltert")
            return False

        combined_text = paper.get("Abstract", "").lower()
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            combined_text += " " + r.text.lower()
        except Exception as e:
            self.update_log(f"Local Filter: Fehler beim Abruf URL {url}: {e}")

        self.update_log(f"[Filter] Prüfe Paper: {paper.get('Title','(kein Titel)')} mit URL: {url}")

        result_local = False
        result_perplexity = False
        result_deepseek = False
        result_chatgpt = False
        result_extra = False
        result_gene_excel = False

        if self.filter_local_var.get():
            keywords = ["genotype", "phenotype", "snp", "genotyp", "phänotyp", "gene"]
            for kw in keywords:
                if kw in combined_text:
                    self.update_log(f"  -> Lokaler Filter: Schlüsselwort '{kw}' in (Abstract/Seite) gefunden -> PAPER OK")
                    result_local = True
                    break

        if self.filter_perplexity_var.get():
            result_perplexity = self.filter_with_perplexity(url, paper.get("Abstract", ""))
            self.update_log(f"  Perplexity Filter: {result_perplexity}")

        if self.filter_deepseek_var.get():
            result_deepseek = self.filter_with_deepseek(url, paper.get("Abstract", ""))
            self.update_log(f"  DeepSeek Filter: {result_deepseek}")

        if self.filter_chatgpt_var.get():
            criteria = []
            if self.chatgpt_genotype_var.get():
                criteria.append("genotype")
            if self.chatgpt_phenotype_var.get():
                criteria.append("phenotype")
            if self.chatgpt_snp_var.get():
                criteria.append("SNP")
            result_chatgpt = self.filter_with_chatgpt(url, criteria, paper.get("Abstract", ""))
            self.update_log(f"  ChatGPT Filter: {result_chatgpt}")

        if self.filter_extra_var.get():
            term = self.extra_filter_string.get().lower().strip()
            if term:
                if term in combined_text:
                    self.update_log(f"  -> Extra Filterbegriff '{term}' im kombinierten Text gefunden -> PAPER OK")
                    result_extra = True

        if self.filter_gene_excel_var.get() and self.gene_excel_genes:
            found_match = False
            title_lower = paper.get("Title", "").lower()
            for gene_candidate in self.gene_excel_genes:
                g_lower = gene_candidate.lower()
                if g_lower in title_lower or g_lower in combined_text:
                    found_match = True
                    break

            if found_match:
                self.update_log(f"  -> Gene-Excel-Filter: '{gene_candidate}' in Titel/Abstract/Seite gefunden -> PAPER OK")
                result_gene_excel = True
            else:
                self.update_log(f"  Gene-Excel-Filter: Keine Übereinstimmung")

        keep_paper = (result_local or result_perplexity or result_deepseek or result_chatgpt or result_extra or result_gene_excel)
        self.update_log(f"Ergebnis Filter für '{paper.get('Title','')}' -> {keep_paper}")
        return keep_paper

    def filter_with_perplexity(self, url, abstract_text):
        api_url = "https://api.perplexity.ai/chat/completions"
        context_snippet = abstract_text[:500]
        payload = {
            "model": "sonar",
            "frequency_penalty": 0.1,
            "max_tokens": 1000,
            "return_images": True,
            "return_related_questions": True,
            "search_recency_filter": "hour",
            "stream": False,
            "temperature": 0,
            "top_k": 0,
            "top_p": 1,
            "messages": [
                {
                    "role": "user",
                    "content": (
                        f"Bitte prüfe anhand des folgenden Abstract-Ausschnitts "
                        f"und ggf. der Seite {url} (wenn frei zugänglich), "
                        f"ob Informationen zu Genotype, Phenotype oder SNP enthalten sind. "
                        f"Abstract-Auszug:\n{context_snippet}\n\n"
                        f"Antworte mit JA oder NEIN."
                    )
                }
            ]
        }
        headers = {
            "Authorization": f"Bearer {PERPLEXITY_API_KEY}",
            "Content-Type": "application/json"
        }
        try:
            response = requests.post(api_url, json=payload, headers=headers, timeout=30)
            if response.status_code == 200:
                data = response.json()
                answer = data.get("choices", [{}])[0].get("message", {}).get("content", "").lower()
                return "ja" in answer
            else:
                self.update_log(f"Perplexity API Fehler: {response.status_code} - {response.text}")
                return False
        except requests.exceptions.RequestException as e:
            self.update_log(f"Perplexity API Fehler: {e}")
            return False

    def filter_with_deepseek(self, url, abstract_text):
        api_url = "https://api.deepseek.ai/chat/completions"
        context_snippet = abstract_text[:500]
        payload = {
            "model": "deep-model-v1",
            "max_tokens": 1000,
            "temperature": 0,
            "messages": [
                {
                    "role": "user",
                    "content": (
                        f"Überprüfe diese Infos (Abstract-Ausschnitt + Seite {url}) "
                        f"auf Genotype, Phenotype oder SNP:\n"
                        f"{context_snippet}\n\n"
                        f"Antworte mit JA oder NEIN."
                    )
                }
            ]
        }
        headers = {
            "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
            "Content-Type": "application/json"
        }
        try:
            response = requests.post(api_url, json=payload, headers=headers, timeout=30)
            if response.status_code == 200:
                data = response.json()
                answer = data.get("choices", [{}])[0].get("message", {}).get("content", "").lower()
                return "ja" in answer
            else:
                self.update_log(f"DeepSeek API Fehler: {response.status_code} - {response.text}")
                return False
        except requests.exceptions.RequestException as e:
            self.update_log(f"DeepSeek API Fehler: {e}")
            return False

    def filter_with_chatgpt(self, url, criteria_list, abstract_text):
        try:
            snippet = abstract_text[:1000]
            criteria_str = ", ".join(criteria_list)
            prompt = (
                f"Analysiere diesen Textauszug (Abstract) und ggf. die Seite {url}, "
                f"um festzustellen, ob es Informationen über {criteria_str} gibt. "
                f"Antworte mit 'Yes' oder 'No'.\n\nText: {snippet}"
            )
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}],
                temperature=0
            )
            answer = response.choices[0].message.content.strip().lower()
            return "yes" in answer
        except Exception as e:
            self.update_log(f"ChatGPT API Fehler: {e}")
            return False

    def search_europe_pmc(self, base_query):
        q = base_query
        if self.search_pop:
            q += " OR popul"
        if self.search_year:
            raw_year = self.search_year.split(':')[0]
            q += f" AND PUB_YEAR:{raw_year}"
        if self.search_journal:
            q += f' AND JOURNAL:"{self.search_journal}"'
        url = "https://www.ebi.ac.uk/europepmc/webservices/rest/search"
        params = {"query": q, "format": "json", "pageSize": 1000, "resultType": "core"}
        try:
            resp = requests.get(url, params=params, timeout=15)
            resp.raise_for_status()
            data = resp.json()
            if "resultList" not in data or "result" not in data["resultList"]:
                return
            for item in data["resultList"]["result"]:
                title = item.get("title", "n/a")
                authors = item.get("authorString", "n/a")
                journ = item.get("journalTitle", "n/a")
                year = item.get("pubYear", "n/a")
                pmid = item.get("pmid", "n/a")
                pmcid = item.get("pmcid", "")
                doi = item.get("doi", "")
                abstract_text = item.get("abstractText", "")
                if pmcid:
                    if not pmcid.startswith("PMC"):
                        pmcid = "PMC" + pmcid
                    url_article = f"https://europepmc.org/articles/{pmcid}"
                elif pmid:
                    url_article = f"https://europepmc.org/article/MED/{pmid}"
                elif doi:
                    url_article = f"https://doi.org/{doi}"
                else:
                    url_article = f"https://europepmc.org/search?query={title}"
                self.all_results.append({
                    "Source": "Europe PMC",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": journ,
                    "Year": year,
                    "PMID": pmid,
                    "DOI": doi,
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            messagebox.showerror("Fehler", f"Europe PMC: {e}")
            self.update_log(f"Europe PMC Error: {e}")

    def search_pubmed(self, base_query):
        q = f"({base_query})"
        if self.search_pop:
            q += " OR population"
        if self.search_year:
            q += f" AND {self.search_year}"
        if self.search_journal:
            q += f' AND "{self.search_journal}"[journal]'
        esearch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        params = {"db": "pubmed", "term": q, "retmode": "json", "retmax": 0}
        try:
            r = self.pubmed_session.get(esearch_url, params=params, timeout=10)
            r.raise_for_status()
            js = r.json()
            total_str = js.get("esearchresult", {}).get("count", "0")
            total = int(total_str)
            if total == 0:
                return
            max_to_fetch = min(total, 1000)
            step = 100
            fetched = 0
            all_ids = []
            while fetched < max_to_fetch:
                retmax = step
                retstart = fetched
                if (fetched + step) > max_to_fetch:
                    retmax = max_to_fetch - fetched
                block_params = {
                    "db": "pubmed",
                    "term": q,
                    "retmode": "json",
                    "retmax": retmax,
                    "retstart": retstart
                }
                block_resp = self.pubmed_session.get(esearch_url, params=block_params, timeout=10)
                block_resp.raise_for_status()
                block_data = block_resp.json()
                block_ids = block_data.get("esearchresult", {}).get("idlist", [])
                if not block_ids:
                    break
                all_ids.extend(block_ids)
                fetched += len(block_ids)
                if len(block_ids) < retmax:
                    break
            if not all_ids:
                return
            esummary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
            step_sum = 100
            idx = 0
            while idx < len(all_ids):
                subset = all_ids[idx:idx + step_sum]
                idx += step_sum
                id_str = ",".join(subset)
                sum_params = {"db": "pubmed", "id": id_str, "retmode": "json"}
                sum_resp = self.pubmed_session.get(esummary_url, params=sum_params, timeout=10)
                sum_resp.raise_for_status()
                sum_data = sum_resp.json()
                for uid in subset:
                    summ = sum_data.get("result", {}).get(uid, {})
                    if not summ:
                        continue
                    title = summ.get("title", "n/a")
                    authors_list = summ.get("authors", [])
                    authors = ", ".join(a.get("name", "") for a in authors_list)
                    journ = summ.get("fulljournalname", "n/a")
                    pub_date = summ.get("pubdate", "n/a")[:4]
                    pmid = summ.get("uid", "n/a")
                    doi = summ.get("elocationid", "n/a")
                    url_article = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
                    abstract_text = self._fetch_pubmed_abstract(pmid)
                    self.all_results.append({
                        "Source": "PubMed",
                        "Title": title,
                        "Authors/Description": authors,
                        "Journal/Organism": journ,
                        "Year": pub_date,
                        "PMID": pmid,
                        "DOI": doi,
                        "URL": url_article,
                        "Abstract": abstract_text
                    })
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Fehler", f"PubMed-Verbindungsfehler:\n{e}")
            self.update_log(f"PubMed Error: {e}")

    def search_pubmed_ids(self):
        self.pubmed_id_search = True
        ids_input = self.pubmed_id_entry.get().strip()
        if not ids_input:
            messagebox.showwarning("Hinweis", "Bitte mindestens eine PubMed-ID eingeben!")
            return
        pubmed_ids = [pmid.strip() for pmid in ids_input.split(',')]
        self.all_results.clear()
        for ch in self.left_tree.get_children():
            self.left_tree.delete(ch)
        for pmid in pubmed_ids:
            self.search_pubmed_id(pmid)
        self._display_page()

    def search_pubmed_id(self, pmid):
        esummary_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
        params = {"db": "pubmed", "id": pmid, "retmode": "json"}
        try:
            response = self.pubmed_session.get(esummary_url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            if "result" in data and pmid in data["result"]:
                summ = data["result"][pmid]
                title = summ.get("title", "n/a")
                authors_list = summ.get("authors", [])
                authors = ", ".join([a.get("name", "") for a in authors_list])
                journal = summ.get("fulljournalname", "n/a")
                pubdate = summ.get("pubdate", "n/a")
                year = pubdate[:4] if pubdate != "n/a" else "n/a"
                doi = summ.get("elocationid", "n/a")
                url_article = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
                abstract_text = self._fetch_pubmed_abstract(pmid)
                self.all_results.append({
                    "Source": "PubMed (ID-Suche)",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": journal,
                    "Year": year,
                    "PMID": pmid,
                    "DOI": doi,
                    "URL": url_article,
                    "Abstract": abstract_text
                })
            else:
                messagebox.showwarning("Warnung", f"Keine Daten für PubMed ID {pmid} gefunden.")
                self.update_log(f"PubMed ID {pmid}: keine Daten gefunden.")
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Fehler", f"Fehler beim Abrufen der PubMed ID {pmid}:\n{e}")
            self.update_log(f"PubMed ID Error: {e}")

    def search_openalex(self, base_query):
        url = f"https://api.openalex.org/works?filter=title.search:{base_query}&sort=cited_by_count:desc&per-page=25"
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            data = r.json()
            results = data.get("results", [])
            for work in results:
                title = work.get("title", "n/a")
                auths = work.get("authorships", [])
                authors = ", ".join(a["author"]["display_name"] for a in auths[:3])
                if len(auths) > 3:
                    authors += " et al."
                year = work.get("publication_year", "n/a")
                doi = work.get("doi", "n/a")
                url_article = "n/a"
                if doi != "n/a":
                    url_article = "https://doi.org/" + doi.replace("https://doi.org/", "")
                host_venue = work.get("host_venue", {})
                journ = host_venue.get("display_name", "n/a") if host_venue else "n/a"
                abstract_text = work.get("abstract", "")
                self.all_results.append({
                    "Source": "OpenAlex",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": journ,
                    "Year": str(year),
                    "PMID": "n/a",
                    "DOI": doi,
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            messagebox.showerror("Fehler", f"OpenAlex: {e}")
            self.update_log(f"OpenAlex Error: {e}")

    def search_google_scholar(self, base_query):
        try:
            from scholarly import scholarly
            search_results = scholarly.search_pubs(base_query)
            for _ in range(5):
                result = next(search_results)
                title = result['bib'].get('title', "n/a")
                authors = result['bib'].get('author', "n/a")
                year = result['bib'].get('pub_year', "n/a")
                url_article = result.get('url_scholarbib', "n/a")
                abstract_text = result['bib'].get('abstract', "")
                self.all_results.append({
                    "Source": "Google Scholar",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": "n/a",
                    "Year": year,
                    "PMID": "n/a",
                    "DOI": "n/a",
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            messagebox.showerror("Fehler", f"Google Scholar: {e}")
            self.update_log(f"Google Scholar Error: {e}")

    def search_semantic_scholar(self, base_query):
        try:
            url = "https://api.semanticscholar.org/graph/v1/paper/search"
            headers = {"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
            params = {"query": base_query, "limit": 5, "fields": "title,authors,year,abstract,doi,paperId"}
            response = requests.get(url, headers=headers, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            for paper in data.get("data", []):
                title = paper.get("title", "n/a")
                authors = ", ".join([author.get("name", "") for author in paper.get("authors", [])])
                year = paper.get("year", "n/a")
                doi = paper.get("doi", "n/a")
                paper_id = paper.get("paperId", "")
                abstract_text = paper.get("abstract", "")
                url_article = f"https://www.semanticscholar.org/paper/{paper_id}" if paper_id else "n/a"
                self.all_results.append({
                    "Source": "Semantic Scholar",
                    "Title": title,
                    "Authors/Description": authors,
                    "Journal/Organism": "n/a",
                    "Year": year,
                    "PMID": "n/a",
                    "DOI": doi,
                    "URL": url_article,
                    "Abstract": abstract_text
                })
        except Exception as e:
            messagebox.showerror("Fehler", f"Semantic Scholar: {e}")
            self.update_log(f"Semantic Scholar Error: {e}")

    def search_core_aggregate(self, base_query):
        API_KEY = "LmAMxdYnK6SDJsPRQCpGgwN7f5yTUBHF"
        core = CoreAPI(API_KEY)
        filters = {"yearPublished": ">=2020,<=2024", "language.name": "English"}
        sort = "citationCount:desc"
        limit = 5
        try:
            result = core.search_publications(query=base_query, filters=filters, sort=sort, limit=limit)
            if result and "results" in result:
                for pub in result["results"]:
                    title = pub.get("title", "n/a")
                    doi = pub.get("doi", "n/a")
                    year = pub.get("yearPublished", "n/a")
                    authors = ", ".join(pub.get("authors", []))
                    downloadUrl = pub.get("downloadUrl", "n/a")
                    abstract = pub.get("abstract", "")
                    publisher = pub.get("publisher", "n/a")
                    self.all_results.append({
                        "Source": "CORE Aggregate",
                        "Title": title,
                        "Authors/Description": authors,
                        "Journal/Organism": publisher,
                        "Year": year,
                        "PMID": "n/a",
                        "DOI": doi,
                        "URL": downloadUrl,
                        "Abstract": abstract
                    })
        except Exception as e:
            messagebox.showerror("Fehler", f"CORE Aggregate: {e}")
            self.update_log(f"CORE Aggregate Error: {e}")

    def search_ensembl_direct(self):
        species_val = self.ensembl_species_var.get().strip()
        gene_val = self.ensembl_gene_var.get().strip()
        if not species_val or not gene_val:
            messagebox.showwarning("Fehler", "Bitte Species und Gene eingeben!")
            return
        self.all_results.clear()
        self.current_page = 0
        for ch in self.left_tree.get_children():
            self.left_tree.delete(ch)
        self._search_ensembl_api(species_val, gene_val)
        self._display_page()

    def _search_ensembl_api(self, species, gene):
        base_url = "https://rest.ensembl.org"
        endpoint = f"/lookup/symbol/{species}/{gene}"
        headers = {"Content-Type": "application/json"}
        try:
            resp = requests.get(base_url + endpoint, headers=headers, timeout=10)
            resp.raise_for_status()
            data = resp.json()
            desc_str = f"chr: {data.get('seq_region_name', 'n/a')}, strand: {data.get('strand', 'n/a')}"
            ensembl_url = f"https://www.ensembl.org/{species}/Gene/Summary?g={data.get('id','')}"
            self.all_results.append({
                "Source": "Ensembl REST",
                "Title": data.get("display_name", gene),
                "Authors/Description": desc_str,
                "Journal/Organism": data.get("species", "n/a"),
                "Year": "n/a",
                "PMID": "n/a",
                "DOI": "n/a",
                "URL": ensembl_url,
                "Abstract": ""
            })
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Fehler", f"Ensembl REST: {e}")
            self.update_log(f"Ensembl Error: {e}")

    def search_uniprot_direct(self):
        protein_name = self.uniprot_protein_var.get().strip()
        if not protein_name:
            messagebox.showwarning("Hinweis", "Bitte Protein auswählen oder manuell eingeben!")
            return
        self.all_results.clear()
        self.current_page = 0
        for ch in self.left_tree.get_children():
            self.left_tree.delete(ch)
        self._search_uniprot(protein_name)
        self._display_page()

    def _search_uniprot(self, protein):
        try:
            url_search = f"https://rest.uniprot.org/uniprotkb/search?query={protein}&format=json&size=1"
            resp1 = requests.get(url_search, timeout=10)
            resp1.raise_for_status()
            data_search = resp1.json()
            results = data_search.get("results", [])
            if not results:
                return
            accession = results[0].get("primaryAccession", "")
            if not accession:
                return
            url_detail = f"https://rest.uniprot.org/uniprotkb/{accession}.json"
            resp2 = requests.get(url_detail, timeout=10)
            resp2.raise_for_status()
            data_detail = resp2.json()
            protein_name = data_detail.get("proteinDescription", {}).get("recommendedName", {}).get("fullName", {}).get("value", "n/a")
            organism = data_detail.get("organism", {}).get("scientificName", "n/a")
            seq_len = data_detail.get("sequence", {}).get("length", "n/a")
            url_article = f"https://www.uniprot.org/uniprotkb/{accession}"
            authors_description = f"Organismus: {organism}; Länge: {seq_len}"
            self.all_results.append({
                "Source": "UniProt",
                "Title": protein_name,
                "Authors/Description": authors_description,
                "Journal/Organism": organism,
                "Year": "n/a",
                "PMID": "n/a",
                "DOI": "n/a",
                "URL": url_article,
                "Abstract": ""
            })
        except Exception as e:
            messagebox.showerror("Fehler", f"UniProt: {e}")
            self.update_log(f"UniProt Error: {e}")

    def search_dbsnp_direct(self):
        rs_id = self.dbsnp_rs_entry.get().strip()
        if not rs_id:
            messagebox.showwarning("Fehler", "Bitte eine RS-ID eingeben!")
            return
        YOUR_EMAIL = "erwin.schimak@novogenia.com"
        API_KEY = "7c60b310bfd435eed965f2b203f4a1900208"
        dbsnp = DBSnpAPI(email=YOUR_EMAIL, api_key=API_KEY)
        result = dbsnp.get_snp_info(rs_id)
        if result:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            pdf.cell(0, 10, "Detaillierte SNP-Informationen", ln=True, align="C")
            pdf.ln(10)
            for key, value in result.items():
                if isinstance(value, list):
                    value_str = ", ".join(value)
                else:
                    value_str = str(value)
                pdf.cell(0, 10, f"{key}: {value_str}", ln=True)
            pdf_filename = f"dbSNP_{rs_id}.pdf"
            pdf.output(pdf_filename)
            self.update_log(f"dbSNP PDF erstellt: {pdf_filename}")
            try:
                webbrowser.open(pdf_filename)
            except Exception as e:
                self.update_log(f"Öffnen PDF Fehler: {e}")
        else:
            messagebox.showinfo("Ergebnis", "Keine Informationen gefunden.")
            self.update_log("dbSNP: Keine Informationen gefunden.")

    def show_abstract_popup(self, event):
        sel = self.left_tree.selection()
        if not sel:
            return
        item_id = self.left_tree.item(sel[0], "values")
        source, title, authors, _, _, pmid, _, _ = item_id
        paper = None
        for p in self.all_results:
            if p["Title"] == title and p.get("PMID", None) == pmid:
                paper = p
                break
        if paper is None:
            abstract_text = "(Kein Abstract verfügbar)"
        else:
            abstract_text = paper.get("Abstract", "(Kein Abstract verfügbar)")
        popup = tk.Toplevel(self.root)
        popup.title("Paper-Details")
        tk.Label(popup, text=f"Titel: {title}", font=("Arial", 11, "bold")).pack(anchor="w", padx=10, pady=5)
        tk.Label(popup, text=f"Autoren: {authors}", font=("Arial", 10)).pack(anchor="w", padx=10, pady=5)
        st = scrolledtext.ScrolledText(popup, wrap="word", width=80, height=15)
        st.pack(fill="both", expand=True, padx=10, pady=5)
        st.insert("1.0", abstract_text)
        st.config(state="disabled")

    def _fetch_pubmed_abstract(self, pmid):
        url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
        params = {"db": "pubmed", "id": pmid, "retmode": "xml"}
        try:
            r = self.pubmed_session.get(url, params=params, timeout=10)
            r.raise_for_status()
            root = ET.fromstring(r.content)
            abs_text = ""
            for elem in root.findall(".//AbstractText"):
                if elem.text:
                    abs_text += (elem.text + "\n")
            if not abs_text.strip():
                abs_text = "(Kein Abstract via eFetch)"
            return abs_text.strip()
        except Exception as e:
            return f"(Fehler eFetch: {e})"

    def _get_esummary_data(self, pmid):
        url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esummary.fcgi"
        params = {"db": "pubmed", "id": pmid, "retmode": "json"}
        out = {}
        try:
            r = self.pubmed_session.get(url, params=params, timeout=10)
            r.raise_for_status()
            js = r.json()
            results = js.get("result", {})
            if pmid in results:
                summ = results[pmid]
                out["Title"] = summ.get("title", "n/a")
                out["FullJournalName"] = summ.get("fulljournalname", "n/a")
                out["PubDate"] = summ.get("pubdate", "n/a")
                auth_list = summ.get("authors", [])
                auth_str = ", ".join(a.get("name", "") for a in auth_list)
                out["Authors"] = auth_str
                out["DOI"] = summ.get("elocationid", "n/a")
            else:
                out["Error"] = "Keine eSummary Daten"
        except Exception as e:
            out["Error"] = f"Fehler eSummary: {e}"
        return out

    def auto_select_papers(self):
        pop_checked = self.var_populationsgroesse.get()
        chosen_year = self.year_var.get()
        chosen_journal_label = self.journal_var.get()
        chosen_journal_title = ""
        if chosen_journal_label and chosen_journal_label in self.journal_lookup:
            chosen_journal_title = self.journal_lookup[chosen_journal_label]
        for item_id in self.left_tree.get_children():
            vals = self.left_tree.item(item_id, "values")
            if self._is_item_in_tree(self.right_tree, vals):
                continue
            if pop_checked:
                text_concat = (vals[1] + " " + vals[2]).lower()
                if "popul" not in text_concat:
                    continue
            if chosen_year:
                if vals[4] != chosen_year:
                    continue
            if chosen_journal_title:
                if vals[3].lower() != chosen_journal_title.lower():
                    continue
            self.right_tree.insert("", tk.END, values=vals)

    def add_selected_papers(self):
        sel = self.left_tree.selection()
        for s in sel:
            vals = self.left_tree.item(s, "values")
            if not self._is_item_in_tree(self.right_tree, vals):
                self.right_tree.insert("", tk.END, values=vals)

    def remove_selected_papers(self):
        sel = self.right_tree.selection()
        for s in sel:
            self.right_tree.delete(s)

    def _is_item_in_tree(self, tree, values):
        for child in tree.get_children():
            if tree.item(child, "values") == values:
                return True
        return False

    def next_page(self):
        max_page = (len(self.all_results) - 1) // self.page_size
        if self.current_page < max_page:
            self.current_page += 1
            self._display_page()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self._display_page()

    def _display_page(self):
        for row_id in self.left_tree.get_children():
            self.left_tree.delete(row_id)
        start_idx = self.current_page * self.page_size
        end_idx = start_idx + self.page_size
        page_items = self.all_results[start_idx:end_idx]
        for res in page_items:
            vals = (
                res.get("Source", ""),
                res.get("Title", ""),
                res.get("Authors/Description", ""),
                res.get("Journal/Organism", ""),
                res.get("Year", ""),
                res.get("PMID", ""),
                res.get("DOI", ""),
                res.get("URL", "")
            )
            self.left_tree.insert("", tk.END, values=vals)
        total_count = len(self.all_results)
        if total_count > 0:
            self.found_frame.config(bg="green")
            self.total_label.config(text=f"Insgesamt {total_count} Paper gefunden (Seite {self.current_page+1})", bg="green", fg="white")
        else:
            self.found_frame.config(bg="red")
            self.total_label.config(text="Insgesamt 0 Paper gefunden", bg="red", fg="white")

    def download_papers(self):
        selected_items = self.right_tree.get_children()
        if not selected_items:
            messagebox.showwarning("Hinweis", "Bitte mindestens ein Paper auswählen.")
            return
        storage_dir = filedialog.askdirectory(title="Speicherort für PDFs und Excel-Dateien")
        if not storage_dir:
            messagebox.showwarning("Hinweis", "Kein Speicherort ausgewählt.")
            return
        curr = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        if self.pubmed_id_search:
            folder_name = f"PubMedID_Search_{curr}"
        else:
            query = self.get_query_string()
            folder_name = f"{query}_{curr}" if query else f"Ergebnisse_{curr}"
        output_dir = os.path.join(storage_dir, folder_name)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.update_log(f"Ordner erstellt: {output_dir}")
        else:
            self.update_log(f"Ordner existiert bereits: {output_dir}")
        for sel in selected_items:
            vals = self.right_tree.item(sel, "values")
            result = {
                "Source": vals[0],
                "Title": vals[1],
                "Authors/Description": vals[2],
                "Journal/Organism": vals[3],
                "Year": vals[4],
                "PMID": vals[5],
                "DOI": vals[6],
                "URL": vals[7]
            }
            title = result.get("Title", "article")
            sanitized_title = self._sanitize_filename(title)
            abstract_pdf_filename = f"{sanitized_title}_abstract.pdf"
            webpage_pdf_filename = f"{sanitized_title}_webpage_chrome.pdf"
            abstract_pdf_path = os.path.join(output_dir, abstract_pdf_filename)
            webpage_pdf_path = os.path.join(output_dir, webpage_pdf_filename)
            article_url = result.get("URL", "")
            pmid = result.get("PMID", "n/a")
            doi = result.get("DOI", "n/a")
            self._create_abstract_pdf(result, article_url, abstract_pdf_path)
            self.update_log(f"PDF erstellt: {abstract_pdf_path}")
            if article_url and article_url != "n/a":
                try:
                    time.sleep(5)
                    self._save_page_via_headless_chrome(article_url, webpage_pdf_path)
                    self.update_log(f"Webseite als PDF gespeichert: {webpage_pdf_path}")
                except Exception as e:
                    self.update_log(f"Fehler Chrome Headless: {e}")
                    messagebox.showerror("Fehler", f"Fehler beim Drucken der Webseite via Chrome Headless:\n{e}")
                    continue
            else:
                self.update_log("Keine gültige URL -> skip.")
                continue
            doi_links = self._extract_doi_links_from_pdf(webpage_pdf_path)
            if not doi_links and doi and doi != "n/a":
                doi_link = doi if doi.startswith("http") else f"https://doi.org/{doi}"
                if doi_link.startswith("https://doi.org/doi:"):
                    doi_link = doi_link.replace("https://doi.org/doi:", "https://doi.org/")
                doi_links.append(doi_link)
            if doi_links:
                for dl in doi_links:
                    sanitized_doi_filename = self._sanitize_filename(dl) + ".pdf"
                    doi_pdf_path = os.path.join(output_dir, sanitized_doi_filename)
                    try:
                        self._save_page_via_headless_chrome(dl, doi_pdf_path)
                        self.update_log(f"DOI-Link als PDF gespeichert: {doi_pdf_path}")
                        pdf_download_link = self._find_pdf_download_link(dl)
                        if pdf_download_link:
                            downloaded_pdf_filename = f"{sanitized_doi_filename[:-4]}_downloaded.pdf"
                            downloaded_pdf_path = os.path.join(output_dir, downloaded_pdf_filename)
                            self._download_pdf(pdf_download_link, downloaded_pdf_path)
                            self.update_log(f"Download PDF: {downloaded_pdf_path}")
                        else:
                            self.update_log(f"Kein direkter PDF-Download-Link: {dl}")
                    except Exception as e:
                        self.update_log(f"Fehler Druck DOI-Link: {e}")
                        messagebox.showerror("Fehler", f"Fehler beim Drucken des DOI-Links:\n{e}")
            else:
                self.update_log("Keine DOI-Links gefunden.")
            try:
                webbrowser.open(article_url)
            except Exception as e:
                self.update_log(f"Fehler Browser öffnen URL: {e}")
            if doi_links:
                for dl in doi_links:
                    try:
                        webbrowser.open(dl)
                    except Exception as e:
                        self.update_log(f"Fehler Browser öffnen DOI: {e}")
            details_entry = {
                "Download Name": abstract_pdf_filename,
                "Abstract": result.get("Authors/Description", ""),
                "PMID": pmid,
                "DOI": doi,
                "eSummary": "n/a"
            }
            self.excel_details.append(details_entry)
            main_entry = {
                "Title": title,
                "PMID": pmid,
                "DOI": doi
            }
            self.excel_main.append(main_entry)
            time.sleep(2)
        self._create_excel_files(output_dir)
        messagebox.showinfo("Fertig", "Alle ausgewählten Paper wurden verarbeitet.")
        self.update_log("Download + PDF-Erstellung beendet.")

    def _create_excel_files(self, output_dir):
        details_path = os.path.join(output_dir, "details.xlsx")
        wb_details = openpyxl.Workbook()
        ws_details = wb_details.active
        ws_details.title = "Details"
        headers_details = ["Download Name", "Abstract", "PMID", "DOI", "eSummary"]
        ws_details.append(headers_details)
        for entry in self.excel_details:
            row = [
                entry.get("Download Name", "n/a"),
                entry.get("Abstract", "n/a"),
                entry.get("PMID", "n/a"),
                entry.get("DOI", "n/a"),
                entry.get("eSummary", "n/a")
            ]
            ws_details.append(row)
        try:
            wb_details.save(details_path)
            self.update_log(f"Excel details.xlsx erstellt: {details_path}")
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern details.xlsx: {e}")
            self.update_log(f"Fehler details.xlsx: {e}")
        main_path = os.path.join(output_dir, "AllPapers.xlsx")
        wb = openpyxl.Workbook()
        ws_main = wb.active
        ws_main.title = "Main"
        headers_main = ["Title", "Year", "PMID", "DOI", "Abstract", "Link", "Populationsgröße?", "Journal"]
        ws_main.append(headers_main)
        selected_items = self.right_tree.get_children()
        for sel in selected_items:
            vals = self.right_tree.item(sel, "values")
            title = vals[1]
            year = vals[4]
            pmid = vals[5]
            doi = vals[6]
            abstract = vals[2]
            link = vals[7]
            pop_found = "Ja" if ("popul" in (vals[1] + " " + vals[2]).lower()) else "Nein"
            journal = vals[3]
            ws_main.append([title, year, pmid, doi, abstract, link, pop_found, journal])
            sheet_name = self._sanitize_filename(title)[:31]
            if sheet_name in wb.sheetnames:
                idx = 2
                orig = sheet_name
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{orig}_{idx}"
                    idx += 1
            ws_detail = wb.create_sheet(title=sheet_name)
            ws_detail.append(["Titel", title])
            ws_detail.append(["Year", year])
            ws_detail.append(["PMID", pmid])
            ws_detail.append(["DOI", doi])
            ws_detail.append(["Journal", journal])
            ws_detail.append(["Populationsgröße?", pop_found])
            if pmid and pmid != "n/a":
                e_sum = self._get_esummary_data(pmid)
                e_abs = self._fetch_pubmed_abstract(pmid)
                ws_detail.append([])
                ws_detail.append(["** eSummary **", ""])
                for k, v in e_sum.items():
                    ws_detail.append([k, v])
                ws_detail.append([])
                ws_detail.append(["** eFetch Abstract **", e_abs])
            else:
                ws_detail.append([])
                ws_detail.append(["(Keine PMID, daher keine eSummary/eFetch-Daten verfügbar)"])
        try:
            wb.save(main_path)
            self.update_log(f"Excel AllPapers.xlsx erstellt: {main_path}")
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern AllPapers.xlsx: {e}")
            self.update_log(f"Fehler AllPapers.xlsx: {e}")

    def open_article_page(self):
        sel = self.right_tree.get_children()
        if not sel:
            messagebox.showwarning("Hinweis", "Bitte mindestens eine Zeile auswählen.")
            return
        for s in sel:
            vals = self.right_tree.item(s, "values")
            url = vals[7]
            if "search?query=" in url:
                messagebox.showinfo("Info", "Keine pmcid/pmid/doi -> Europe PMC-Suche.")
            elif url == "n/a" or not url:
                messagebox.showinfo("Info", "Keine URL.")
                continue
            try:
                webbrowser.open(url)
            except Exception as e:
                self.update_log(f"Fehler open URL: {e}")

    def execute_selected_modules(self):
        messagebox.showinfo("Info", "Ausgewählte Module ausgeführt.")
        self.update_log("Module ausgeführt.")

    def _save_page_via_headless_chrome(self, url, output_pdf):
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36")
        prefs = {
            "download.default_directory": os.path.dirname(output_pdf),
            "plugins.always_open_pdf_externally": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        with webdriver.Chrome(options=chrome_options) as driver:
            driver.get(url)
            try:
                WebDriverWait(driver, 20).until(lambda d: d.execute_script("return document.readyState") == "complete")
            except TimeoutException:
                self.update_log(f"Timeout Laden: {url}")
            time.sleep(5)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            try:
                pdf_data = driver.execute_cdp_cmd("Page.printToPDF", {"printBackground": True, "paperWidth": 8.27, "paperHeight": 11.69})
                pdf_bytes = base64.b64decode(pdf_data["data"])
                with open(output_pdf, "wb") as f:
                    f.write(pdf_bytes)
            except Exception as e:
                self.update_log(f"Chrome printToPDF Fehler: {e}")

    def _create_abstract_pdf(self, result, article_url, output_file):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        try:
            pdf.add_font(self.font_name, "", self.font_file, uni=True)
            pdf.set_font(self.font_name, "", 12)
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler Schriftart: {e}")
            return
        page_width = pdf.w - pdf.l_margin - pdf.r_margin
        title_txt = f"** Titel **\n{result.get('Title', 'Keine Angabe')}"
        pdf.multi_cell(page_width, 10, self._wrap_long_words(title_txt))
        pdf.ln(5)
        desc = result.get("Authors/Description", "") or "Keine Beschreibung verfügbar."
        desc_txt = f"** Abstract/Description **\n{desc}"
        pdf.multi_cell(page_width, 10, self._wrap_long_words(desc_txt))
        pdf.ln(5)
        url_txt = "** URL **\n" + article_url
        pdf.multi_cell(page_width, 8, self._wrap_long_words(url_txt))
        pdf.ln(5)
        pdf.cell(0, 10, "** Metadaten **", ln=1)
        for key, val in result.items():
            if key in ["Source", "Title", "Authors/Description", "Journal/Organism", "Year", "URL"]:
                continue
            line = f"{key}: {str(val)}"
            pdf.multi_cell(page_width, 6, self._wrap_long_words(line))
        try:
            pdf.output(output_file)
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern PDF: {e}")

    def _wrap_long_words(self, text, max_len=80):
        words = text.split()
        result = []
        while words:
            w = words.pop(0)
            while len(w) > max_len:
                result.append(w[:max_len])
                w = w[max_len:]
            result.append(w)
        return " ".join(result)

    def _sanitize_filename(self, filename):
        invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
        for c in invalid_chars:
            filename = filename.replace(c, '_')
        return filename

    def _extract_doi_links_from_pdf(self, pdf_path):
        doi_links = []
        try:
            doc = fitz.open(pdf_path)
            for page in doc:
                text = page.get_text()
                matches = re.findall(r'(https?://(?:dx\.)?doi\.org/\S+)', text)
                for m in matches:
                    if m not in doi_links:
                        doi_links.append(m.strip())
            doc.close()
        except Exception as e:
            self.update_log(f"Fehler PyMuPDF: {e}")
        return doi_links

    def _find_pdf_download_link(self, url):
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            match = re.search(r'href=["\']([^"\']+\.pdf)["\']', r.text, re.IGNORECASE)
            if match:
                link = match.group(1)
                if link.startswith("/"):
                    from urllib.parse import urljoin
                    link = urljoin(url, link)
                return link
            return None
        except Exception as e:
            self.update_log(f"Fehler PDF-Link: {e}")
            return None

    def _download_pdf(self, pdf_url, output_path):
        try:
            r = requests.get(pdf_url, timeout=10)
            r.raise_for_status()
            with open(output_path, "wb") as f:
                f.write(r.content)
        except Exception as e:
            self.update_log(f"Fehler PDF-Download: {e}")

    def create_complete_excel_list(self):
        selected_items = self.right_tree.get_children()
        if not selected_items:
            messagebox.showwarning("Hinweis", "Bitte mindestens ein Paper auswählen.")
            return
        storage_dir = filedialog.askdirectory(title="Speicherort für Excel-Dateien")
        if not storage_dir:
            messagebox.showwarning("Hinweis", "Kein Speicherort gewählt.")
            return
        curr = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        if self.pubmed_id_search:
            folder_name = f"PubMedID_Search_{curr}"
        else:
            query = self.get_query_string()
            folder_name = f"{query}_{curr}" if query else f"Ergebnisse_{curr}"
        output_dir = os.path.join(storage_dir, folder_name)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            self.update_log(f"Ordner erstellt: {output_dir}")
        else:
            self.update_log(f"Ordner existiert bereits: {output_dir}")
        self._create_excel_files(output_dir)
        messagebox.showinfo("Fertig", "Die komplette Excel-Liste wurde erstellt.")
        self.update_log("Komplette Excel-Liste erstellt.")


def main():
    root = tk.Tk()
    app = App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
